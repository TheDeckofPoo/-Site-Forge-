#!/usr/bin/env python3
"""
fortna_autogen.py — Python port of the Excel VBA PLC Autogen (autogen_VBS_test.xlsm).

Reads conveyor/IO tables (from Excel Inputdata/IO sheets or JSON) + O'Reilly L5X library,
then generates a Studio 5000 .L5X project by cloning library template tags/rungs and
renaming them to site conveyor names (same approach as the Excel tool).

Why Excel felt "locked":
  - VBA macros + ActiveX buttons require Trust Center macro enable
  - Dict / IO Dict / Template sheets are formula engines (not for manual edit)
  - Only Inputdata + IO are meant as inputs
  - OneDrive + huge formula sheets make Excel sluggish

Usage:
  py tools/scripts/fortna_autogen.py from-excel path/to/autogen.xlsm --library tools/libraries/OReilly_Library_v3.L5X
  py tools/scripts/fortna_autogen.py from-json path/to/input.json --library tools/libraries/OReilly_Library_v3.L5X
  py tools/scripts/fortna_autogen.py inspect-excel path/to/autogen.xlsm
"""
from __future__ import annotations

import argparse
import json
import re
import shutil
import sys
from copy import deepcopy
from dataclasses import dataclass, field, asdict
from datetime import datetime, timezone
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parent.parent
sys.path.insert(0, str(SCRIPT_DIR))
DEFAULT_LIBRARY = REPO_ROOT / "tools" / "libraries" / "OReilly_Library_v3.L5X"
DEFAULT_SAMPLE_XLS = REPO_ROOT / "tools" / "libraries" / "autogen_VBS_test.xlsm"
DEFAULT_RUN = REPO_ROOT / "workspace" / "active" / "RUN"
PROGRAM_LIBRARY_DIR = REPO_ROOT / "tools" / "libraries" / "programs"
# Side-loaded AOI exports (e.g. Slow_Flt_AOI.L5X from source-key re-export).
# Replaces same-named AOI defs in the main library during build_l5x.
AOI_OVERLAY_DIR = REPO_ROOT / "tools" / "libraries"

# Gold program exports from Excel/Studio (Desktop/Autogen) — always or optional
ALWAYS_PROGRAMS: dict[str, str] = {
    "Sys": "Sys_Program.L5X",
    "IO_MAP": "IO_MAP_Program.L5X",  # full gold map; replaces RUN scaffold when present
}
OPTIONAL_PROGRAMS: dict[str, str] = {
    "ShippingSorter_Area_L3": "ShippingSorter_Area_L3_Program.L5X",
    "WCS_Interface_TCP_IP": "WCS_Interface_TCP_IP_Program.L5X",
    "Sorter_Track": "Sorter_Track_Program.L5X",
    # PLC4-class collector/sawtooth merge (optional template; not auto from tar)
    "Sawtooth_Merge": "Sawtooth_Merge_Program.L5X",
}

# Fortna ASC mechanical types → Excel autogen TYPE strings
FORTNA_TYPE_TO_AUTOGEN = {
    "STRAIGHT": "Transport with MS",
    "BELT": "Transport with MS",
    "CURVE": "Transport with MS",
    "MERGE": "Transport with MS",
    "SKEW": "Transport with MS",
    "SPUR": "Transport with MS",
    "TRIANG": "Transport with MS",
    "ACCUM": "Accumulation with MS",
    "ZEROPRESSURE": "Accumulation with MS",
}
FORTNA_TYPE_TO_AUTOGEN_VFD = {
    "STRAIGHT": "Transport with VFD",
    "BELT": "Transport with VFD",
    "CURVE": "Transport with VFD",
    "MERGE": "Transport with VFD",
    "SKEW": "Transport with VFD",
    "SPUR": "Transport with VFD",
    "TRIANG": "Transport with VFD",
    "ACCUM": "Accumulation with VFD",
    "ZEROPRESSURE": "Accumulation with VFD",
}
CONVEYOR_ASC_TYPES = frozenset(FORTNA_TYPE_TO_AUTOGEN.keys())

# Excel Config maps TYPE -> template base name. Library uses P1000–P4000 (not P7000/P8000).
# We prefer library-real templates; Config names remapped when missing.
TYPE_TO_TEMPLATE = {
    "transport with vfd": "P1000_Conv",
    "accumulation with vfd": "P2000_Conv",
    "transport with ms": "P3000_Conv",  # library: MS = P3000 (Excel said P7000)
    "accumulation with ms": "P4000_Conv",  # library: MS accum = P4000 (Excel said P8000)
    "transport with mdr": "P4000_Conv",
    "accumulation with mdr": "P3000_Conv",
    "gravity": "P3000_Conv",
}

# Excel Config TYPE string -> preferred template (from Config sheet)
EXCEL_CONFIG_OVERRIDES = {
    "transport with vfd": "P1000_Conv",
    "accumulation with vfd": "P2000_Conv",
    "transport with ms": "P7000_Conv",  # may not exist in lib — fallback applied
    "accumulation with ms": "P8000_Conv",
    "transport with mdr": "P4000_Conv",
    "accumulation with mdr": "P3000_Conv",
    "gravity": "P5000_Conv",
}


@dataclass
class ConveyorRow:
    number: int
    system: str = ""
    main_area: str = ""
    safety_zone: str = ""
    conveyor: str = ""
    type: str = ""
    downstream: str = ""
    track: str = ""
    exit_pe: str = ""
    entry_pe: str = ""
    induct: str = ""
    jam: str = ""
    height: str = ""
    full: str = ""
    add_pe: str = ""
    motor_starter: str = ""
    espb: str = ""
    espc: str = ""
    control_station: str = ""
    power_supply: str = ""
    # Real PE tag names from RUN (Excel-parity wiring)
    exit_pe_tag: str = ""          # product/discharge PE for Fast_Conv
    add_pe_tag: str = ""           # second PE for Fast_Conv add slot
    jam_pe_tags: list = field(default_factory=list)   # up to 5 for Slow_Jam
    full_pe_tags: list = field(default_factory=list)  # Full_PE AOIs
    product_pe_tags: list = field(default_factory=list)  # PE_Logic product eyes
    all_pe_tags: list = field(default_factory=list)   # every PE on this conv

    @property
    def clean_name(self) -> str:
        return re.sub(r"\s+", "", (self.conveyor or "").strip())

    @property
    def template_key(self) -> str:
        return (self.type or "").strip().lower()


@dataclass
class IoModule:
    name: str
    type: str
    slot: str = ""
    ip: str = ""
    parent: str = ""
    rack: str = ""
    connection: str = ""


@dataclass
class IoPoint:
    device_name: str
    device_type: str = ""
    source_module: str = ""
    port: str = ""
    direction: str = ""  # I / O
    slot_or_port: str = ""
    fortna_bank: str = ""
    fortna_bit: str = ""
    conveyor: str = ""
    description: str = ""
    pe_role: str = ""  # product | full | jam | other


@dataclass
class AutogenInput:
    project_name: str = "Autogen_Project"
    processor: str = "1756-L83E"
    major_rev: str = "35"
    minor_rev: str = "00"
    areas: list[str] = field(default_factory=list)
    safety_zones: list[str] = field(default_factory=list)
    conveyors: list[ConveyorRow] = field(default_factory=list)
    modules: list[IoModule] = field(default_factory=list)
    io_points: list[IoPoint] = field(default_factory=list)
    # EIP adapters from eipcfg (for Modules tree)
    eip_adapters: list = field(default_factory=list)
    eip_interface_ip: str = ""
    # Named topology: [{rio_name, rack, ip, children:[{flex_slot, type, word, ...}]}]
    eip_topology: list = field(default_factory=list)
    # Fortna octal Word → {rio_name, flex_slot, catalog, direction}
    io_word_map: dict = field(default_factory=dict)
    # Configio.asc: Fortna Octal_Word → [{bank, lohi, desc}, ...] (Reno primary map)
    configio_octal_map: dict = field(default_factory=dict)
    # All photoeyes (for PE_UDT tags + IO_MAP)
    pe_devices: list = field(default_factory=list)
    # Optional gold programs to merge (keys from OPTIONAL_PROGRAMS)
    include_programs: list = field(default_factory=list)
    # Sys gold is OK (timers/nulls).
    # include_io_map: emit RUN/tar.gz IO_MAP program (any site). OFF = no IO_MAP in L5X.
    # include_io_map_gold: optional Greensboro Excel merge (CLI only; blocked if site is CP1–CP4).
    include_sys: bool = True
    include_io_map: bool = True
    include_io_map_gold: bool = False
    # Sorter build UI config (induct / tracking / encoders / divert count)
    sorter_build: dict = field(default_factory=dict)
    # 2:1 merges (PLC2-class transport) — list of dicts from workbook UI
    # keys: name, area, lane_a, lane_b, discharge, pe_a, pe_b, jam_pe
    merges_2to1: list = field(default_factory=list)
    # Tar equipment inventory + build plan (fortna_equipment_plan)
    equipment_plan: dict = field(default_factory=dict)
    # Transport Build stubs not present in active RUN (PRISM twin gaps)
    transport_stub_tags: list = field(default_factory=list)


def load_program_export(path: Path) -> dict | None:
    """
    Load a Studio Program-export L5X (TargetType=Program) from tools/libraries/programs.

    Returns {name, program_xml, tags: [Tag blocks], datatypes_xml?, aois_xml?}.
    """
    path = Path(path)
    if not path.is_file():
        return None
    text = path.read_text(encoding="utf-8", errors="replace").lstrip("\ufeff")
    m = re.search(r'<Program\s+Use="Target"\s+([^>]+)>(.*?)</Program>', text, re.S)
    if not m:
        # Some exports may omit Use="Target"
        m = re.search(r'<Program\s+Name="([^"]+)"([^>]*)>(.*?)</Program>', text, re.S)
        if not m:
            return None
        name = m.group(1)
        attrs = f'Name="{name}"' + m.group(2)
        body = m.group(3)
    else:
        attrs = m.group(1)
        body = m.group(2)
        nm = re.search(r'Name="([^"]+)"', attrs)
        name = nm.group(1) if nm else path.stem.replace("_Program", "")

    # Controller L5X programs are not Use=Target
    program_xml = f"<Program {attrs}>{body}</Program>"
    program_xml = re.sub(r'\s*Use="Target"', "", program_xml, count=1)

    tags: list[str] = []
    # Context tags (full Base tags with Data) sit under Controller Use=Context
    ctx = re.search(r'<Tags\s+Use="Context"[^>]*>(.*?)</Tags>', text, re.S)
    if ctx:
        for tm in re.finditer(r"<Tag\b[^>]*>.*?</Tag>", ctx.group(1), re.S):
            tags.append(tm.group(0))
    # Also program-scoped tags if present
    prog_tags = re.search(r"<Program[^>]*>\s*<Tags>(.*?)</Tags>", text, re.S)
    if prog_tags and prog_tags.group(1).strip():
        for tm in re.finditer(r"<Tag\b[^>]*>.*?</Tag>", prog_tags.group(1), re.S):
            tags.append(tm.group(0))

    # Program exports use <DataTypes Use="Context"> / <AddOnInstructionDefinitions Use="Context">
    # — must allow attributes or WCS/Sorter packs merge ZERO types/AOIs (Studio: Data type not found).
    dt = re.search(r"<DataTypes\b[^>]*>.*?</DataTypes>", text, re.S)
    aoi = re.search(
        r"<AddOnInstructionDefinitions\b[^>]*>.*?</AddOnInstructionDefinitions>",
        text,
        re.S,
    )
    # Normalize wrappers to plain tags so controller merge accepts them
    dt_xml = ""
    if dt:
        dt_xml = re.sub(
            r"<DataTypes\b[^>]*>",
            "<DataTypes>",
            dt.group(0),
            count=1,
        )
    aoi_xml = ""
    if aoi:
        aoi_xml = re.sub(
            r"<AddOnInstructionDefinitions\b[^>]*>",
            "<AddOnInstructionDefinitions>",
            aoi.group(0),
            count=1,
        )
    return {
        "name": name,
        "program_xml": program_xml,
        "tags": tags,
        "datatypes_xml": dt_xml,
        "aois_xml": aoi_xml,
        "source": str(path),
        "tag_count": len(tags),
    }


def resolve_program_exports(
    include_optional: list[str] | None = None,
    *,
    include_sys: bool = True,
    include_io_map_gold: bool = False,
    programs_dir: Path | None = None,
) -> list[dict]:
    """Load Sys / optional gold programs. Gold IO_MAP only if explicitly requested."""
    pdir = Path(programs_dir) if programs_dir else PROGRAM_LIBRARY_DIR
    wanted: list[tuple[str, str]] = []
    if include_sys and "Sys" in ALWAYS_PROGRAMS:
        wanted.append(("Sys", ALWAYS_PROGRAMS["Sys"]))
    if include_io_map_gold and "IO_MAP" in ALWAYS_PROGRAMS:
        wanted.append(("IO_MAP", ALWAYS_PROGRAMS["IO_MAP"]))
    for key in include_optional or []:
        k = (key or "").strip()
        if not k:
            continue
        # Accept short aliases
        aliases = {
            "shippingsorter": "ShippingSorter_Area_L3",
            "shipping_sorter": "ShippingSorter_Area_L3",
            "shippingsorter_shoe": "ShippingSorter_Area_L3",
            "shippingsorter_shoesorter": "ShippingSorter_Area_L3",
            # PopUp Divert — no gold L5X yet; skip silently (UI option only)
            "shippingsorter_popup_divert": "",
            "shippingsorter_popupdivert": "",
            "wcs": "WCS_Interface_TCP_IP",
            "wcs_interface": "WCS_Interface_TCP_IP",
            "sorter_track": "Sorter_Track",
            "sortertrack": "Sorter_Track",
            "sawtooth": "Sawtooth_Merge",
            "sawtooth_merge": "Sawtooth_Merge",
        }
        k2 = aliases.get(k.lower().replace(" ", "_").replace("-", "_"), k)
        if k2 == "" or k in (
            "ShippingSorter_PopUp_Divert",
            "ShippingSorter_PopUpDivert",
        ):
            # Placeholder pack — no program file to merge yet
            continue
        # Sorter_Track gold pack is Greensboro-fixed (~15 diverts). Live build
        # replaces it when sorter_build config is present (handled in build_l5x).
        if k2 == "Sorter_Track":
            continue
        fname = OPTIONAL_PROGRAMS.get(k2) or OPTIONAL_PROGRAMS.get(k)
        if not fname:
            # allow exact filename
            if (pdir / k).is_file():
                loaded = load_program_export(pdir / k)
                if loaded:
                    wanted.append((loaded["name"], k))
            continue
        wanted.append((k2, fname))

    out: list[dict] = []
    seen_names: set[str] = set()
    for _key, fname in wanted:
        path = pdir / fname
        loaded = load_program_export(path)
        if not loaded:
            continue
        if loaded["name"] in seen_names:
            continue
        seen_names.add(loaded["name"])
        out.append(loaded)
    return out


def _safe(s: str) -> str:
    """Studio/Logix tag name: letters, digits, underscore; must not start with a digit."""
    t = re.sub(r"[^A-Za-z0-9_]", "_", (s or "").strip())
    t = re.sub(r"_+", "_", t).strip("_")
    if not t:
        return "Tag"
    # 1PBSTART / 7ES are valid Fortna names but illegal Logix identifiers
    if t[0].isdigit():
        t = f"T_{t}"
    return t[:40]


def _xml_escape(s: str) -> str:
    return (
        (s or "")
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


# ---------------------------------------------------------------------------
# Excel reader
# ---------------------------------------------------------------------------

def inspect_excel(path: Path) -> dict:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=False, keep_vba=True)
    info = {
        "path": str(path),
        "sheets": wb.sheetnames,
        "has_vba": "vbaProject.bin" in path.read_bytes()[:100] or True,
        "notes": [
            "Edit only Inputdata (conveyors) and IO (modules/points).",
            "Config / Dict / IO Dict / Template are formula engines — look locked.",
            "Enable macros (Trust Center) for VBA Generate buttons.",
            "ActiveX controls require Excel desktop (not Excel Online).",
        ],
    }
    if "Inputdata" in wb.sheetnames:
        ws = wb["Inputdata"]
        info["project_name"] = ws["B2"].value
        info["processor"] = ws["B3"].value
        info["version"] = f"{ws['B4'].value}.{ws['C4'].value}"
        n = 0
        for r in range(17, 5000):
            if ws.cell(r, 5).value:
                n += 1
            elif n and r > 30:
                break
        info["conveyor_rows"] = n
    return info


def load_from_excel(path: Path) -> AutogenInput:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, keep_vba=False)
    if "Inputdata" not in wb.sheetnames:
        raise ValueError("Excel missing Inputdata sheet")

    ws = wb["Inputdata"]
    inp = AutogenInput(
        project_name=str(ws["B2"].value or "Autogen_Project").strip(),
        processor=str(ws["B3"].value or "1756-L83E").strip(),
        major_rev=str(ws["B4"].value or "35").strip(),
        minor_rev=str(ws["C4"].value or "00").strip().zfill(2),
    )

    # Areas row 7 C..V
    seen_areas: set[str] = set()
    for c in range(3, 23):
        v = ws.cell(7, c).value
        if not v:
            continue
        s = str(v).strip()
        if not s or "Enter" in s or "Area From" in s:
            continue
        if s not in seen_areas:
            seen_areas.add(s)
            inp.areas.append(s)

    # Safety zones row 10 C..
    for c in range(3, 25):
        v = ws.cell(10, c).value
        if v and str(v).strip():
            s = str(v).strip()
            if s not in inp.safety_zones and "Safe Zone" not in s:
                inp.safety_zones.append(s)

    # Conveyor table from row 17
    for r in range(17, 9000):
        conv = ws.cell(r, 5).value
        if not conv:
            if r > 50 and not any(ws.cell(r, c).value for c in range(1, 8)):
                # allow sparse but stop after long empty streak
                empty = all(not ws.cell(rr, 5).value for rr in range(r, min(r + 20, 9000)))
                if empty:
                    break
            continue
        row = ConveyorRow(
            number=int(ws.cell(r, 1).value or (r - 16)),
            system=str(ws.cell(r, 2).value or "").strip(),
            main_area=str(ws.cell(r, 3).value or "").strip(),
            safety_zone=str(ws.cell(r, 4).value or "").strip(),
            conveyor=str(conv).strip(),
            type=str(ws.cell(r, 6).value or "").strip(),
            downstream=str(ws.cell(r, 7).value or "").strip(),
            track=str(ws.cell(r, 8).value or "").strip(),
            exit_pe=str(ws.cell(r, 9).value or "").strip(),
            entry_pe=str(ws.cell(r, 10).value or "").strip(),
            induct=str(ws.cell(r, 11).value or "").strip(),
            jam=str(ws.cell(r, 12).value or "").strip(),
            height=str(ws.cell(r, 13).value or "").strip(),
            full=str(ws.cell(r, 14).value or "").strip(),
            add_pe=str(ws.cell(r, 15).value or "").strip(),
            motor_starter=str(ws.cell(r, 16).value or "").strip(),
            espb=str(ws.cell(r, 17).value or "").strip(),
            espc=str(ws.cell(r, 18).value or "").strip(),
            control_station=str(ws.cell(r, 19).value or "").strip(),
            power_supply=str(ws.cell(r, 20).value or "").strip(),
        )
        if row.clean_name:
            inp.conveyors.append(row)
            if row.main_area and row.main_area not in inp.areas:
                inp.areas.append(row.main_area)

    if "IO" in wb.sheetnames:
        iows = wb["IO"]
        for r in range(6, 2000):
            mod = iows.cell(r, 2).value
            if mod:
                inp.modules.append(
                    IoModule(
                        name=str(mod).strip(),
                        type=str(iows.cell(r, 3).value or "").strip(),
                        slot=str(iows.cell(r, 4).value or "").strip(),
                        ip=str(iows.cell(r, 5).value or "").strip(),
                        parent=str(iows.cell(r, 6).value or "").strip(),
                    )
                )
            dev = iows.cell(r, 8).value
            if dev:
                inp.io_points.append(
                    IoPoint(
                        device_name=str(dev).strip(),
                        device_type=str(iows.cell(r, 9).value or "").strip(),
                        source_module=str(iows.cell(r, 10).value or "").strip(),
                        port=str(iows.cell(r, 11).value or "").strip(),
                        direction=str(iows.cell(r, 13).value or "").strip().upper()[:1],
                        slot_or_port=str(iows.cell(r, 14).value or "").strip(),
                    )
                )

    return inp


def load_from_json(path: Path) -> AutogenInput:
    data = json.loads(path.read_text(encoding="utf-8"))
    convs = [ConveyorRow(**c) if isinstance(c, dict) else c for c in data.get("conveyors", [])]
    mods = [IoModule(**m) if isinstance(m, dict) else m for m in data.get("modules", [])]
    pts = [IoPoint(**p) if isinstance(p, dict) else p for p in data.get("io_points", [])]
    return AutogenInput(
        project_name=data.get("project_name", "Autogen_Project"),
        processor=data.get("processor", "1756-L83E"),
        major_rev=str(data.get("major_rev", "35")),
        minor_rev=str(data.get("minor_rev", "00")).zfill(2),
        areas=list(data.get("areas") or []),
        safety_zones=list(data.get("safety_zones") or []),
        conveyors=convs,
        modules=mods,
        io_points=pts,
    )


def _area_from_conveyor_name(name: str, fallback: str = "Main_Area") -> str:
    """P309 → Zone3_Area style bucket (autogen wants Main_Area names)."""
    m = re.match(r"^P(\d)", (name or "").upper())
    if m:
        return f"Zone{m.group(1)}_Area"
    return fallback


def _classify_pe_role(name: str, desc: str = "") -> str:
    """Classify Fortna PE into product | full | jam | other (Excel-style roles)."""
    n = (name or "").upper()
    d = (desc or "").upper()
    # Full eyes first (EZPE116_F, PE704A_F, EZPE530_F1)
    if re.search(r"_F\d*$", n) or n.endswith("_F") or "FULL EYE" in d or "FULL DETECTION" in d:
        if "FULL" in d or re.search(r"_F\d*$", n) or n.endswith("_F"):
            # jam-full hybrid suffixes like _JF stay jam
            if "_JF" in n or "_FDJ" in n:
                return "jam"
            return "full"
    if "_JF" in n or "_FDJ" in n or re.search(r"_J\d*$", n) or n.endswith("_J") or "JAM" in d:
        return "jam"
    if (
        re.search(r"_P\d*$", n)
        or n.endswith("_P")
        or "PRODUCT" in d
        or "DISCHARGE" in d
        or "PRESENT" in d
    ):
        return "product"
    return "other"


def _link_pe_to_conveyor(p: dict) -> str:
    """Resolve PE → P### conveyor name from RUN description / tag."""
    link = (p.get("conveyor") or "").upper()
    if link and re.match(r"^P\d", link):
        return link
    desc = (p.get("description") or "").upper()
    name = (p.get("fortna_name") or p.get("io_name") or "").upper()
    m = re.search(r"\bP(\d{2,4}[A-Z]?)\b", desc)
    if m:
        return f"P{m.group(1)}"
    m = re.search(r"(?:EZ)?PE(\d{2,4}[A-Z]?)", name)
    if m:
        return f"P{m.group(1)}"
    return ""


def _pe_wiring_for_conv(pe_rows: list[dict]) -> dict:
    """
    Build Excel-parity PE wiring for one conveyor from RUN photoeye rows.

    Returns tag names (not Yes/No options) for Fast_Conv / Slow_Jam / Full_PE.
    """
    product: list[str] = []
    full: list[str] = []
    jam: list[str] = []
    other: list[str] = []
    for p in pe_rows:
        raw = (p.get("fortna_name") or p.get("io_name") or "").strip()
        if not raw:
            continue
        tag = _safe(raw)
        role = _classify_pe_role(raw, p.get("description") or "")
        if role == "full":
            full.append(tag)
        elif role == "jam":
            jam.append(tag)
        elif role == "product":
            product.append(tag)
        else:
            other.append(tag)

    # Slow_Jam slots: product eyes first, then jam, then other (max 5); full eyes are Full_PE
    jam_slots: list[str] = []
    for src in (product, jam, other):
        for t in src:
            if t not in jam_slots:
                jam_slots.append(t)
            if len(jam_slots) >= 5:
                break
        if len(jam_slots) >= 5:
            break

    exit_tag = product[0] if product else (jam_slots[0] if jam_slots else "")
    add_tag = product[1] if len(product) > 1 else ""
    all_tags = list(dict.fromkeys(product + full + jam + other))

    return {
        "exit_pe_tag": exit_tag,
        "add_pe_tag": add_tag,
        "jam_pe_tags": jam_slots[:5],
        "full_pe_tags": full,
        "product_pe_tags": product,
        "all_pe_tags": all_tags,
        "full_opt": "Yes Standard Logic" if full else "",
        "exit_opt": "Yes Standard when Jam reset and PE clear" if exit_tag else "",
        "jam_opt": "Yes" if jam_slots else "",
    }


# Library templates → Rockwell catalog + AB: config datatype (from OReilly_Library_v3)
# 1794 = Flex I/O (Greensboro gold). 1734 = POINT I/O (MSC Reno tar EIPModules).
EIP_CHILD_TEMPLATE = {
    # Flex I/O (1794) — library IO_1N90_*
    "1794-IA16": "IO_1N90_1",   # AB:1794_DI_Delay16:C:0, catalog 1794-IA16/A
    "1794-OA8I": "IO_1N90_2",   # AB:1794_DO8:C:0
    "1794-OW8": "IO_1N90_3",    # AB:1794_DO8:C:0
    "1794-IB16": "IO_1N90_4",   # AB:1794_IB16:C:0
    "1794-OB16P": "IO_1N90_5",  # AB:1794_DO16:C:0
    # POINT I/O (1734) — library IO_1N80_*
    "1734-IB8": "IO_1N80_1",    # AB:1734_DI8:C:0
    "1734-OB8E": "IO_1N80_2",   # AB:1734_DOB8:C:0
    "1734-OB8": "IO_1N80_2",    # closest — no bare OB8 template
    "1734-IA4": "IO_1N80_3",    # AB:1734_DI4:C:0
    "1734-OA4": "IO_1N80_4",    # AB:1734_DO4:C:0
}
EIP_CATALOG = {
    "1794-AENT": "1794-AENT",
    "1794-IA16": "1794-IA16/A",
    "1794-OA8I": "1794-OA8I/A",
    "1794-OW8": "1794-OW8/A",
    "1794-IB16": "1794-IB16/A",
    "1794-OB16P": "1794-OB16P/A",
    # POINT — library uses AENTR revision suffix
    "1734-AENT": "1734-AENTR/C",
    "1734-AENTR": "1734-AENTR/C",
    "1734-IB8": "1734-IB8/C",
    "1734-OB8E": "1734-OB8E/C",
    "1734-OB8": "1734-OB8E/C",
    "1734-IA4": "1734-IA4/C",
    "1734-OA4": "1734-OA4/C",
}
EIP_PARENT_TEMPLATE = {
    "1794": "IO_1N90",   # AB:1794_AEN_8SLOT — Bus Size 8
    "1734": "IO_1N80",   # AB:1734_40SLOT — Bus Size 40
}
EIP_PARENT_BUS_SIZE = {
    "1794": 8,
    "1734": 40,
}
EIP_PARENT_CATALOG = {
    "1794": "1794-AENT",
    "1734": "1734-AENTR/C",
}
EIP_PARENT_TYPE = {
    "1794": "1794-AENT",
    "1734": "1734-AENT",
}


def _eip_family_from_types(types: list[str]) -> str:
    """Return '1734' (POINT) or '1794' (Flex) from module type strings in the tar."""
    joined = " ".join(types or []).upper()
    if "1734" in joined or "1738" in joined:
        return "1734"
    if "1794" in joined:
        return "1794"
    # Default Flex (legacy Greensboro path)
    return "1794"


def _fortna_bit_to_data_bit(bit: str | int, *, max_bit: int = 15) -> int | None:
    """Fortna PE bits → Logix Data bit.

    Prefer PLC-5 octal (0-7, 10-17 → 0-15) for 16-pt Flex cards.
    For 1734 4/8-pt POINT cards, clamp to max_bit (3 or 7); if octal
    overshoots, retry as decimal.
    """
    s = str(bit or "").strip()
    if not s:
        return None
    val: int | None = None
    try:
        val = int(s, 8)
    except ValueError:
        try:
            val = int(s, 10)
        except ValueError:
            return None
    if val is None:
        return None
    if val > max_bit:
        try:
            val = int(s, 10)
        except ValueError:
            return None
    if val < 0 or val > max_bit:
        return None
    return val


def _point_card_max_bit(mod_type: str) -> int:
    """Max Data bit index for a POINT/Flex card type."""
    mt = (mod_type or "").upper()
    if "IA4" in mt or "OA4" in mt or "IB4" in mt or "OB4" in mt:
        return 3
    if "IB8" in mt or "OB8" in mt or "IA8" in mt or "OA8" in mt:
        return 7
    return 15


def _io_point_want_dir(device_name: str, device_type: str, direction: str) -> str:
    """Desired card direction for a field device (I=input card, O=output card)."""
    n = (device_name or "").upper()
    dt = (device_type or "").lower()
    # Solenoids / pusher outputs before generic rules
    if dt in ("digital_out", "solenoid", "output") or re.search(r"SSV", n):
        return "O"
    # Feedback / sense points are always inputs
    if n.endswith("_AUX") or dt in ("photoeye", "pushbutton", "digital_in", "encoder"):
        return "I"
    if n.startswith("ENC") or "ENCODER" in n:
        return "I"
    # E-stop PB / pullcord feedbacks are inputs; MCR/ES *coil* tags are outputs
    if dt in ("estop", "e-stop", "e_stop", "es"):
        if re.search(r"MCR\d*$", n) or re.match(r"^\d*ES\d+$", n) or re.match(r"^ES\d+$", n):
            # 14MCR1 / 14ES1 coil (no _AUX) → output; 14ES1_AUX already caught above
            if not n.endswith("_AUX"):
                d = (direction or "").upper()
                if d in ("O", "OUT", "OUTPUT"):
                    return "O"
                # Default: bare MCR/ES coil is an energize output on Fortna prints
                if "MCR" in n:
                    return "O"
        return "I"
    if dt in ("beacon",):
        return "O"
    d = (direction or "").upper()
    if d in ("O", "OUT", "OUTPUT"):
        return "O"
    return "I"


def _fortna_bit_is_high(bit: str | int) -> bool:
    """True when Fortna IO_Address_Bit is the high half (octal 10-17 → Data 8-15)."""
    s = str(bit or "").strip()
    if not s:
        return False
    try:
        v = int(s, 8)
    except ValueError:
        try:
            v = int(s, 10)
        except ValueError:
            return False
    return v >= 8


def _rio_numeric_key(rio: str) -> tuple:
    """Sort key so AENTR3 < AENTR5 < AENTR13 < AENTR14 < AENTR14RP1."""
    s = (rio or "").strip().upper()
    m = re.match(r"AENTR(\d+)(?:RP(\d+))?", s)
    if m:
        return (int(m.group(1)), int(m.group(2) or 0), s)
    m = re.search(r"(\d+)", s)
    if m:
        return (int(m.group(1)), 0, s)
    return (9999, 0, s)


def _load_configio_octal_map(run_dir: Path, machine: str = "") -> dict[int, list[dict]]:
    """Load FORTNA/Configio.asc[.MACHINE] → Octal_Word → [{bank, lohi, desc}].

    This is the authoritative Fortna word→EIP bank map when EIPCSV is empty
    (MSC Reno). Example: Octal 400 Low→bank 312 (AENTR14 IA4), 1047→banks 15/16.
    """
    from fortna_asc import read_asc

    run_dir = Path(run_dir)
    if (run_dir / "RUN" / "project.cfg").is_file():
        run_dir = run_dir / "RUN"
    fortna = run_dir / "FORTNA"
    mach = (machine or "").strip()
    candidates: list[Path] = []
    if mach:
        candidates.append(fortna / f"Configio.asc.{mach}")
    candidates.append(fortna / "Configio.asc")
    path = next((p for p in candidates if p.is_file()), None)
    if not path:
        return {}
    try:
        _, rows = read_asc(path)
    except Exception:
        return {}
    out: dict[int, list[dict]] = {}
    for r in rows:
        try:
            octal = int(float(r.get("Octal_Word") or 0))
            bank = int(float(r.get("Bank") or -1))
        except Exception:
            continue
        # Octal_Word 0 rows are unused placeholders. Bank 0 is valid —
        # AENTR3 slot14 OB8E uses OutputBank 0 (MX1–MX6 SSV pushers).
        if octal <= 0 or bank < 0:
            continue
        lohi = (r.get("LoHi") or "").strip() or "Low"
        desc = (r.get("Desc") or "").strip()
        out.setdefault(octal, []).append({
            "bank": bank,
            "lohi": lohi,
            "desc": desc,
        })
    return out


def _build_eip_bank_index(topology: list[dict]) -> dict[int, list[dict]]:
    """Map Fortna/EIP bank number → candidate module slots (may be multiple)."""
    bank_map: dict[int, list[dict]] = {}
    for ad in topology or []:
        rio = ad.get("rio_name") or ""
        family = ad.get("family") or "1794"
        for c in ad.get("children") or []:
            direction = (c.get("direction") or "").upper()
            # Index input cards by InputBank only; output cards by OutputBank only.
            # OB8E status InputBank must NOT steal discrete input words (e.g. 400).
            keys: list[str] = []
            if direction == "I":
                keys = ["input_bank", "word"]
            elif direction == "O":
                keys = ["output_bank", "word"]
            else:
                keys = ["input_bank", "output_bank", "word"]
            for key in keys:
                raw = c.get(key)
                if raw is None or str(raw).strip() == "":
                    continue
                try:
                    b = int(raw)
                except Exception:
                    continue
                # InputBank 0 is unused placeholder; OutputBank 0 is valid (AENTR3 OB8E).
                if key == "input_bank" and b <= 0:
                    continue
                if key == "output_bank" and b < 0:
                    continue
                if key == "word" and b < 0:
                    continue
                if direction == "I" and b <= 0:
                    continue
                bank_map.setdefault(b, []).append({
                    "rio_name": rio,
                    "child_name": c.get("name") or "",
                    "flex_slot": int(c.get("flex_slot") or 0),
                    "type": c.get("type") or "",
                    "direction": direction,
                    "family": family or c.get("family") or "1794",
                    "via": key,
                })
    return bank_map


# Configio Desc → AENTR rio name (bank numbers collide across adapters!)
_CONFIGIO_DESC_TO_RIO = {
    "EP3RP": "AENTR3",
    "EP4RP": "AENTR4",
    "EP5RP": "AENTR5",
    "EP6RP": "AENTR6",
    "EP7RP": "AENTR7",
    "CP13": "AENTR13",
    "13RP1": "AENTR13RP1",
    "13RP2": "AENTR13RP2",
    "CP14": "AENTR14",
    "14RP1": "AENTR14RP1",
    "14RP2": "AENTR14RP2",
}


def _configio_desc_to_rio(desc: str) -> str:
    d = (desc or "").strip().upper()
    if not d:
        return ""
    for key, rio in _CONFIGIO_DESC_TO_RIO.items():
        if d == key.upper():
            return rio
    # Fallback: EP3 → AENTR3, CP14 → AENTR14
    m = re.match(r"^EP(\d+)RP$", d)
    if m:
        return f"AENTR{m.group(1)}"
    m = re.match(r"^CP(\d+)$", d)
    if m:
        return f"AENTR{m.group(1)}"
    m = re.match(r"^(\d+)RP(\d+)$", d)
    if m:
        return f"AENTR{m.group(1)}RP{m.group(2)}"
    return ""


def _resolve_via_configio(
    word: int,
    bit: str,
    *,
    want_dir: str,
    bank_index: dict[int, list[dict]],
    configio_map: dict[int, list[dict]],
) -> dict | None:
    """Resolve Fortna Octal_Word via Configio → EIP bank → module.

    Critical:
      1) Scope by Configio Desc→AENTR* (bank 68 is IB on AENTR3 AND OB on AENTR14RP2).
      2) Filter by want_dir so mixed I/O words (1022 Low=IB High=OB) never mirror.
      3) When both halves match want_dir, pick by bit half using bank NUMBER order.
    """
    entries = configio_map.get(word) or []
    if not entries:
        return None
    want = (want_dir or "I").upper()

    # (lohi, bank, info) for cards on the Configio adapter with matching direction
    matched: list[tuple[str, int, dict]] = []
    for e in entries:
        side = (e.get("lohi") or "").strip() or "Low"
        try:
            b = int(e.get("bank"))
        except Exception:
            continue
        if b < 0:
            continue
        expect_rio = _configio_desc_to_rio(e.get("desc") or "")
        for info in bank_index.get(b) or []:
            if (info.get("direction") or "").upper() != want:
                continue
            rio = (info.get("rio_name") or "").upper()
            if expect_rio and rio != expect_rio.upper():
                continue
            matched.append((side, b, info))
            break

    if not matched:
        return None

    banks = sorted({b for _s, b, _i in matched})
    if len(banks) >= 2:
        target = banks[1] if _fortna_bit_is_high(bit) else banks[0]
    else:
        highs = [b for s, b, _i in matched if s == "High"]
        lows = [b for s, b, _i in matched if s == "Low"]
        if _fortna_bit_is_high(bit) and highs:
            target = highs[0]
        elif not _fortna_bit_is_high(bit) and lows:
            target = lows[0]
        else:
            target = banks[0]

    for _side, b, info in matched:
        if b == target:
            out = dict(info)
            out["resolved_bank"] = target
            out["resolve_how"] = "configio"
            return out
    out = dict(matched[0][2])
    out["resolved_bank"] = matched[0][1]
    out["resolve_how"] = "configio"
    return out


def _resolve_fortna_bank(
    bank: str | int,
    *,
    want_dir: str,
    bank_index: dict[int, list[dict]],
    bit: str = "",
    configio_map: dict[int, list[dict]] | None = None,
) -> dict | None:
    """Resolve Conveyor.asc IO_Address_Word → EIP module slot.

    Priority (Reno / empty EIPCSV):
      1) Configio.asc Octal_Word → Bank (authoritative)
      2) direct bank number
      3) fragile heuristics (m1000/m800/oct) — last resort only
    Only accepts cards whose direction matches want_dir when possible.
    """
    try:
        w = int(float(str(bank).strip()))
    except Exception:
        return None
    if w <= 0:
        return None
    want = (want_dir or "I").upper()
    if configio_map:
        hit = _resolve_via_configio(
            w, bit, want_dir=want, bank_index=bank_index, configio_map=configio_map
        )
        if hit:
            return hit
    trials: list[tuple[str, int]] = [("direct", w)]
    # Heuristics only when Configio has no row for this word
    if not (configio_map and w in configio_map):
        if w >= 1000:
            trials.append(("m1000", w - 1000))
        if w >= 800:
            trials.append(("m800", w - 800))
        s = str(w)
        if re.fullmatch(r"[0-7]+", s):
            try:
                trials.append(("oct", int(s, 8)))
            except Exception:
                pass
    for _how, val in trials:
        if val <= 0:
            continue
        for info in bank_index.get(val) or []:
            if (info.get("direction") or "").upper() == want:
                out = dict(info)
                out["resolved_bank"] = val
                out["resolve_how"] = _how
                return out
    return None


def _safe_rio_name(adapter_name: str, fallback: str) -> str:
    """Prefer Fortna adapter name (AENTR13) over invented CPxRIOn."""
    raw = (adapter_name or "").strip()
    if not raw or raw.upper() in ("N/A", "INVALID", "NONE"):
        return fallback
    t = _safe(raw)
    if not t or t.lower() in ("tag", "n_a"):
        return fallback
    return t[:40]


# MSC Reno SHIP: ASC/eipcfg uses AENTR-1..4; prints use AENTR15 / RP1 / RP2 / AENTR16
_SHIP_AENTR_PRINT_NAMES = {
    "AENTR-1": "AENTR15",
    "AENTR_1": "AENTR15",
    "AENTR1": "AENTR15",
    "AENTR-2": "AENTR15RP1",
    "AENTR_2": "AENTR15RP1",
    "AENTR2": "AENTR15RP1",
    "AENTR-3": "AENTR15RP2",
    "AENTR_3": "AENTR15RP2",
    "AENTR3": "AENTR15RP2",
    "AENTR-4": "AENTR16",
    "AENTR_4": "AENTR16",
    "AENTR4": "AENTR16",
}


def _print_rio_name(adapter_name: str, machine: str = "") -> str:
    """Map tar adapter names to print / Studio names when they differ."""
    raw = (adapter_name or "").strip()
    mach = (machine or "").upper()
    if "SHIP" in mach:
        hit = _SHIP_AENTR_PRINT_NAMES.get(raw) or _SHIP_AENTR_PRINT_NAMES.get(
            raw.replace("_", "-")
        ) or _SHIP_AENTR_PRINT_NAMES.get(_safe(raw))
        if hit:
            return hit
    return _safe_rio_name(raw, raw[:40] if raw else "RIO")


# Gold Greensboro controller stems that appear in Sys / AOI context exports
_GOLD_SITE_NAME_RE = re.compile(
    r"ORLY_Greensboro_NC_(?:PLC|CP)\d+"
    r"|OReillyGreensboro_ORNCCP\d+"
    r"|OReillyGreensboro_[A-Za-z0-9]+"
    r"|ORLY\s+Greensboro\s+NC\s+PLC\d+",
    re.I,
)


def _retarget_gold_site_names(xml: str, site_name: str) -> str:
    """Rewrite gold Greensboro controller/tag names to this site's project name.

    Sys_Program.L5X and AOI context exports embed ORLY_Greensboro_NC_PLC5_System, etc.
    """
    site = _safe(site_name) or "Site"
    if not xml:
        return xml
    return _GOLD_SITE_NAME_RE.sub(site, xml)


def _infer_rack_from_ip(ip: str, fallback: str = "CP5") -> str:
    ip = (ip or "").strip()
    m = re.search(r"\.(\d+)$", ip)
    if not m:
        return fallback
    last = int(m.group(1))
    if 51 <= last <= 53:
        return "CP5"
    if 54 <= last <= 56:
        return "CP6"
    if 57 <= last <= 58:
        return "CP7"
    return fallback


def _synthesize_point_io_banks(adapters: list[dict]) -> None:
    """Fill InputBank/OutputBank when EIPModules left them at 0 (MSC Reno SHIP).

    PACK stamps real banks on every card. SHIP leaves IB=0/OB=0, so Configio
    octal→bank cannot resolve to a slot. Mirror Fortna RTA layout:
      first child InputBank = adapter.InputAddress + 8
      first child OutputBank = adapter.OutputAddress
      each IB8/IA4/IM consumes one InputBank
      each OA4/OB consumes one OutputBank
      OB8E also consumes one InputBank (channel status word)
    Only runs when every bridged card on an adapter still has IB=0 and OB=0.
    """
    for ad in adapters or []:
        mods = list(ad.get("modules") or [])
        bridged = sorted(
            [
                m for m in mods
                if (m.get("connection") or "").upper() != "HEADNODE"
                and "AENT" not in (m.get("type") or "").upper()
            ],
            key=lambda m: int(m.get("slot") or 0),
        )
        if not bridged:
            continue
        all_zero = True
        for m in bridged:
            try:
                ib = int(m.get("input_bank") or 0)
                ob = int(m.get("output_bank") or 0)
            except Exception:
                ib = ob = 0
            if ib != 0 or ob != 0:
                all_zero = False
                break
        if not all_zero:
            continue
        try:
            in_addr = int(float(ad.get("input_address") or 0))
        except Exception:
            in_addr = 0
        try:
            out_addr = int(float(ad.get("output_address") or 0))
        except Exception:
            out_addr = 0
        if in_addr <= 0 and out_addr < 0:
            continue
        next_ib = in_addr + 8 if in_addr > 0 else 0
        next_ob = out_addr if out_addr >= 0 else 0
        for m in bridged:
            mt = (m.get("type") or "").upper()
            is_in = any(x in mt for x in ("IA", "IB", "IM"))
            is_out = any(x in mt for x in ("OA", "OB", "OW"))
            if is_in:
                m["input_bank"] = next_ib
                next_ib += 1
            if is_out:
                # OB8/OB8E carry a status input image word in Fortna/EIP layout
                if "OB8" in mt or "OB16" in mt:
                    m["input_bank"] = next_ib
                    next_ib += 1
                m["output_bank"] = next_ob
                next_ob += 1


def load_eip_topology(run_dir: Path, *, machine: str = "") -> dict:
    """
    Build named remote I/O tree + Fortna Word→module map from RUN.

    Naming matches edited gold: CP5RIO0, CP5RIO0_0, CP5RIO0_1, …
    Family is taken from EIPModules / eipcfg:
      1794 Flex → parent AB:1794_AEN_8SLOT (Bus 8)
      1734 POINT → parent AB:1734_40SLOT (Bus 40) — MSC Reno

    Fortna Conveyor.asc IO_Address_Word maps via EIPCSV when present; otherwise
    module InputBank/OutputBank is used as the word key (Reno EIPCSV often empty).
    """
    from fortna_asc import read_asc

    run_dir = Path(run_dir)
    if (run_dir / "RUN" / "project.cfg").is_file():
        run_dir = run_dir / "RUN"

    if not machine:
        try:
            from fortna_io_extract import read_project_meta
            machine = (read_project_meta(run_dir).get("machine_name") or "").strip()
        except Exception:
            machine = ""

    result: dict = {
        "interface_ip": "",
        "adapters_raw": [],
        "topology": [],  # named CPxRIOn with children
        "word_map": {},  # str(word) -> mapping
        "modules_flat": [],
        "machine": machine,
    }

    # --- Adapters + modules from ASC (most complete for slot/bank) ---
    adapters: list[dict] = []
    proj = run_dir / "PROJECT"
    ad_path = next(iter(sorted(proj.glob("EIPAdapters.asc*"))), None)
    mod_path = next(iter(sorted(proj.glob("EIPModules.asc*"))), None)
    csv_path = next(iter(sorted(proj.glob("EIPCSV.asc*"))), None)

    # Prefer THIS machine's eipcfg for adapter inventory. Multi-PLC Reno tars
    # often ship a polluted PROJECT/EIPAdapters.asc (PICK contains PACK's
    # AENTR3–7). When machine eipcfg lists AENTR*, treat it as the allowlist.
    eipcfg_allow_names: set[str] = set()
    eipcfg_allow_ips: set[str] = set()
    eipcfg_is_authoritative = False
    try:
        from fortna_ignition_build import load_eip_modules

        eip = load_eip_modules(run_dir)
        result["interface_ip"] = eip.get("interface_ip") or ""
        sources = [str(s).upper() for s in (eip.get("sources") or [])]
        mach_u = (machine or "").upper()
        eipcfg_is_authoritative = bool(
            mach_u
            and any(mach_u in s and "EIPCFG" in s for s in sources)
            and (eip.get("adapters") or [])
        )
        if eip.get("adapters"):
            for a in eip["adapters"]:
                nm = (a.get("name") or "").strip()
                ip = (a.get("ip") or "").strip()
                adapters.append({
                    "name": nm,
                    "ip": ip,
                    "rack": (a.get("rack") or "").strip(),
                    "input_address": a.get("input_address") or "",
                    "output_address": a.get("output_address") or "",
                    "modules": list(a.get("modules") or []),
                })
                if nm:
                    eipcfg_allow_names.add(nm)
                    eipcfg_allow_names.add(nm.replace("-", "_"))
                if ip:
                    eipcfg_allow_ips.add(ip)
    except Exception:
        pass

    # Detect polluted EIPAdapters.asc (PICK tar copies PACK's AENTR3–7 and
    # omits this machine's AENTR1/AENTR2 from eipcfg). If most eipcfg AENTR
    # names are missing from ASC, ignore ASC-only adapter rows.
    skip_asc_new_adapters = False
    if eipcfg_is_authoritative and eipcfg_allow_names and ad_path and ad_path.is_file():
        try:
            _, _asc_probe = read_asc(ad_path)
            asc_names_probe = {
                (r.get("Name") or "").strip()
                for r in _asc_probe
                if (r.get("Name") or "").strip()
                and (r.get("Name") or "").strip().upper() not in ("N/A", "INVALID")
            }
            eip_aentr = {
                n for n in eipcfg_allow_names
                if n.upper().startswith("AENTR")
            }
            missing = eip_aentr - asc_names_probe
            if eip_aentr and len(missing) * 2 >= len(eip_aentr):
                skip_asc_new_adapters = True
        except Exception:
            skip_asc_new_adapters = False

    if ad_path and ad_path.is_file():
        _, rows = read_asc(ad_path)
        by_name = {a["name"]: a for a in adapters if a.get("name")}
        for r in rows:
            name = (r.get("Name") or "").strip()
            if not name or name.upper() in ("N/A", "INVALID"):
                continue
            ip = (r.get("TargetIP") or "").strip()
            rack = (r.get("Rack") or "").strip()
            if name in by_name:
                # Enrich existing eipcfg adapter — do not steal IP from another panel
                if ip and (
                    not eipcfg_is_authoritative
                    or not by_name[name].get("ip")
                    or by_name[name].get("ip") == ip
                ):
                    by_name[name]["ip"] = ip
                if rack:
                    by_name[name]["rack"] = rack
                if not by_name[name].get("input_address"):
                    by_name[name]["input_address"] = r.get("InputAddress") or ""
                if not by_name[name].get("output_address"):
                    by_name[name]["output_address"] = r.get("OutputAddress") or ""
            else:
                if skip_asc_new_adapters:
                    continue
                adapters.append({
                    "name": name,
                    "ip": ip,
                    "rack": rack,
                    "input_address": r.get("InputAddress") or "",
                    "output_address": r.get("OutputAddress") or "",
                    "modules": [],
                })
                by_name[name] = adapters[-1]

    if mod_path and mod_path.is_file():
        _, rows = read_asc(mod_path)
        by_name = {a["name"]: a for a in adapters if a.get("name")}
        for r in rows:
            name = (r.get("Name") or "").strip()
            if not name or name.upper() in ("N/A", "INVALID"):
                continue
            mt = (r.get("Type") or "").strip()
            ad = (r.get("Adapter") or "").strip()
            conn = (r.get("Connection") or "").strip()
            try:
                slot = int(float(r.get("Slot") or 0))
            except Exception:
                slot = 0
            try:
                ib = int(float(r.get("InputBank") or 0))
            except Exception:
                ib = 0
            try:
                ob = int(float(r.get("OutputBank") or 0))
            except Exception:
                ob = 0
            rec = {
                "name": name,
                "type": mt,
                "slot": slot,
                "connection": conn,
                "input_bank": ib,
                "output_bank": ob,
                "word": "",  # filled from EIPCSV
            }
            if ad in by_name:
                existing = by_name[ad].setdefault("modules", [])
                # Machine eipcfg already listed cards for this adapter — do not
                # merge polluted EIPModules from another panel (PICK AENTR3 vs
                # PACK AENTR3 share a name but different IPs/slots).
                if (
                    eipcfg_is_authoritative
                    and eipcfg_allow_names
                    and ad in eipcfg_allow_names
                    and any(
                        (m.get("connection") or "").upper() != "HEADNODE"
                        and "AENT" not in (m.get("type") or "").upper()
                        for m in existing
                    )
                ):
                    continue
                # replace empty modules list from xml with ASC detail
                # avoid dupes by slot
                if not any(int(m.get("slot") or -1) == slot for m in existing):
                    existing.append(rec)
                else:
                    for m in existing:
                        if int(m.get("slot") or -1) == slot:
                            m.update({k: v for k, v in rec.items() if v not in ("", None)})
                            break
            elif conn == "HEADNODE":
                pass

    # Attach words from EIPCSV (Word is what Conveyor.asc IO_Address_Word uses)
    word_to_mod: dict[str, dict] = {}
    if csv_path and csv_path.is_file():
        _, rows = read_asc(csv_path)
        # track last non-empty type/name for continuation rows
        last_type = ""
        last_name = ""
        last_rack = ""
        last_ip = ""
        for r in rows:
            name = (r.get("Name") or "").strip()
            mt = (r.get("Type") or "").strip()
            rack = (r.get("Rack") or "").strip()
            ip = (r.get("IP") or "").strip()
            word = str(r.get("Word") or "").strip()
            io_dir = (r.get("IO") or "").strip().upper()
            bit_rng = (r.get("Bit") or "").strip()
            if name and name.upper() not in ("N/A", "INVALID"):
                last_name = name
            else:
                name = last_name
            if mt:
                last_type = mt
            else:
                mt = last_type
            if rack:
                last_rack = rack
            else:
                rack = last_rack
            if ip:
                last_ip = ip
            else:
                ip = last_ip
            if not word or word in ("0", ""):
                continue
            # Only keep first (low) half for mapping word → card; both halves share word
            key = word
            if key not in word_to_mod or (io_dir == "I" and "00" in bit_rng):
                word_to_mod[key] = {
                    "word": word,
                    "type": mt,
                    "name": name,
                    "rack": rack,
                    "ip": ip,
                    "io": io_dir,
                    "bit_range": bit_rng,
                    "bank": str(r.get("Bank") or "").strip(),
                }

    # SHIP (and similar): EIPModules leaves InputBank/OutputBank at 0 for every
    # POINT card. Configio still has the real Fortna banks. Synthesize from
    # adapter InputAddress/OutputAddress using the same +8 head offset PACK uses
    # (InAddr 48 → first child IB 56). OB8E also consumes a status InputBank.
    _synthesize_point_io_banks(adapters)

    # Stamp words onto adapter modules when type+rack match
    for ad in adapters:
        for m in ad.get("modules") or []:
            mt = (m.get("type") or "").strip()
            for w, info in word_to_mod.items():
                if info.get("type") == mt and info.get("rack") == (ad.get("rack") or info.get("rack")):
                    # match by input bank if available
                    try:
                        ib = int(info.get("bank") or -1)
                    except Exception:
                        ib = -1
                    if ib >= 0 and int(m.get("input_bank") or -2) == ib:
                        m["word"] = w
                        break
            if not m.get("word"):
                # fallback: first word of same type on same rack not yet assigned
                for w, info in word_to_mod.items():
                    if info.get("type") == mt and (
                        not ad.get("rack") or info.get("rack") == ad.get("rack")
                    ):
                        if not any(
                            (om.get("word") == w)
                            for om in (ad.get("modules") or [])
                        ):
                            m["word"] = w
                            break

    # When PROJECT/EIPAdapters.asc lists Fortna AENTR* racks, drop synthetic
    # eipcfg leftovers from other machines in the same tar (1734_AENT_51, VU4…).
    asc_aentr = {
        (a.get("name") or "").strip()
        for a in adapters
        if (a.get("name") or "").upper().startswith("AENTR")
        and (a.get("modules") or [])
    }
    # Also count names that came from ASC file specifically
    asc_names: set[str] = set()
    if ad_path and ad_path.is_file():
        try:
            _, _ad_rows = read_asc(ad_path)
            for r in _ad_rows:
                n = (r.get("Name") or "").strip()
                if n and n.upper() not in ("N/A", "INVALID"):
                    asc_names.add(n)
        except Exception:
            asc_names = set()
    if asc_names:
        fortna_asc = [n for n in asc_names if n.upper().startswith("AENTR")]
        if fortna_asc:
            adapters = [
                a for a in adapters
                if (a.get("name") or "").strip() in asc_names
                or (
                    (a.get("name") or "").upper().startswith("AENTR")
                    and (a.get("modules") or [])
                )
            ]

    # Name adapters CPxRIOn and children CPxRIOn_k
    by_rack: dict[str, list] = {}
    for ad in adapters:
        rack = (ad.get("rack") or "").strip().upper()
        if not rack:
            rack = _infer_rack_from_ip(ad.get("ip") or "")
            ad["rack"] = rack
        by_rack.setdefault(rack or "CP5", []).append(ad)

    topology: list[dict] = []
    word_map: dict[str, dict] = {}
    modules_flat: list[IoModule] = []

    for rack in sorted(by_rack.keys()):
        # Gold Excel / IO_MAP names CP7 heads CP7RIO1 + CP7RIO2 (no CP7RIO0).
        # CP5/CP6 use RIO0..n. Mismatch left gold IO_MAP OTE(CP7RIO2:…) undefined.
        rio_start = 1 if str(rack).upper() in ("CP7", "PLC7") else 0
        # Dedupe adapters that share an IP (eipcfg synthetic 1734_AENT_51 + ASC AENTR13).
        # Prefer Fortna print names: AENTR* over invented 1734_AENT_* / CPxRIO*.
        ads_in = list(by_rack[rack])
        by_ip: dict[str, dict] = {}
        no_ip: list[dict] = []

        def _name_rank(n: str) -> tuple:
            u = (n or "").upper()
            if u.startswith("AENTR"):
                return (0, u)
            if "1734_AENT" in u or "1738_AENT" in u:
                return (2, u)
            return (1, u)

        for ad in ads_in:
            ip_k = (ad.get("ip") or "").strip()
            if not ip_k:
                no_ip.append(ad)
                continue
            prev = by_ip.get(ip_k)
            if not prev or _name_rank(ad.get("name") or "") < _name_rank(prev.get("name") or ""):
                # Keep richer module list
                if prev and len(prev.get("modules") or []) > len(ad.get("modules") or []):
                    ad = {**ad, "modules": prev.get("modules") or ad.get("modules") or []}
                elif prev and not ad.get("modules") and prev.get("modules"):
                    ad = {**ad, "modules": prev["modules"]}
                by_ip[ip_k] = ad
            elif prev and len(ad.get("modules") or []) > len(prev.get("modules") or []):
                by_ip[ip_k] = {**prev, "modules": ad["modules"], "name": prev.get("name") or ad.get("name")}
        ads_deduped = list(by_ip.values()) + no_ip

        used_rio_names: set[str] = set()
        for idx, ad in enumerate(ads_deduped):
            # Prints / eipcfg use Fortna names (AENTR13, AENTR13RP1) — prefer those.
            # SHIP ASC uses AENTR-1..4 → print names AENTR15 / RP1 / RP2 / AENTR16.
            fallback = f"{rack}RIO{idx + rio_start}"
            raw_name = ad.get("name") or ""
            rio = _print_rio_name(raw_name, machine) if raw_name else ""
            if not rio or rio.lower() in ("tag", "n_a"):
                rio = _safe_rio_name(raw_name, fallback)
            if rio in used_rio_names:
                rio = f"{rio}_{idx + rio_start}"
            used_rio_names.add(rio)
            ip = (ad.get("ip") or "").strip()
            children = []
            # Bridged modules only; Flex address = EIP slot - 1 when slot0 is AENT headnode
            bridged = [
                m for m in sorted(ad.get("modules") or [], key=lambda x: int(x.get("slot") or 0))
                if (m.get("connection") or "").upper() != "HEADNODE"
                and "AENT" not in (m.get("type") or "").upper()
            ]
            # If modules list has no connection flags, skip pure AENT types
            if not bridged:
                bridged = [
                    m for m in sorted(ad.get("modules") or [], key=lambda x: int(x.get("slot") or 0))
                    if "AENT" not in (m.get("type") or "").upper()
                ]
            family = _eip_family_from_types(
                [(m.get("type") or "") for m in (ad.get("modules") or [])]
                + [(m.get("type") or "") for m in bridged]
            )
            for m in bridged:
                eip_slot = int(m.get("slot") or 0)
                # Chassis slot must match prints / Studio POINT addressing:
                #   print "13:1:I.Data" → PointIO Address=1 → parent Data[1]
                # Do NOT subtract 1 (that shifted every card and broke IO vs drawings).
                flex = eip_slot if eip_slot >= 1 else eip_slot
                mt = (m.get("type") or "").strip()
                child_name = f"{rio}_{flex}"
                catalog = EIP_CATALOG.get(mt, mt)
                word = str(m.get("word") or "").strip()
                # Prefer EIPCSV word for this type@rack@bank
                if not word:
                    for w, info in word_to_mod.items():
                        if info.get("type") == mt and info.get("rack") == rack:
                            try:
                                if int(info.get("bank") or -1) == int(m.get("input_bank") or -2):
                                    word = w
                                    break
                            except Exception:
                                pass
                # Direction first — needed to pick the right bank key below
                child_dir = (
                    "I" if any(x in mt for x in ("IA", "IB", "IM")) else (
                        "O" if any(x in mt for x in ("OA", "OB", "OW")) else ""
                    )
                )
                # Reno / empty EIPCSV: Fortna bank on the module IS the address key.
                # Output cards: prefer OutputBank (0 is valid — AENTR3 OB8E).
                # Input cards: prefer InputBank (>0). Never use OB8E status IB as word.
                if not word:
                    ib = m.get("input_bank")
                    ob = m.get("output_bank")
                    try:
                        if child_dir == "O" and ob is not None and int(ob) >= 0:
                            word = str(int(ob))
                        elif child_dir == "I" and ib is not None and int(ib) > 0:
                            word = str(int(ib))
                        elif ib is not None and int(ib) > 0:
                            word = str(int(ib))
                        elif ob is not None and int(ob) >= 0:
                            word = str(int(ob))
                    except Exception:
                        word = str(ib or ob or "").strip()
                child = {
                    "name": child_name,
                    "type": mt,
                    "catalog": catalog,
                    "template": EIP_CHILD_TEMPLATE.get(mt, ""),
                    "eip_slot": eip_slot,
                    "flex_slot": flex,
                    "word": word,
                    "input_bank": m.get("input_bank"),
                    "output_bank": m.get("output_bank"),
                    "family": family,
                    "direction": child_dir,
                }
                children.append(child)
                if word:
                    word_map[word] = {
                        "rio_name": rio,
                        "child_name": child_name,
                        "flex_slot": flex,
                        "type": mt,
                        "catalog": catalog,
                        "rack": rack,
                        "ip": ip,
                        "direction": child["direction"],
                        "family": family,
                    }
                # Also index output cards by OutputBank (including 0)
                try:
                    ob = int(m.get("output_bank"))
                except Exception:
                    ob = -1
                if child["direction"] == "O" and ob >= 0:
                    word_map.setdefault(str(ob), {
                        "rio_name": rio,
                        "child_name": child_name,
                        "flex_slot": flex,
                        "type": mt,
                        "catalog": catalog,
                        "rack": rack,
                        "ip": ip,
                        "direction": "O",
                        "family": family,
                    })
                modules_flat.append(
                    IoModule(
                        name=child_name,
                        type=mt,
                        slot=str(flex),
                        ip="",
                        parent=rio,
                        rack=rack,
                        connection="BRIDGED",
                    )
                )

            # Fix flex slots if EIP used 0-based without headnode (min flex became -1)
            if children and min(c["flex_slot"] for c in children) < 0:
                for c in children:
                    c["flex_slot"] = c["eip_slot"]
                    c["name"] = f"{rio}_{c['flex_slot']}"

            # Also map any EIPCSV words for this rack/type that match child types
            for c in children:
                if c.get("word"):
                    continue
                for w, info in word_to_mod.items():
                    if info.get("rack") == rack and info.get("type") == c["type"] and w not in word_map:
                        c["word"] = w
                        word_map[w] = {
                            "rio_name": rio,
                            "child_name": c["name"],
                            "flex_slot": c["flex_slot"],
                            "type": c["type"],
                            "catalog": c["catalog"],
                            "rack": rack,
                            "ip": ip,
                            "direction": c["direction"],
                        }
                        break

            topology.append({
                "rio_name": rio,
                "adapter_name": ad.get("name") or "",
                "rack": rack,
                "ip": ip,
                "family": family,
                "children": children,
            })
            modules_flat.insert(
                0,
                IoModule(
                    name=rio,
                    type=EIP_PARENT_TYPE.get(family, "1794-AENT"),
                    slot="0",
                    ip=ip,
                    parent="",
                    rack=rack,
                    connection="HEADNODE",
                ),
            )

    # Final word_map pass: map every EIPCSV input word of known type on known rack
    # to the matching child by order of input cards
    for rack, ads in by_rack.items():
        for idx, ad in enumerate(ads):
            rio = f"{rack}RIO{idx}"
            topo = next((t for t in topology if t["rio_name"] == rio), None)
            if not topo:
                continue
            in_children = [c for c in topo["children"] if c["direction"] == "I"]
            # words for this rack that are input modules
            rack_words = [
                (w, info) for w, info in word_to_mod.items()
                if info.get("rack") == rack and info.get("io") == "I"
                and any(x in (info.get("type") or "") for x in ("IA", "IB"))
            ]
            rack_words.sort(key=lambda x: int(x[0]) if str(x[0]).isdigit() else 0)
            # assign by sequence if not already mapped
            used_flex = {word_map[w]["flex_slot"] for w in word_map if word_map[w].get("rio_name") == rio}
            for w, info in rack_words:
                if w in word_map:
                    continue
                mt = info.get("type") or ""
                for c in in_children:
                    if c["type"] == mt and c["flex_slot"] not in used_flex:
                        word_map[w] = {
                            "rio_name": rio,
                            "child_name": c["name"],
                            "flex_slot": c["flex_slot"],
                            "type": c["type"],
                            "catalog": c["catalog"],
                            "rack": rack,
                            "ip": topo["ip"],
                            "direction": "I",
                        }
                        c["word"] = w
                        used_flex.add(c["flex_slot"])
                        break

    result["adapters_raw"] = adapters
    result["topology"] = topology
    result["word_map"] = word_map
    result["modules_flat"] = modules_flat
    return result


def _load_eip_adapters(run_dir: Path) -> tuple[list[dict], str, list[IoModule], list, dict]:
    """Backward-compatible wrapper → topology + word map."""
    topo = load_eip_topology(run_dir)
    return (
        topo.get("adapters_raw") or [],
        topo.get("interface_ip") or "",
        list(topo.get("modules_flat") or []),
        list(topo.get("topology") or []),
        dict(topo.get("word_map") or {}),
    )


def load_from_run(run_dir: Path, *, processor: str = "1756-L83E") -> AutogenInput:
    """
    Build autogen input from a Fortna RUN package (tar.gz already extracted).

    This replaces the manual step: pull tags from prints → type into Excel.
    Conveyor.asc mechanical rows become Inputdata conveyor lines; PE rows attach
    real PE tag names for Fast_Conv/Slow_Jam/PE_Logic; EIP modules from eipcfg.
    """
    from fortna_asc import read_asc
    from fortna_io_extract import (
        equipment_kind,
        extract_io_points,
        normalize_io_name,
        read_project_meta,
    )

    run_dir = Path(run_dir)
    if (run_dir / "RUN" / "project.cfg").is_file():
        run_dir = run_dir / "RUN"
    if not (run_dir / "project.cfg").is_file():
        raise FileNotFoundError(f"No Fortna RUN at {run_dir}")

    meta = read_project_meta(run_dir)
    machine = meta.get("machine_name") or "Machine"
    project = meta.get("project_name") or machine

    conv_path = run_dir / "FORTNA" / "Conveyor.asc"
    if not conv_path.is_file():
        raise FileNotFoundError(f"Missing Conveyor.asc under {run_dir}")

    _, rows = read_asc(conv_path)
    points = extract_io_points(run_dir, include_spares=False)

    # EIP first — word_map scopes devices to this master PLC's RIO network
    eip_adapters, eip_ip, eip_modules, eip_topology, io_word_map = _load_eip_adapters(run_dir)
    # Configio: Fortna Octal_Word → EIP Bank (authoritative when EIPCSV empty)
    configio_octal_map = _load_configio_octal_map(run_dir, machine)

    from fortna_io_extract import belongs_to_controller, row_machine_matches

    # --- Scope I/O + PE to this controller only (Greensboro has 3 masters) ---
    pe_by_conv: dict[str, list[dict]] = {}
    io_points: list[IoPoint] = []
    pe_devices: list[dict] = []
    linked_conveyors: set[str] = set()  # P### owned via PE/IO on this machine

    # VFD tags on this controller only (for conveyor MS vs VFD classification)
    vfd_num_keys: set[str] = set()
    for row in rows:
        rn = (row.get("IO_Name") or "").strip()
        if not re.match(r"^VFD", rn, re.I):
            continue
        if not belongs_to_controller(
            machine_name=(row.get("Machine_Name") or "").strip(),
            io_word=str(row.get("IO_Address_Word") or "").strip(),
            controller=machine,
            word_map=io_word_map,
        ):
            continue
        core = normalize_io_name(rn)
        m = re.match(r"^VFD[\s\-_]*([A-Z0-9]*\d[A-Z0-9]{0,6})", core, re.I)
        if m:
            c = re.sub(r"(_EN|_AUX|_FLT|_RUN|_OK)$", "", m.group(1), flags=re.I).upper()
            if c:
                vfd_num_keys.add(c)
                dm = re.search(r"(\d{2,4})", c)
                if dm:
                    vfd_num_keys.add(dm.group(1))

    # Optional print OCR VFD ids (from last OCR run) → same number matching
    try:
        ocr_path = REPO_ROOT / "workspace" / "ocr-last-result.json"
        if ocr_path.is_file():
            ocr = json.loads(ocr_path.read_text(encoding="utf-8"))
            for p in ocr.get("print_vfd_params") or []:
                val = str(p.get("value") or p.get("device_id") or "")
                m = re.match(r"^VFD[\s\-_]*([A-Z0-9]*\d[A-Z0-9]{0,6})", val, re.I)
                if m:
                    core = m.group(1).upper()
                    vfd_num_keys.add(core)
                    dm = re.search(r"(\d{2,4})", core)
                    if dm:
                        vfd_num_keys.add(dm.group(1))
    except Exception:
        pass

    for p in points:
        if not belongs_to_controller(
            machine_name=str(p.get("machine_name") or ""),
            io_word=str(p.get("fortna_bank") or ""),
            controller=machine,
            word_map=io_word_map,
        ):
            continue
        kind = p.get("equipment_kind") or equipment_kind(
            p.get("fortna_name") or "", p.get("device_type") or "", p.get("description") or ""
        )
        name = p.get("fortna_name") or ""
        if kind == "photoeye":
            link = _link_pe_to_conveyor(p)
            role = _classify_pe_role(name, p.get("description") or "")
            pe_rec = {
                "name": _safe(name),
                "fortna_name": name,
                "conveyor": link,
                "role": role,
                "bank": str(p.get("fortna_bank") or ""),
                "bit": str(p.get("fortna_bit") or ""),
                "description": (p.get("description") or "")[:120],
                "machine_name": str(p.get("machine_name") or ""),
            }
            pe_devices.append(pe_rec)
            if link:
                pe_by_conv.setdefault(link.upper(), []).append(p)
                linked_conveyors.add(link.upper())
            io_points.append(
                IoPoint(
                    device_name=name,
                    device_type="photoeye",
                    source_module=p.get("module") or "",
                    direction="I",
                    fortna_bank=str(p.get("fortna_bank") or ""),
                    fortna_bit=str(p.get("fortna_bit") or ""),
                    conveyor=link,
                    description=(p.get("description") or "")[:120],
                    pe_role=role,
                )
            )
        elif p.get("fortna_bank") and kind not in ("conveyor", "spare", "invalid"):
            # Every bank-addressed field device from tar.gz (PB, ES, motor, beacon,
            # digital_in, scanner, power_supply, …) — not just PE.
            if kind == "vfd":
                dm = re.search(r"(\d{2,4})", name)
                if dm:
                    linked_conveyors.add(f"P{dm.group(1)}")
            io_dir = str(p.get("io_type") or "").upper()
            direction = "O" if io_dir in ("OUT", "O", "OUTPUT") else "I"
            # Encoders are inputs (pulse) — never force to output even if Type=BEACON
            if kind in ("encoder",) or re.match(r"^ENC\d", name, re.I) or re.search(
                r"ENCODER", name + " " + str(p.get("description") or ""), re.I
            ):
                direction = "I"
            # Solenoids / pusher SSV outputs (Type=TRIANG often mislabeled)
            elif kind in ("digital_out", "solenoid") or re.search(r"SSV", name, re.I):
                direction = "O"
            # Beacons / horns are almost always outputs even if Type mislabeled
            elif kind in ("beacon",) or re.search(r"WH\d|HORN|BEACON|LAMP", name, re.I):
                direction = "O"
            io_points.append(
                IoPoint(
                    device_name=name,
                    device_type=kind or str(p.get("device_class") or "io"),
                    source_module=p.get("module") or "",
                    direction=direction,
                    fortna_bank=str(p.get("fortna_bank") or ""),
                    fortna_bit=str(p.get("fortna_bit") or ""),
                    description=(p.get("description") or "")[:120],
                )
            )

    # Conveyors: explicit Machine_Name match OR linked from this controller's PE/VFD
    # (plant-wide ASC often leaves conveyors as Machine_Name=N/A)
    conveyors: list[ConveyorRow] = []
    areas: list[str] = []
    n = 0
    for row in rows:
        typ = (row.get("Type") or "").strip().upper()
        if typ not in CONVEYOR_ASC_TYPES:
            continue
        raw_name = (row.get("IO_Name") or "").strip()
        if not raw_name or raw_name.upper() in ("INVALID", "N/A", "SPARE", "ALWAYSON", "NEVERON"):
            continue
        kind = equipment_kind(raw_name, typ, row.get("General_Description") or "")
        if kind in ("photoeye", "pushbutton", "estop", "beacon", "motor", "vfd"):
            continue
        name = normalize_io_name(raw_name)
        if not re.match(r"^P\d{2,4}[A-Z0-9_]*$", name, re.I):
            continue
        if re.search(r"_AUX$|_FLT$|_OK$|_RUN$", name, re.I):
            continue
        if not name:
            continue

        row_mach = (row.get("Machine_Name") or "").strip()
        name_u = name.upper()
        # Exclude conveyors tagged to a different master
        if row_mach and row_mach.upper() not in ("N/A", "INVALID", "", "NONE", "ALL"):
            if not row_machine_matches(row_mach, machine):
                continue
        else:
            # Untagged: only if PE/VFD on this controller linked to this belt
            # also allow base P### match for P600C-style PE linking to P600
            base_m = re.match(r"^(P\d{2,4})", name_u)
            base = base_m.group(1) if base_m else name_u
            if name_u not in linked_conveyors and base not in linked_conveyors:
                # any linked name starts with this conveyor (P600 ← P600C PE)
                if not any(
                    lc == name_u or lc.startswith(name_u) or name_u.startswith(lc)
                    for lc in linked_conveyors
                ):
                    continue

        desc = (row.get("General_Description") or "").strip()
        drive = (row.get("Drive") or "").strip()
        is_vfd = bool(re.search(r"\bVFD\b", f"{raw_name} {desc} {drive}", re.I))
        if drive in ("0", "1", " ", "N/A", "N", "~", ""):
            pass
        elif any(c.isalpha() for c in drive) or len(drive) > 2:
            is_vfd = True
        if not is_vfd:
            pm = re.match(r"^P(\d{2,4}[A-Z]?)", name, re.I)
            if pm:
                pkey = pm.group(1).upper()
                pdig = re.search(r"(\d{2,4})", pkey)
                if pkey in vfd_num_keys or (pdig and pdig.group(1) in vfd_num_keys):
                    is_vfd = True

        type_map = FORTNA_TYPE_TO_AUTOGEN_VFD if is_vfd else FORTNA_TYPE_TO_AUTOGEN
        ag_type = type_map.get(typ, "Transport with MS")

        # One L5X per controller → one area named for the machine (MSCRENOPACK_Area).
        # Do NOT invent Zone1–Zone9 from P-number prefixes.
        area = f"{_safe(machine)}_Area"
        if area not in areas:
            areas.append(area)

        pe = _pe_wiring_for_conv(
            pe_by_conv.get(name_u, []) or pe_by_conv.get(name, [])
        )
        n += 1
        conveyors.append(
            ConveyorRow(
                number=n,
                system=machine,
                main_area=area,
                safety_zone=f"{_safe(machine)}_ESZone1",
                conveyor=name,
                type=ag_type,
                downstream="",  # topology not reliable in ASC
                exit_pe=pe["exit_opt"],
                full=pe["full_opt"],
                jam=pe["jam_opt"] or pe["exit_opt"],
                motor_starter="" if is_vfd else "Yes",
                exit_pe_tag=pe["exit_pe_tag"],
                add_pe_tag=pe["add_pe_tag"],
                jam_pe_tags=pe["jam_pe_tags"],
                full_pe_tags=pe["full_pe_tags"],
                product_pe_tags=pe["product_pe_tags"],
                all_pe_tags=pe["all_pe_tags"],
            )
        )

    # Dedupe by conveyor name (ASC sometimes lists the same P### twice → duplicate AOI refs)
    seen_conv: set[str] = set()
    deduped: list[ConveyorRow] = []
    for c in sorted(conveyors, key=lambda x: x.conveyor):
        key = (c.conveyor or "").upper()
        if not key or key in seen_conv:
            continue
        seen_conv.add(key)
        deduped.append(c)
    conveyors = deduped
    for i, c in enumerate(conveyors, start=1):
        c.number = i

    equipment_plan: dict = {}
    try:
        from fortna_equipment_plan import inventory_and_plan

        equipment_plan = inventory_and_plan(run_dir, machine_name=machine)
        plan = equipment_plan.get("plan") or {}
        _emit_progress(
            f"Equipment[{machine}]: {equipment_plan.get('inventory', {}).get('counts', {})} "
            f"profile={plan.get('profile')} packs={plan.get('packs')}",
            18,
        )
    except Exception as ex:
        equipment_plan = {"error": str(ex)}

    return AutogenInput(
        project_name=f"{project}_{machine}",
        processor=processor,
        major_rev="35",
        minor_rev="00",
        areas=areas,
        safety_zones=[f"{a.replace('_Area', '')}_ESZone1" for a in areas],
        conveyors=conveyors,
        modules=eip_modules,
        io_points=io_points,
        eip_adapters=eip_adapters,
        eip_interface_ip=eip_ip,
        eip_topology=eip_topology,
        io_word_map=io_word_map,
        configio_octal_map=configio_octal_map,
        pe_devices=pe_devices,
        equipment_plan=equipment_plan,
    )


# ---------------------------------------------------------------------------
# Library template resolution
# ---------------------------------------------------------------------------

def resolve_template(conv_type: str, library_text: str) -> str:
    key = (conv_type or "").strip().lower()
    preferred = EXCEL_CONFIG_OVERRIDES.get(key) or TYPE_TO_TEMPLATE.get(key) or "P1000_Conv"
    fallbacks = [
        preferred,
        TYPE_TO_TEMPLATE.get(key, "P1000_Conv"),
        "P1000_Conv",
        "P3000_Conv",
        "P2000_Conv",
        "P4000_Conv",
    ]
    for t in fallbacks:
        if t and f'Tag Name="{t}"' in library_text:
            return t
    return "P1000_Conv"


def extract_tag_block(library_text: str, tag_name: str) -> str | None:
    """Extract full <Tag Name="...">...</Tag> block."""
    pat = rf'<Tag Name="{re.escape(tag_name)}"[^>]*>.*?</Tag>'
    m = re.search(pat, library_text, re.S)
    return m.group(0) if m else None


def extract_routine_block(library_text: str, routine_name: str) -> str | None:
    pat = rf'<Routine Name="{re.escape(routine_name)}"[^>]*>.*?</Routine>'
    m = re.search(pat, library_text, re.S)
    return m.group(0) if m else None


def _pad_pe_slots(tags: list[str], n: int = 5) -> list[str]:
    """Pad PE tag list to n slots with NO_PE (library PE_UDT null object)."""
    out = [t for t in (tags or []) if t][:n]
    while len(out) < n:
        out.append("NO_PE")
    return out


def clone_template_for_conveyor(
    library_text: str,
    template: str,
    conveyor: str,
    area: str,
    safety_zone: str,
    downstream: str,
    *,
    is_vfd: bool = False,
    exit_pe_tag: str = "",
    add_pe_tag: str = "",
    jam_pe_tags: list | None = None,
    full_pe_tags: list | None = None,
    product_pe_tags: list | None = None,
) -> dict:
    """
    Clone library template tags + Excel-style AOI rungs for one conveyor.

    Matches edited gold L5X patterns:
      Fast_Conv(…, Next_or_NO_Conv, ExitPE_or_NO_PE, AddPE_or_NO_PE, …)
      Slow_Jam(…, PE1..PE5 with NO_PE padding)
      PE_Logic / Full_PE per eye
      Slow_Flt for motor/VFD fault
    """
    conv = _safe(conveyor)
    area_s = _safe(area) or "Main_Area"
    safe_s = _safe(safety_zone) or f"{area_s}_Safe"
    if not (safe_s.endswith("_Safe") or "ESZone" in safe_s):
        safe_s = f"{safe_s}_Safe" if safe_s else f"{area_s}_Safe"

    # Next conveyor: real Pxxx_Conv or NO_Conv (never fake Next_Conv tag)
    next_tag = "NO_Conv"
    if downstream:
        dn = _safe(downstream)
        if dn and dn not in ("Next_Conv", "NO_Conv"):
            next_tag = f"{dn}_Conv" if not dn.endswith("_Conv") else dn

    base = template  # P1000_Conv / P3000_Conv / …
    aoi = f"{base}_AOI"
    new_base = f"{conv}_Conv"
    new_aoi = f"{conv}_Conv_AOI"
    # Gold finished PLC5 (ORLY … PLC5Finished): tags named P###_VFD are
    # Motor_Starter_UDT (aux / contactor I/O), NOT Ethernet VFD_UDT.
    # Slow_Flt(IO_VFD=NO_VFD, IO_MS=P###_VFD) and Fast_Conv uses NO_VFD.
    # True PowerFlex program params live in print OCR / docs, not this UDT.
    vfd_tag = "NO_VFD"
    ms_tag = f"{conv}_VFD" if is_vfd else f"{conv}_MS"

    exit_pe = _safe(exit_pe_tag) if exit_pe_tag else "NO_PE"
    add_pe = _safe(add_pe_tag) if add_pe_tag else "NO_PE"
    jam_slots = _pad_pe_slots([_safe(t) for t in (jam_pe_tags or []) if t], 5)

    replacements = [
        (base, new_base),
        (aoi, new_aoi),
        ("Conv_Area_Safe", safe_s),
        ("Conv_Area", area_s),
        ("Main_Area", area_s),
    ]

    tags = []
    # Conv_UDT + Conv_AOI backing tag + optional brake tags (same as Excel autogen)
    for name in (base, aoi, f"{base}_BrakeReleased", f"{base}_BrakeReleased_Aux"):
        block = extract_tag_block(library_text, name)
        if not block:
            continue
        cloned = block
        for old, new in sorted(replacements, key=lambda x: -len(x[0])):
            cloned = cloned.replace(old, new)
        tags.append(cloned)

    # Motor_Starter_UDT for Slow_Flt IO_MS — MS conveyors use P###_MS; VFD-fed
    # discrete I/O conveyors use P###_VFD (same UDT, gold naming).
    ms_src = extract_tag_block(library_text, "NO_MS")
    if ms_src:
        tags.append(ms_src.replace("NO_MS", ms_tag))

    # Real PE_UDT + PE_Logic / Full_PE AOI backing tags (cloned from library templates)
    pe_udt_src = (
        extract_tag_block(library_text, "PE1000_P")
        or extract_tag_block(library_text, "PEExit_P")
        or extract_tag_block(library_text, "NO_PE")
    )
    pe_logic_src = extract_tag_block(library_text, "PE1000_P_AOI")
    full_pe_src = extract_tag_block(library_text, "PE1000_F") or pe_udt_src
    full_aoi_src = extract_tag_block(library_text, "PE1000_F_AOI")

    pe_tags_needed = list(
        dict.fromkeys(
            [t for t in jam_slots if t != "NO_PE"]
            + [exit_pe, add_pe]
            + [_safe(t) for t in (full_pe_tags or []) if t]
            + [_safe(t) for t in (product_pe_tags or []) if t]
        )
    )
    pe_tags_needed = [t for t in pe_tags_needed if t and t != "NO_PE"]

    for pe_name in pe_tags_needed:
        role_full = pe_name in {_safe(t) for t in (full_pe_tags or [])} or re.search(
            r"_F\d*$", pe_name, re.I
        )
        if role_full and full_pe_src:
            tags.append(full_pe_src.replace("PE1000_F", pe_name).replace("PEExit_P", pe_name).replace("NO_PE", pe_name))
            if full_aoi_src:
                tags.append(
                    full_aoi_src.replace("PE1000_F_AOI", f"{pe_name}_AOI").replace(
                        "PE1000_F", pe_name
                    )
                )
        else:
            if pe_udt_src:
                tags.append(
                    pe_udt_src.replace("PE1000_P", pe_name)
                    .replace("PEExit_P", pe_name)
                    .replace("NO_PE", pe_name)
                )
            if pe_logic_src:
                tags.append(
                    pe_logic_src.replace("PE1000_P_AOI", f"{pe_name}_AOI").replace(
                        "PE1000_P", pe_name
                    )
                )

    # --- Rungs (Excel / edited gold mnemonics) ---
    fast_text = (
        f"Fast_Conv({new_aoi}.Fast,{new_base},{area_s},{safe_s},{next_tag},"
        f"{exit_pe},{add_pe},{new_base}.Type,{vfd_tag},0,1,0,0,HMIColor,HMI_StatsClear);"
    )
    jam_text = (
        f"Slow_Jam({new_aoi}.Jam,{new_base},{area_s},"
        f"{','.join(jam_slots)});"
    )
    # Slow_Flt — motor/VFD fault (Type2 = standard MS type code from library)
    flt_text = (
        f"Slow_Flt({new_aoi}.Flt,{new_base},{area_s},{vfd_tag},NO_Enc,Type2,"
        f"{ms_tag},NO_PS,NO_AirPress,NO_AdditionalFlt,{area_s}.MtrFlt_Reset);"
    )

    # Excel Autogen rung comments (tilde banner — clean in Studio ladder view)
    # "with VFD" = discrete VFD feeder (P###_VFD Motor_Starter_UDT), not Ethernet VFD_UDT
    conv_kind = "Accumulation Conv with VFD" if is_vfd else "Transport Conv with MS"
    if not is_vfd and any(
        x in (template or "").upper() for x in ("P3000", "P4000", "ACCUM", "ZERO")
    ):
        conv_kind = "Accumulation Conv with MS"
    elif is_vfd and any(x in (template or "").upper() for x in ("P3000", "P4000")):
        conv_kind = "Accumulation Conv with VFD"

    def _excel_comment(*lines: str) -> str:
        body = "\n".join(lines)
        return f"~~~~~~~~~~~\n{body}\n~~~~~~~~~~~"

    rungs = [
        {
            "label": "Fast",
            "text": fast_text,
            "comment": _excel_comment(f"{new_base} Fast logic", f"({conv_kind})"),
        },
        {
            "label": "Jam",
            "text": jam_text,
            "comment": _excel_comment(
                f"{new_base} Jam logic",
                "Jam Sum logic & Jam Silence logic",
            ),
        },
        {
            "label": "Flt",
            "text": flt_text,
            "comment": _excel_comment(
                f"{new_base} Motor/VFD Fault logic",
                "Standard Logic",
            ),
        },
    ]

    # Per-eye PE_Logic / Full_PE rungs
    for pe_name in pe_tags_needed:
        is_full = pe_name in {_safe(t) for t in (full_pe_tags or [])} or bool(
            re.search(r"_F\d*$", pe_name, re.I)
        )
        aoi_tag = f"{pe_name}_AOI"
        if is_full:
            # Gold ModuleB_Area_Fast has separate Conv_Full routine
            rungs.append(
                {
                    "label": "Full",
                    "text": f"Full_PE({aoi_tag},{pe_name},{new_base},1,0,HMI_StatsClear);",
                    "comment": _excel_comment(
                        f"{pe_name} Full PELogic",
                        "Note: Some of the AOI values are coming from the Full_Init Routine from Slow Task",
                        "Standard Logic",
                    ),
                }
            )
        else:
            rungs.append(
                {
                    "label": "PE",
                    "text": (
                        f"PE_Logic({aoi_tag},{pe_name},{new_base},"
                        f"{area_s}.PI.JamRst,0,0,0,{aoi_tag}.I_AutoClearTmr,HMI_StatsClear);"
                    ),
                    "comment": _excel_comment(
                        f"{pe_name} Debounce and Jam Logic",
                        "Standard Logic -> Jam Reset only when PE is Clear",
                    ),
                }
            )

    if not pe_tags_needed:
        rungs.append(
            {
                "label": "PE",
                "text": "NOP();",
                "comment": f"{conv} PE",
            }
        )

    return {
        "conveyor": conv,
        "template": template,
        "area": area_s,
        "safety_zone": safe_s,
        "is_vfd": is_vfd,
        "exit_pe": exit_pe,
        "jam_pes": jam_slots,
        "pe_tags": pe_tags_needed,
        "ms_tag": ms_tag,
        "tags": tags,
        "rungs": rungs,
        "tag_names": [new_base, new_aoi] + pe_tags_needed,
    }


# ---------------------------------------------------------------------------
# L5X assembly
# ---------------------------------------------------------------------------

def _load_gold_plc2_text() -> str:
    """Prefer Compare-programs / Desktop gold PLC2 for CommDiag tag/UDT templates."""
    candidates = [
        Path(r"C:\Users\curtiskricke\OneDrive - Fortna Inc\Documents\Studio 5000\Projects\Compare programs\ORLY_Greensboro_NC_PLC2.L5X"),
        Path(r"C:\Users\curtiskricke\Desktop\ORielly Green\1 PLC2\ORLY_Greensboro_NC_PLC2s.L5X"),
        Path(r"C:\Users\curtiskricke\Desktop\ORielly Green\1 PLC2\ORLY_Greensboro_NC_PLC2.L5X"),
    ]
    for p in candidates:
        try:
            if p.is_file():
                return p.read_text(encoding="utf-8", errors="ignore")
        except Exception:
            continue
    return ""


def _build_sys_comm_program_xml(
    *,
    inp: "AutogenInput",
    library_text: str,
    _add_tag_block,
    _rung_xml,
    routine,
    seen_tag_names: set,
) -> str:
    """Gold PLC2 `System` program — Devices_Comm_Logic from this site's tar.

    Tag types must match AOI_CommDiag (see Greensboro PLC2), not DINT/BOOL stubs.
    """
    gold = _load_gold_plc2_text() or library_text
    proj = _safe(inp.project_name) or "Site"
    reset_udt = f"{proj}_System"

    devices: list[tuple[str, str | None]] = []
    for node in getattr(inp, "eip_topology", None) or []:
        rio = (node.get("rio_name") or node.get("name") or "").strip()
        if not rio:
            continue
        devices.append((rio, None))
        for ch in node.get("children") or []:
            slot = ch.get("flex_slot")
            child_name = (ch.get("name") or "").strip()
            if not child_name and slot is not None:
                child_name = f"{rio}_{slot}"
            if child_name:
                devices.append((child_name, rio))
    if not devices:
        for m in getattr(inp, "modules", None) or []:
            name = (getattr(m, "name", None) or (m.get("name") if isinstance(m, dict) else "") or "").strip()
            if name:
                devices.append((name, None))

    def _clone_tag(src_name: str, new_name: str | None = None) -> bool:
        """Clone a controller tag block from gold/library under a new name."""
        dest = new_name or src_name
        if dest in seen_tag_names:
            return True
        block = extract_tag_block(gold, src_name) or extract_tag_block(library_text, src_name)
        if not block:
            return False
        if new_name and new_name != src_name:
            block = re.sub(r'Tag Name="[^"]+"', f'Tag Name="{_xml_escape(new_name)}"', block, count=1)
        _add_tag_block(block)
        return True

    def _ensure_typed(name: str, datatype: str, *, dims: str = "", l5k: str = "0") -> None:
        if name in seen_tag_names:
            return
        dim_attr = f' Dimensions="{dims}"' if dims else ""
        if datatype == "BOOL":
            decorated = '<DataValue DataType="BOOL" Value="0"/>'
        elif datatype == "SINT" and dims:
            zeros = ",".join(["0"] * int(str(dims).split()[0] or "20"))
            l5k = f"[{zeros}]"
            decorated = ""  # array decorated optional
        else:
            decorated = f'<DataValue DataType="{datatype}" Radix="Decimal" Value="0"/>'
        deco = f'<Data Format="Decorated">{decorated}</Data>' if decorated else ""
        _add_tag_block(
            f'<Tag Name="{_xml_escape(name)}" TagType="Base" DataType="{datatype}"{dim_attr} '
            f'Radix="Decimal" Constant="false" ExternalAccess="Read/Write">'
            f'<Data Format="L5K"><![CDATA[{l5k}]]></Data>{deco}</Tag>'
        )

    def _ensure_comm_udt(name: str) -> None:
        """Clone Comm_UDT Data matching *library* Comm_UDT (includes CommLoss_Tmr).

        Gold NO_CommDev Data omits CommLoss_Tmr → Studio 'Data type mismatch'.
        Prefer library templates CP2N6_RIO / CP2N7_RIO.
        """
        if name in seen_tag_names:
            return
        for cand in ("CP2N6_RIO", "CP2N7_RIO", "NO_CommDev"):
            # Only accept clone if source is in library (has CommLoss_Tmr) when possible
            src = extract_tag_block(library_text, cand) or (
                extract_tag_block(gold, cand) if cand == "NO_CommDev" else None
            )
            if not src:
                continue
            # Library templates include CommLoss_Tmr — required for lib Comm_UDT
            if cand != "NO_CommDev" or "CommLoss_Tmr" in src:
                block = re.sub(r'Tag Name="[^"]+"', f'Tag Name="{_xml_escape(name)}"', src, count=1)
                # Zero out site-specific MAC/IP from template
                block = re.sub(
                    r"(<DataValueMember Name=\"LEN\"[^>]*Value=\")\d+(\")",
                    r"\g<1>0\2",
                    block,
                )
                _add_tag_block(block)
                return
        # Explicit library-shaped Decorated Data (CommLoss_Tmr + Flt + …)
        _add_tag_block(
            f'<Tag Name="{_xml_escape(name)}" TagType="Base" DataType="Comm_UDT" '
            f'Constant="false" ExternalAccess="Read/Write">'
            f"<Data Format=\"L5K\"><![CDATA[[[0,0,0],[0],0,0.00000000e+000,"
            f"[0,'$00$00$00$00$00$00$00$00$00$00$00$00$00$00$00$00$00$00$00$00'],"
            f"[0,'$00$00$00$00$00$00$00$00$00$00$00$00$00$00$00']]]]></Data>"
            f'<Data Format="Decorated"><Structure DataType="Comm_UDT">'
            f'<StructureMember Name="CommLoss_Tmr" DataType="TIMER">'
            f'<DataValueMember Name="PRE" DataType="DINT" Radix="Decimal" Value="0"/>'
            f'<DataValueMember Name="ACC" DataType="DINT" Radix="Decimal" Value="0"/>'
            f'<DataValueMember Name="EN" DataType="BOOL" Value="0"/>'
            f'<DataValueMember Name="TT" DataType="BOOL" Value="0"/>'
            f'<DataValueMember Name="DN" DataType="BOOL" Value="0"/>'
            f"</StructureMember>"
            f'<StructureMember Name="Flt" DataType="Comm_Flt">'
            f'<DataValueMember Name="CommLoss" DataType="BOOL" Value="0"/>'
            f'<DataValueMember Name="UpStrmCommLoss" DataType="BOOL" Value="0"/>'
            f"</StructureMember>"
            f'<DataValueMember Name="Comm_Code" DataType="DINT" Radix="Decimal" Value="0"/>'
            f'<DataValueMember Name="Firmware" DataType="REAL" Radix="Float" Value="0.0"/>'
            f'<StructureMember Name="MACId" DataType="String_20">'
            f'<DataValueMember Name="LEN" DataType="DINT" Radix="Decimal" Value="0"/>'
            f'<DataValueMember Name="DATA" DataType="String_20" Radix="ASCII"><![CDATA[]]></DataValueMember>'
            f"</StructureMember>"
            f'<StructureMember Name="IP_Address" DataType="String_15">'
            f'<DataValueMember Name="LEN" DataType="DINT" Radix="Decimal" Value="0"/>'
            f'<DataValueMember Name="DATA" DataType="String_15" Radix="ASCII"><![CDATA[]]></DataValueMember>'
            f"</StructureMember>"
            f"</Structure></Data></Tag>"
        )

    def _ensure_aoi_comm(name: str) -> None:
        if name in seen_tag_names:
            return
        for cand in ("PLC2_ENET1_CommLoss_AOI", "CP2RIO0_CommLoss_AOI", "CP2RIO0_0_CommLoss_AOI"):
            if _clone_tag(cand, name):
                return
        # Last resort: skip creating broken empty AOI tag — rung will NOP if missing
        return

    def _ensure_commdiag_group(g: int, index_max: int) -> None:
        name = f"CommsDiag_Group{g}"
        if name in seen_tag_names:
            return
        if _clone_tag("CommsDiag_Group1", name):
            return
        _add_tag_block(
            f'<Tag Name="{_xml_escape(name)}" TagType="Base" DataType="CommDiag_UDT" '
            f'Constant="false" ExternalAccess="Read/Write">'
            f'<Data Format="L5K"><![CDATA[[0,{int(index_max)},0]]]></Data>'
            f'<Data Format="Decorated"><Structure DataType="CommDiag_UDT">'
            f'<DataValueMember Name="Index" DataType="SINT" Radix="Decimal" Value="0"/>'
            f'<DataValueMember Name="IndexMax" DataType="SINT" Radix="Decimal" Value="{int(index_max)}"/>'
            f'<DataValueMember Name="Init" DataType="BOOL" Value="0"/>'
            f"</Structure></Data></Tag>"
        )

    # Shared infra — exact gold types
    _ensure_typed("DeviceInfo_Read", "BOOL")
    if not _clone_tag("MACID_Bytes"):
        _ensure_typed("MACID_Bytes", "SINT", dims="20")
    if not _clone_tag("Firmware_Bytes"):
        _ensure_typed("Firmware_Bytes", "SINT", dims="20")
    if not _clone_tag("GET_Firmware"):
        _add_tag_block(
            '<Tag Name="GET_Firmware" TagType="Base" DataType="MESSAGE" ExternalAccess="Read/Write">'
            "<Data Format=\"Message\"><MessageParameters MessageType=\"CIP Generic\" RequestedLength=\"0\" "
            "ConnectedFlag=\"2\" CommTypeCode=\"0\" ServiceCode=\"16#000e\" ObjectType=\"16#0001\" "
            "TargetObject=\"1\" AttributeNumber=\"16#0004\" LocalIndex=\"0\" "
            "DestinationTag=\"Firmware_Bytes[0]\" LargePacketUsage=\"false\"/></Data></Tag>"
        )
    if not _clone_tag("GET_MACID"):
        _add_tag_block(
            '<Tag Name="GET_MACID" TagType="Base" DataType="MESSAGE" ExternalAccess="Read/Write">'
            "<Data Format=\"Message\"><MessageParameters MessageType=\"CIP Generic\" RequestedLength=\"0\" "
            "ConnectedFlag=\"2\" CommTypeCode=\"0\" ServiceCode=\"16#000e\" ObjectType=\"16#00f6\" "
            "TargetObject=\"1\" AttributeNumber=\"16#0003\" LocalIndex=\"0\" "
            "DestinationTag=\"MACID_Bytes[0]\" LargePacketUsage=\"false\"/></Data></Tag>"
        )
    _ensure_comm_udt("NO_CommDev")
    if reset_udt not in seen_tag_names:
        if not _clone_tag("ORLY_Greensboro_NC_PLC2_System", reset_udt):
            # System_UDT with Reset bit
            _add_tag_block(
                f'<Tag Name="{_xml_escape(reset_udt)}" TagType="Base" DataType="System_UDT" '
                f'Constant="false" ExternalAccess="Read/Write">'
                f'<Data Format="Decorated"><Structure DataType="System_UDT"/></Data></Tag>'
            )

    n_groups = max(1, (len(devices) + 9) // 10) if devices else 1
    for g in range(1, n_groups + 1):
        imax = 10 if g < n_groups else max(1, len(devices) - (g - 1) * 10)
        _ensure_commdiag_group(g, imax)

    has_aoi = "AOI_CommDiag" in library_text or "AOI_CommDiag" in gold
    has_commdiag_udt = "CommDiag_UDT" in library_text or "CommDiag_UDT" in gold
    rungs: list[str] = []
    rungs.append(
        _rung_xml(
            0,
            "NOP();",
            f"#############################\n{proj} / RIO Devices (System)\n"
            f"Gold: Devices_Comm_Logic under P11_Slow_200ms\n"
            f"Devices from this site tar: {len(devices)}\n######################################",
        )
    )
    for g in range(1, n_groups + 1):
        imax = 10 if g < n_groups else max(1, len(devices) - (g - 1) * 10)
        rungs.append(
            _rung_xml(
                len(rungs),
                f"[EQU(CommsDiag_Group{g}.Index,CommsDiag_Group{g}.IndexMax) ,"
                f"XIC(S:FS) MOV({imax},CommsDiag_Group{g}.IndexMax) ]"
                f"CLR(CommsDiag_Group{g}.Index)NOP();",
                f"Clear Count group {g}",
            )
        )
    rungs.append(_rung_xml(len(rungs), "NOP()OTU(DeviceInfo_Read);", "Pulse DeviceInfo_Read"))

    # If types missing, emit NOP list instead of 600 Studio errors
    safe_emit = has_aoi and has_commdiag_udt and ("Comm_UDT" in library_text or "Comm_UDT" in gold)

    for i, (dev, parent) in enumerate(devices):
        g = (i // 10) + 1
        dsafe = _safe(dev)
        aoi_tag = f"{dsafe}_CommLoss_AOI"
        _ensure_comm_udt(dsafe)
        parent_arg = _safe(parent) if parent else "NO_CommDev"
        if parent_arg != "NO_CommDev":
            _ensure_comm_udt(parent_arg)
        _ensure_aoi_comm(aoi_tag)
        info_read = "DeviceInfo_Read" if parent is None else "0"
        reset_arg = f"{reset_udt}.Reset"
        # Only emit live AOI call when backing AOI tag was successfully cloned
        if safe_emit and aoi_tag in seen_tag_names and dsafe in seen_tag_names:
            text = (
                f"AOI_CommDiag({aoi_tag},{parent_arg},{dsafe},{dsafe},"
                f"GET_Firmware,GET_MACID,MACID_Bytes,Firmware_Bytes,{info_read},"
                f"{reset_arg},CommsDiag_Group{g}.Index);"
            )
            comment = f"CommDiag {dsafe}"
        else:
            text = "NOP();"
            comment = f"TODO CommDiag {dsafe} — clone AOI_CommDiag instance from gold library"
        rungs.append(_rung_xml(len(rungs), text, comment))

    main = [
        _rung_xml(0, "JSR(Devices_Comm_Logic,0);", "Devices_Comm_Logic"),
        _rung_xml(1, "JSR(NTP,0);", "NTP scaffold"),
        _rung_xml(2, "JSR(System_Logic,0);", "System_Logic scaffold"),
    ]
    return (
        f'<Program Name="System" TestEdits="false" MainRoutineName="Main_Routine" '
        f'Disabled="false" UseAsFolder="false">'
        f"<Tags/>"
        f"<Routines>"
        f'{routine("Main_Routine", main)}'
        f'{routine("Devices_Comm_Logic", rungs)}'
        f'{routine("NTP", [_rung_xml(0, "NOP();", "NTP / SNTP — enable when AOI_SNTP_QUERY wired")])}'
        f'{routine("System_Logic", [_rung_xml(0, "NOP();", "System_Logic — site customize")])}'
        f"</Routines></Program>"
    )


# Preferred short AOI titles (Studio Description field — one line only).
_AOI_SHORT_DESC: dict[str, str] = {
    "Fast_Conv": "Conv Fast Routine Logic AOI",
    "Slow_Jam": "Conveyor Jam Logic AOI",
    "Slow_Flt": "Conveyor Fault Logic AOI",
    "PE_Logic": "PE Clear and Jam Logic AOI",
    "Full_PE": "Full PE Logic AOI",
    "Full_PE_Init": "Full PE Init AOI",
    "Fast_Pulse": "Pulse Logic AOI",
    "Fast_TimeStamp": "Date Time Stamp AOI",
    "AB_VFD35": "PowerFlex 35 VFD Logic AOI",
    "AB_VFD525": "PowerFlex 525 VFD Logic AOI",
    "AB_VFD750": "PowerFlex 750 VFD Logic AOI",
    "ES_PI10": "E-Stop PI 10 Logic AOI",
    "ES_PI20": "E-Stop PI 20 Logic AOI",
    "ES_SIL1_Cat1": "E-Stop SIL1 Cat1 Logic AOI",
    "Enc_CounterCard": "Encoder Counter Card Logic AOI",
    "Enc_RIOCard": "Encoder RIO Card Logic AOI",
    "Enc_Virtual_DistBased": "Virtual Encoder Distance AOI",
    "Gapper_Basic": "Gapper Basic Logic AOI",
    "Merge_2to1": "2:1 Merge Logic AOI",
    "Merge_2to1_RAT": "2:1 RAT Merge Logic AOI",
    "AOI_CommDiag": "Comm Loss Diagnostic AOI",
    "AOI_TIME_ADD": "DateTime Add Time AOI",
    "AOI_TIME_DIFFERENCE": "DateTime Difference AOI",
    "AOI_SNTP_QUERY": "SNTP Clock Sync AOI",
}


def _aoi_short_description(name: str, current: str = "") -> str:
    """One-line AOI purpose for Studio. Prefer curated map, else clean/fallback."""
    if name in _AOI_SHORT_DESC:
        return _AOI_SHORT_DESC[name]
    cur = (current or "").strip()
    # Drop copyright / parameter-list essays
    junk = (
        "copyright" in cur.lower()
        or "unauthorized" in cur.lower()
        or "all rights reserved" in cur.lower()
        or cur.count(",") >= 3
        or len(cur) > 72
        or "\n" in cur
    )
    if cur and not junk:
        # Keep a short existing title (strip trailing punctuation noise)
        first = re.split(r"[\r\n:]", cur, maxsplit=1)[0].strip()
        if 3 <= len(first) <= 72 and "copyright" not in first.lower():
            if not first.upper().endswith("AOI"):
                return f"{first} AOI" if not first.endswith("AOI") else first
            return first
    # Name → readable title
    pretty = name.replace("_", " ").strip()
    if not pretty.upper().endswith("AOI"):
        pretty = f"{pretty} AOI"
    return pretty


def _shorten_aoi_descriptions(aoi_xml: str) -> str:
    """
    Rewrite unsealed AOI <Description> text for Studio.

    CRITICAL: never touch <EncodedData>…</EncodedData> sealed bodies.
    Mutating sealed AOI XML (even Description CDATA) invalidates SignatureID
    and produces Studio: "Invalid signature ID. Reseal instruction to resolve."
    """
    if not aoi_xml:
        return aoi_xml

    def _repl_desc(m: re.Match) -> str:
        open_tag = m.group(1)
        name = m.group(2)
        old = m.group(3) or ""
        short = _aoi_short_description(name, old)
        return f"{open_tag}{short}{m.group(4)}"

    # Plain (unsealed) AOI definitions only — sealed EncodedData left byte-for-byte
    aoi_xml = re.sub(
        r'(<AddOnInstructionDefinition\b[^>]*\bName="([^"]+)"[^>]*>\s*'
        r'<Description>\s*<!\[CDATA\[)(.*?)(\]\]>\s*</Description>)',
        _repl_desc,
        aoi_xml,
        flags=re.S,
    )
    return aoi_xml



def _filter_aois_to_used(aoi_xml: str, keep: set[str]) -> str:
    """Keep only named AOI defs (sealed EncodedData or plaintext) that are used."""
    if not keep:
        return "<AddOnInstructionDefinitions/>"
    parts: list[str] = []
    for m in re.finditer(
        r'<EncodedData\s+EncodedType="AddOnInstructionDefinition"\s+Name="([^"]+)"[^>]*>.*?</EncodedData>',
        aoi_xml,
        re.S,
    ):
        if m.group(1) in keep:
            parts.append(m.group(0))
    for m in re.finditer(
        r'<AddOnInstructionDefinition\s+Name="([^"]+)"[^>]*>.*?</AddOnInstructionDefinition>',
        aoi_xml,
        re.S,
    ):
        if m.group(1) in keep:
            parts.append(m.group(0))
    if not parts:
        return aoi_xml  # safety: don't wipe library if keep-set mismatched
    return "<AddOnInstructionDefinitions>\n" + "\n".join(parts) + "\n</AddOnInstructionDefinitions>"


def _overlay_aoi_exports(aoi_xml: str, overlay_dir: Path | None = None) -> str:
    """Replace AOI defs with matching *_AOI.L5X exports (source-key re-seals).

    Example: tools/libraries/Slow_Flt_AOI.L5X replaces library Slow_Flt EncodedData
    so Studio signature matches the key used to export it.
    """
    odir = Path(overlay_dir) if overlay_dir else AOI_OVERLAY_DIR
    if not odir.is_dir():
        return aoi_xml
    replaced: list[str] = []
    for path in sorted(odir.glob("*_AOI.L5X")):
        try:
            text = path.read_text(encoding="utf-8", errors="ignore")
        except Exception:
            continue
        # Prefer EncodedData (typical sealed export), else plaintext definition
        blocks = re.findall(
            r'<EncodedData\s+EncodedType="AddOnInstructionDefinition"\s+Name="([^"]+)"[^>]*>.*?</EncodedData>',
            text,
            re.S,
        )
        # re.findall with one group returns names only — re-extract full blocks
        for m in re.finditer(
            r'<EncodedData\s+EncodedType="AddOnInstructionDefinition"\s+Name="([^"]+)"[^>]*>.*?</EncodedData>',
            text,
            re.S,
        ):
            name, block = m.group(1), m.group(0)
            pat = re.compile(
                rf'<EncodedData\s+EncodedType="AddOnInstructionDefinition"\s+Name="{re.escape(name)}"[^>]*>.*?</EncodedData>',
                re.S,
            )
            if pat.search(aoi_xml):
                aoi_xml = pat.sub(block, aoi_xml, count=1)
                replaced.append(name)
            else:
                # Insert before closing wrapper
                if aoi_xml.rstrip().endswith("</AddOnInstructionDefinitions>"):
                    aoi_xml = (
                        aoi_xml.rstrip()[: -len("</AddOnInstructionDefinitions>")]
                        + "\n"
                        + block
                        + "\n</AddOnInstructionDefinitions>"
                    )
                    replaced.append(f"{name}+")
        for m in re.finditer(
            r'<AddOnInstructionDefinition\s+Name="([^"]+)"[^>]*>.*?</AddOnInstructionDefinition>',
            text,
            re.S,
        ):
            name, block = m.group(1), m.group(0)
            # Only overlay plaintext if no EncodedData of that name remains
            if re.search(
                rf'EncodedData[^>]*Name="{re.escape(name)}"', aoi_xml
            ):
                continue
            pat = re.compile(
                rf'<AddOnInstructionDefinition\s+Name="{re.escape(name)}"[^>]*>.*?</AddOnInstructionDefinition>',
                re.S,
            )
            if pat.search(aoi_xml):
                aoi_xml = pat.sub(block, aoi_xml, count=1)
                replaced.append(name)
    if replaced:
        _emit_progress(f"AOI overlay: {', '.join(replaced)}", 86)
    return aoi_xml



def build_l5x(inp: AutogenInput, library_path: Path) -> tuple[str, dict]:
    library_text = library_path.read_text(encoding="utf-8", errors="replace")

    # Pull DataTypes + AOI definitions from library wholesale (allow attributes)
    datatypes = re.search(r"<DataTypes\b[^>]*>.*?</DataTypes>", library_text, re.S)
    aois = re.search(
        r"<AddOnInstructionDefinitions\b[^>]*>.*?</AddOnInstructionDefinitions>",
        library_text,
        re.S,
    )

    stamp = datetime.now(timezone.utc).strftime("%a %b %d %H:%M:%S %Y")
    proj = _safe(inp.project_name) or "Autogen_Project"
    major = str(inp.major_rev or "35")
    minor = str(inp.minor_rev or "00").zfill(2)
    processor = inp.processor or "1756-L83E"

    cloned = []
    by_area: dict[str, list] = {}
    missing_templates = set()
    for row in inp.conveyors:
        tmpl = resolve_template(row.type, library_text)
        if tmpl != (EXCEL_CONFIG_OVERRIDES.get(row.template_key) or tmpl):
            pass
        excel_pref = EXCEL_CONFIG_OVERRIDES.get(row.template_key)
        if excel_pref and f'Tag Name="{excel_pref}"' not in library_text:
            missing_templates.add(excel_pref)
        is_vfd = "vfd" in (row.type or "").lower()
        item = clone_template_for_conveyor(
            library_text,
            tmpl,
            row.clean_name,
            row.main_area or "Main_Area",
            row.safety_zone or "",
            row.downstream or "",
            is_vfd=is_vfd,
            exit_pe_tag=getattr(row, "exit_pe_tag", "") or "",
            add_pe_tag=getattr(row, "add_pe_tag", "") or "",
            jam_pe_tags=list(getattr(row, "jam_pe_tags", None) or []),
            full_pe_tags=list(getattr(row, "full_pe_tags", None) or []),
            product_pe_tags=list(getattr(row, "product_pe_tags", None) or []),
        )
        item["type"] = row.type
        item["number"] = row.number
        cloned.append(item)
        by_area.setdefault(item["area"], []).append(item)

    # Controller tags: all cloned tags + area helpers
    all_tags = []
    seen_tag_names = set()
    for item in cloned:
        for block in item["tags"]:
            m = re.search(r'<Tag Name="([^"]+)"', block)
            if m and m.group(1) not in seen_tag_names:
                seen_tag_names.add(m.group(1))
                all_tags.append(block)

    def _add_tag_block(block: str) -> None:
        if not block:
            return
        m = re.search(r'<Tag Name="([^"]+)"', block)
        if not m or m.group(1) in seen_tag_names:
            return
        seen_tag_names.add(m.group(1))
        all_tags.append(block)

    def _bool_tag(name: str, value: int = 0) -> str:
        return (
            f'<Tag Name="{_xml_escape(name)}" TagType="Base" DataType="BOOL" '
            f'Radix="Decimal" Constant="false" ExternalAccess="Read/Write">'
            f'<Data Format="L5K"><![CDATA[{value}]]></Data>'
            f'<Data Format="Decorated"><DataValue DataType="BOOL" Value="{value}"/></Data></Tag>'
        )

    def _ensure_library_tag(name: str, fallback_bool: bool = False) -> None:
        """Clone a library controller tag by name, or create a BOOL stub."""
        if name in seen_tag_names:
            return
        block = extract_tag_block(library_text, name)
        if block:
            _add_tag_block(block)
            return
        if fallback_bool:
            _add_tag_block(_bool_tag(name, 0))

    # Simple BOOL tags for areas / zones
    for area in sorted({i["area"] for i in cloned} | set(map(_safe, inp.areas))):
        if not area:
            continue
        for suffix, dtype in (("", "Area_UDT"), ("_Safe", "ES_Zone_UDT"), ("_HMI", "Area_HMI")):
            name = f"{area}{suffix}" if suffix else area
            if name in seen_tag_names:
                continue
            # only create simple BOOL if UDT not in library sample
            if dtype == "Area_UDT" and extract_tag_block(library_text, "Main_Area"):
                block = extract_tag_block(library_text, "Main_Area")
                if block:
                    all_tags.append(block.replace("Main_Area", area))
                    seen_tag_names.add(name)
                    continue
            if suffix == "_Safe" and extract_tag_block(library_text, "Main_Area_Safe"):
                block = extract_tag_block(library_text, "Main_Area_Safe")
                if block:
                    all_tags.append(block.replace("Main_Area_Safe", name).replace("Main_Area", area))
                    seen_tag_names.add(name)
                    continue
            all_tags.append(_bool_tag(name, 0))
            seen_tag_names.add(name)

    # --- Shared tags required by Fast_Conv / Slow_Jam / Slow_Flt (Excel gold wiring) ---
    _add_tag_block(_bool_tag("AlwaysOff", 0))
    _add_tag_block(_bool_tag("AlwaysOn", 1))
    # Null objects — critical for unused PE/Conv slots (edited gold uses these heavily)
    for lib_tag in (
        "NO_PE",
        "NO_Conv",
        "NO_VFD",
        "NO_Enc",
        "NO_MS",
        "NO_AirPress",
        "NO_AdditionalFlt",
        "PConv_VFD",
        "PEExit_P",
        "PEAdd_PE",
        "HMIColor",
        "HMI_StatsClear",
        "Type2",
    ):
        _ensure_library_tag(lib_tag, fallback_bool=(lib_tag in ("HMI_StatsClear",)))
    # NO_PS (PS_UDT) — used by Slow_Flt; not always in library, synthesize if missing
    if "NO_PS" not in seen_tag_names:
        ps_block = extract_tag_block(library_text, "NO_PS")
        if ps_block:
            _add_tag_block(ps_block)
        else:
            _add_tag_block(
                '<Tag Name="NO_PS" TagType="Base" DataType="PS_UDT" Constant="false" '
                'ExternalAccess="Read/Write">'
                '<Data Format="L5K"><![CDATA[[[0],[[0,0,0],0],0]]]></Data>'
                '<Data Format="Decorated"><Structure DataType="PS_UDT"/></Data></Tag>'
            )

    # Safety zones referenced by Fast_Conv after rename (e.g. Zone1_ESZone1)
    for item in cloned:
        sz = item.get("safety_zone") or ""
        if not sz or sz in seen_tag_names:
            continue
        if extract_tag_block(library_text, "Main_Area_Safe"):
            _add_tag_block(
                extract_tag_block(library_text, "Main_Area_Safe").replace("Main_Area_Safe", sz)
            )
        elif extract_tag_block(library_text, "Conv_Area_Safe"):
            _add_tag_block(
                extract_tag_block(library_text, "Conv_Area_Safe").replace("Conv_Area_Safe", sz)
            )
        else:
            _add_tag_block(_bool_tag(sz, 0))

    # --- Ensure every RUN photoeye has PE_UDT (+ AOI) even if not on a conveyor ---
    pe_udt_src = (
        extract_tag_block(library_text, "PE1000_P")
        or extract_tag_block(library_text, "PEExit_P")
        or extract_tag_block(library_text, "NO_PE")
    )
    pe_logic_src = extract_tag_block(library_text, "PE1000_P_AOI")
    full_pe_src = extract_tag_block(library_text, "PE1000_F") or pe_udt_src
    full_aoi_src = extract_tag_block(library_text, "PE1000_F_AOI")
    for pe in getattr(inp, "pe_devices", None) or []:
        pe_name = _safe(pe.get("name") or pe.get("fortna_name") or "")
        if not pe_name or pe_name in seen_tag_names:
            continue
        role = pe.get("role") or _classify_pe_role(pe_name, pe.get("description") or "")
        if role == "full" and full_pe_src:
            _add_tag_block(
                full_pe_src.replace("PE1000_F", pe_name)
                .replace("PEExit_P", pe_name)
                .replace("NO_PE", pe_name)
            )
            if full_aoi_src and f"{pe_name}_AOI" not in seen_tag_names:
                _add_tag_block(
                    full_aoi_src.replace("PE1000_F_AOI", f"{pe_name}_AOI").replace(
                        "PE1000_F", pe_name
                    )
                )
        elif pe_udt_src:
            _add_tag_block(
                pe_udt_src.replace("PE1000_P", pe_name)
                .replace("PEExit_P", pe_name)
                .replace("NO_PE", pe_name)
            )
            if pe_logic_src and f"{pe_name}_AOI" not in seen_tag_names:
                _add_tag_block(
                    pe_logic_src.replace("PE1000_P_AOI", f"{pe_name}_AOI").replace(
                        "PE1000_P", pe_name
                    )
                )

    # --- Non-PE I/O as BOOL tags (PEs are PE_UDT already cloned above) ---
    pe_name_set = {
        _safe(p.device_name)
        for p in (inp.io_points or [])
        if (p.device_type or "") == "photoeye"
    }
    for pe in getattr(inp, "pe_devices", None) or []:
        if pe.get("name"):
            pe_name_set.add(_safe(pe["name"]))

    io_tag_rows: list[dict] = []
    for p in inp.io_points or []:
        raw = (p.device_name or "").strip()
        if not raw:
            continue
        # Spares omitted — engineers add later if needed
        _rn = raw.upper()
        _dt = (p.device_type or "").lower()
        if (
            _dt in ("spare", "invalid")
            or _rn in ("SPARE", "INVALID", "N/A")
            or _rn.startswith("SPARE")
            or "_SPARE" in _rn
        ):
            continue
        tname = _safe(raw)[:40]
        if not tname or tname in seen_tag_names:
            # still record PE rows for CSV / IO_MAP even if tag already exists as PE_UDT
            if tname in pe_name_set or (p.device_type or "") == "photoeye":
                io_tag_rows.append({
                    "tag": tname,
                    "fortna_name": raw,
                    "fortna_address": f"Bank{p.fortna_bank}.{p.fortna_bit}" if p.fortna_bank else (p.source_module or ""),
                    "description": (p.description or f"PE {raw}")[:120],
                    "type": "PE_UDT",
                    "device_class": "photoeye",
                })
            continue
        if tname in pe_name_set or (p.device_type or "") == "photoeye":
            # PE_UDT created in clone_template; skip BOOL
            io_tag_rows.append({
                "tag": tname,
                "fortna_name": raw,
                "fortna_address": f"Bank{p.fortna_bank}.{p.fortna_bit}" if p.fortna_bank else "",
                "description": (p.description or raw)[:120],
                "type": "PE_UDT",
                "device_class": "photoeye",
            })
            continue
        desc = f"Fortna {raw} ({p.device_type or 'io'}) {p.direction or ''}".strip()
        desc_c = (desc[:120]).replace("]]>", "]] >")
        dtype = "BOOL"
        dtype_u = (p.device_type or "").lower()
        # Prefer Fortna UDT stubs for device classes that IO_MAP addresses with .I./.O.
        # (BOOL stubs cause Studio 'Invalid member specifier' on ES500.I.ES_OK etc.)
        udt_block = None
        # ALL project VFD discrete points (VFD500_AUX, VFD500_EN, VFD816A_AUX, …)
        # → one Motor_Starter_UDT P###_VFD per drive number (gold style). No BOOL stubs.
        vfd_m = (
            re.match(r"^VFD(\d+[A-Z]?)(?:_.*)?$", tname, re.I)
            or re.match(r"^VFD(\d+[A-Z]?)(?:_.*)?$", raw, re.I)
            or re.match(r"^T_VFD(\d+[A-Z]?)(?:_.*)?$", tname, re.I)
        )
        if not vfd_m and dtype_u in ("vfd", "drive", "powerflex") and re.search(
            r"VFD(\d+[A-Z]?)", raw, re.I
        ):
            vfd_m = re.search(r"VFD(\d+[A-Z]?)", raw, re.I)
        if vfd_m:
            ms_name = f"P{vfd_m.group(1)}_VFD"
            if ms_name not in seen_tag_names:
                ms_src = extract_tag_block(library_text, "NO_MS")
                if ms_src:
                    _add_tag_block(ms_src.replace("NO_MS", ms_name))
                else:
                    _add_tag_block(
                        f'<Tag Name="{_xml_escape(ms_name)}" TagType="Base" '
                        f'DataType="Motor_Starter_UDT" Constant="false" '
                        f'ExternalAccess="Read/Write">'
                        f'<Data Format="Decorated">'
                        f'<Structure DataType="Motor_Starter_UDT"/></Data></Tag>'
                    )
            # Skip BOOL VFD500_AUX — IO_MAP addresses P###_VFD.I/O members
            io_tag_rows.append({
                "tag": ms_name,
                "fortna_name": raw,
                "fortna_address": (
                    f"Bank{p.fortna_bank}.{p.fortna_bit}" if p.fortna_bank else ""
                ),
                "description": f"VFD discrete → {ms_name} (Motor_Starter_UDT)",
                "type": "Motor_Starter_UDT",
                "device_class": "vfd_ms",
            })
            continue
        # ES_UDT for e-stops AND for MCR/ESR aux contacts that IO_MAP addresses as .I.ES_OK.
        # Names may be T_1MCR1_AUX (digit-leading Fortna tags get T_ prefix).
        needs_es_udt = (
            dtype_u in ("estop", "e-stop", "e_stop", "es")
            or re.match(r"^ES\d", tname, re.I)
            or re.match(r"^ESLS", tname, re.I)
            or re.match(r"^T_\d*ES\d", tname, re.I)  # T_1ES, T_4ES…
            or re.search(r"(?:^|_)(?:\d*)?(?:MCR|ESR)\d", tname, re.I)  # T_1MCR1_AUX, T_4ESR3_AUX
            or re.search(r"(?:^|_)(?:\d*)?(?:MCR|ESR)\d", raw, re.I)
        )
        if needs_es_udt:
            src = (
                extract_tag_block(library_text, "NO_ES")
                or extract_tag_block(library_text, "CP5_ES")
            )
            if src:
                # Rename template tag to this device
                udt_block = re.sub(
                    r'Tag Name="[^"]+"',
                    f'Tag Name="{_xml_escape(tname)}"',
                    src,
                    count=1,
                )
                dtype = "ES_UDT"
        if udt_block:
            _add_tag_block(udt_block)
        else:
            _add_tag_block(
                f'<Tag Name="{_xml_escape(tname)}" TagType="Base" DataType="BOOL" '
                f'Radix="Decimal" Constant="false" ExternalAccess="Read/Write">'
                f'<Description><![CDATA[{desc_c}]]></Description>'
                f'<Data Format="L5K"><![CDATA[0]]></Data>'
                f'<Data Format="Decorated"><DataValue DataType="BOOL" Value="0"/></Data></Tag>'
            )
        bank_addr = (
            f"Bank{p.fortna_bank}.{p.fortna_bit}"
            if p.fortna_bank or p.fortna_bit
            else (p.source_module or "")
        )
        io_tag_rows.append({
            "tag": tname,
            "fortna_name": raw,
            "fortna_address": bank_addr,
            "description": desc,
            "type": dtype,
            "device_class": p.device_type or "",
        })

    def _rung_xml(num: int, text: str, comment: str = "") -> str:
        c = f"<Comment><![CDATA[{comment}]]></Comment>" if comment else ""
        return (
            f'<Rung Number="{num}" Type="N">{c}'
            f"<Text><![CDATA[{text}]]></Text></Rung>"
        )

    def routine(name: str, rungs: list[str]) -> str:
        if not rungs:
            rungs = ['<Rung Number="0" Type="N"><Text><![CDATA[NOP();]]></Text></Rung>']
        fixed = []
        for i, r in enumerate(rungs):
            fixed.append(re.sub(r'Number="\d+"', f'Number="{i}"', r, count=1))
        return (
            f'<Routine Name="{name}" Type="RLL"><RLLContent>'
            + "".join(fixed)
            + "</RLLContent></Routine>"
        )

    def st_routine(name: str, lines: list[str]) -> str:
        """Structured Text routine (gold Area_L2 Merge presets)."""
        if not lines:
            lines = ["NOP();"]
        body = []
        for i, line in enumerate(lines):
            body.append(
                f'<Line Number="{i}"><![CDATA[{line}]]></Line>'
            )
        return (
            f'<Routine Name="{name}" Type="ST"><STContent>'
            + "".join(body)
            + "</STContent></Routine>"
        )

    def _merge_time_preset_lines(time_tag: str) -> list[str]:
        """Gold L2 MergeTime.HMI defaults (Clear/NoCartons/Release/ReleaseFull)."""
        if not time_tag or time_tag.startswith("NO_"):
            return []
        return [
            f"{time_tag}.HMI.ClearTime := 8000;",
            f"{time_tag}.HMI.NoCartonsTime := 8000;",
            f"{time_tag}.HMI.ReleaseTime := 10000;",
            f"{time_tag}.HMI.ReleaseTimeFull := 15000;",
        ]

    # Build programs per area — ModuleB-shaped pack (PLC2 gold):
    # Fast / Slow / L1 / L2 (+ Conv_Merge when merges configured)
    programs_xml = []
    pe_wired_count = 0
    flt_count = 0
    include_prog_set = {
        str(x).strip()
        for x in (getattr(inp, "include_programs", None) or [])
        if str(x).strip()
    }
    want_sys_comm = "System" in include_prog_set or "Sys_Comm" in include_prog_set

    def _nop_rung(comment: str = "") -> str:
        return _rung_xml(0, "NOP();", comment or "Site customize")

    for area, items in sorted(by_area.items()):
        rungs_fast: list[str] = []
        rungs_jam: list[str] = []
        rungs_pe: list[str] = []
        rungs_full: list[str] = []
        rungs_flt: list[str] = []
        area_convs: list[str] = []
        for item in items:
            cname = (item.get("conveyor") or "").strip()
            if cname:
                area_convs.append(cname)
            for r in item["rungs"]:
                rx = _rung_xml(0, r["text"], r.get("comment") or "")
                if r["label"] == "Fast":
                    rungs_fast.append(rx)
                elif r["label"] == "Jam":
                    rungs_jam.append(rx)
                elif r["label"] == "Flt":
                    rungs_flt.append(rx)
                    flt_count += 1
                elif r["label"] == "Full":
                    rungs_full.append(rx)
                    pe_wired_count += 1
                elif r["label"] == "PE":
                    rungs_pe.append(rx)
                    if "PE_Logic" in r["text"] or "Full_PE" in r["text"]:
                        pe_wired_count += 1

        # Area Slow — only emit routines Autogen fills (no empty CS / PI / Stacklight scaffolds)
        main_slow = [
            _rung_xml(0, "JSR(Conv_Flt,0);", "Conv_Flt"),
            _rung_xml(1, "JSR(Conv_Jam,0);", "Conv_Jam"),
            _rung_xml(2, "JSR(Conv_PE,0);", "Conv_PE"),
        ]
        # Gold ModuleB_Area_Fast Main JSR chain
        main_fast = [
            _rung_xml(0, "JSR(Conv_Fast,0);", "Conv_Fast"),
            _rung_xml(1, "JSR(Conv_Full,0);", "Conv_Full"),
            _rung_xml(2, "JSR(Conv_Merge,0);", "Conv_Merge"),
            _rung_xml(3, "JSR(Conv_PE,0);", "Conv_PE"),
        ]

        # Merges for this area (gold Conv_Merge + L2 ST). UI stores 2:1 / 3:1+;
        # L5X emit is 2:1 only for now (3:1+ kept in workbook for later).
        rungs_merge: list[str] = []
        merge_st_lines: list[str] = []
        area_merges = [
            m for m in (getattr(inp, "merges_2to1", None) or [])
            if isinstance(m, dict)
            and (
                not (m.get("area") or "").strip()
                or _safe(m.get("area") or "") == _safe(area)
                or (m.get("area") or "").strip() == area
            )
        ]
        for m in area_merges:
            try:
                lane_n = int(m.get("lanes") or m.get("lane_count") or 2)
            except (TypeError, ValueError):
                lane_n = 2
            if lane_n > 2:
                # Config captured from prints; codegen TBD
                continue
            name = _safe(m.get("name") or m.get("merge") or "")
            if not name:
                continue
            if not name.endswith("_Merge"):
                merge_tag = f"{name}_Merge"
            else:
                merge_tag = name
                name = name[: -len("_Merge")] or name
            lane_a = _safe(m.get("lane_a") or m.get("induct") or "")
            lane_b = _safe(m.get("lane_b") or m.get("main") or "")
            discharge = _safe(m.get("discharge") or m.get("out") or name)
            # Empty PE must stay NO_PE — _safe("") becomes "Tag"
            _pe_a = (m.get("pe_a") or "").strip()
            _pe_b = (m.get("pe_b") or "").strip()
            _jam = (m.get("jam_pe") or "").strip()
            # Allow blank / NO_PE / override — never emit a PE name that won't exist.
            # Known PEs = RUN pe_devices + IO photoeyes (already cloned above).
            # Unknown PEs: create PE_UDT stub if allow_undefined_pe, else NO_PE.
            allow_undef_pe = bool(
                m.get("allow_undefined_pe")
                or m.get("create_missing_pe")
                or getattr(inp, "allow_undefined_merge_pe", False)
            )

            def _merge_pe_operand(raw: str) -> str:
                if not raw or raw.upper() in ("NO_PE", "NONE", "-", "N/A"):
                    return "NO_PE"
                pe_name = _safe(raw)
                if pe_name in pe_name_set or pe_name in seen_tag_names:
                    return pe_name
                if allow_undef_pe and pe_udt_src:
                    # Wizard override: still build — clone PE_UDT so Studio resolves
                    _add_tag_block(
                        pe_udt_src.replace("PE1000_P", pe_name)
                        .replace("PEExit_P", pe_name)
                        .replace("NO_PE", pe_name)
                    )
                    if pe_logic_src and f"{pe_name}_AOI" not in seen_tag_names:
                        _add_tag_block(
                            pe_logic_src.replace("PE1000_P_AOI", f"{pe_name}_AOI").replace(
                                "PE1000_P", pe_name
                            )
                        )
                    pe_name_set.add(pe_name)
                    return pe_name
                # Safe default — build succeeds; engineer can wire PE later
                return "NO_PE"

            pe_a = _merge_pe_operand(_pe_a)
            pe_b = _merge_pe_operand(_pe_b)
            jam_pe = _merge_pe_operand(_jam)
            conv_a = f"{lane_a}_Conv" if lane_a and not lane_a.endswith("_Conv") else (lane_a or "NO_Conv")
            conv_b = f"{lane_b}_Conv" if lane_b and not lane_b.endswith("_Conv") else (lane_b or "NO_Conv")
            conv_out = (
                f"{discharge}_Conv"
                if discharge and not discharge.endswith("_Conv")
                else (discharge or "NO_Conv")
            )
            time_a = f"{lane_a}_MergeTime" if lane_a else "NO_MergeTime"
            time_b = f"{lane_b}_MergeTime" if lane_b else "NO_MergeTime"
            # hold_mode: runhold (PLC2/4 BOOL) | stop_next (PLC5 Conv.PI.Stop_Next)
            hold_mode = str(m.get("hold_mode") or m.get("hold") or "runhold").strip().lower()
            if hold_mode in ("stop_next", "stopnext", "pi_stop_next", "plc5"):
                hold_main = f"{conv_a}.PI.Stop_Next"
                hold_induct = f"{conv_b}.PI.Stop_Next"
                make_hold_bools = False
            else:
                hold_main = f"{name}_MainLane_Conv_RunHold"
                hold_induct = f"{name}_InductLane_Conv_RunHold"
                make_hold_bools = True
            # Tags: Merge_2to1 instance + Merge_Time + optional BOOL holds
            if merge_tag not in seen_tag_names:
                _add_tag_block(
                    f'<Tag Name="{_xml_escape(merge_tag)}" TagType="Base" '
                    f'DataType="Merge_2to1" Constant="false" ExternalAccess="Read/Write">'
                    f'<Data Format="Decorated"><Structure DataType="Merge_2to1"/></Data></Tag>'
                )
            for tname in (time_a, time_b):
                if tname.startswith("NO_") or tname in seen_tag_names:
                    continue
                _add_tag_block(
                    f'<Tag Name="{_xml_escape(tname)}" TagType="Base" '
                    f'DataType="Merge_Time" Constant="false" ExternalAccess="Read/Write">'
                    f'<Data Format="Decorated"><Structure DataType="Merge_Time"/></Data></Tag>'
                )
            if make_hold_bools:
                for bname in (hold_main, hold_induct):
                    if bname not in seen_tag_names:
                        _add_tag_block(
                            f'<Tag Name="{_xml_escape(bname)}" TagType="Base" DataType="BOOL" '
                            f'Radix="Decimal" Constant="false" ExternalAccess="Read/Write">'
                            f'<Data Format="L5K"><![CDATA[0]]></Data>'
                            f'<Data Format="Decorated"><DataValue DataType="BOOL" Value="0"/></Data></Tag>'
                        )
            # Gold Merge_2to1 signature (timers/CX members on merge UDT)
            text = (
                f"Merge_2to1({merge_tag},{conv_a},{conv_b},{conv_out},{conv_out},NO_Conv,"
                f"1,1,{pe_a},{pe_b},{jam_pe},0,NO_PE,NO_PE,"
                f"{time_a},{time_b},0,"
                f"{merge_tag}.I_Merge_FltClearTime,{merge_tag}.I_MergeCX_Enable,"
                f"{merge_tag}.I_MergeCX_TimeReset,{hold_main},{hold_induct});"
            )
            rungs_merge.append(
                _rung_xml(
                    0,
                    text,
                    f"~~~~~~~~~~~\n{merge_tag} 2:1 Merge\n(equipment pattern — Site Forge)\n~~~~~~~~~~~",
                )
            )
            # Gold Area_L2 Merge ST presets
            merge_st_lines.extend(
                [
                    f"// {merge_tag}",
                    f"{merge_tag}.I_MergeCX_Enable := 0;",
                    f"{merge_tag}.I_InductLane_AddNotlReadyBit := 0;",
                    f"{merge_tag}.I_MainLane_AddNotReadyBit := 0;",
                    f"{merge_tag}.I_Merge_FltClearTime := 10000;",
                    "",
                    "// Main lane MergeTime",
                    *_merge_time_preset_lines(time_a),
                    "",
                    "// Induct lane MergeTime",
                    *_merge_time_preset_lines(time_b),
                    "",
                ]
            )

        # Only JSR Conv_Merge when we have merge rungs (still list routine as NOP otherwise)
        if not rungs_merge:
            # Remove Conv_Merge JSR from Fast main when no merges for this area
            main_fast = [r for r in main_fast if "Conv_Merge" not in r]

        prog_slow = f"{area}_Slow" if area.endswith("_Area") else f"{area}_Area_Slow"
        prog_fast = f"{area}_Fast" if area.endswith("_Area") else f"{area}_Area_Fast"
        prog_l1 = f"{area}_L1" if area.endswith("_Area") else f"{area}_Area_L1"
        prog_l2 = f"{area}_L2" if area.endswith("_Area") else f"{area}_Area_L2"
        prog_slow = _safe(prog_slow)[:40]
        prog_fast = _safe(prog_fast)[:40]
        prog_l1 = _safe(prog_l1)[:40]
        prog_l2 = _safe(prog_l2)[:40]

        # --- Slow (Flt / Jam / PE only — site CS / PI / Stacklight left for later) ---
        programs_xml.append(
            f'<Program Name="{prog_slow}" TestEdits="false" MainRoutineName="Main_Routine" '
            f'Disabled="false" UseAsFolder="false">'
            f"<Tags/>"
            f"<Routines>"
            f'{routine("Main_Routine", main_slow)}'
            f'{routine("Conv_Flt", rungs_flt)}'
            f'{routine("Conv_Jam", rungs_jam)}'
            f'{routine("Conv_PE", rungs_pe)}'
            f"</Routines></Program>"
        )

        # --- Fast (ModuleB_Area_Fast shape) ---
        fast_routines = (
            f'{routine("Main_Routine", main_fast)}'
            f'{routine("Conv_Fast", rungs_fast)}'
            f'{routine("Conv_Full", rungs_full or [_nop_rung("No Full_PE eyes on this area")])}'
            f'{routine("Conv_Merge", rungs_merge or [_nop_rung("No 2:1 merges for this area")])}'
            f'{routine("Conv_PE", rungs_pe)}'
        )
        programs_xml.append(
            f'<Program Name="{prog_fast}" TestEdits="false" MainRoutineName="Main_Routine" '
            f'Disabled="false" UseAsFolder="false">'
            f"<Tags/>"
            f"<Routines>"
            f"{fast_routines}"
            f"</Routines></Program>"
        )

        # --- L1 ST presets (ModuleB_Area_L1) ---
        l1_area_lines = [
            f"// {_safe(area)} Area presets",
            f"{_safe(area)}.StartTime := 1000;",
        ]
        l1_conv_lines = [f"// Conveyor presets — {len(area_convs)} belt(s)"]
        l1_ms_lines = ["// Motor starter fault times (defaults; Sys Init may override later)"]
        for cn in area_convs:
            base = _safe(cn)
            l1_conv_lines.append(f"// {base}")
            l1_ms_lines.extend(
                [
                    f"//{base}_MS",
                    f"{base}_MS.MtrOverload_FltTime := 5000;",
                    f"{base}_MS.MtrContactor_FltTime := 5000;",
                    "",
                ]
            )
        main_l1 = [
            _rung_xml(0, "JSR(Area,0);", "Area"),
            _rung_xml(1, "JSR(Conv,0);", "Conv"),
            _rung_xml(2, "JSR(CS,0);", "CS"),
            _rung_xml(3, "JSR(MS_Time,0);", "MS_Time"),
            _rung_xml(4, "JSR(PS_Time,0);", "PS_Time"),
            _rung_xml(5, "JSR(PWS_Time,0);", "PWS_Time"),
        ]
        programs_xml.append(
            f'<Program Name="{prog_l1}" TestEdits="false" MainRoutineName="Main_Routine" '
            f'Disabled="false" UseAsFolder="false">'
            f"<Tags/>"
            f"<Routines>"
            f'{routine("Main_Routine", main_l1)}'
            f"{st_routine('Area', l1_area_lines)}"
            f"{st_routine('Conv', l1_conv_lines)}"
            f"{st_routine('CS', ['// Control stations — site customize'])}"
            f"{st_routine('MS_Time', l1_ms_lines)}"
            f"{st_routine('PS_Time', ['// Power supplies — site customize'])}"
            f"{st_routine('PWS_Time', ['// PWS timers — site customize'])}"
            f"</Routines></Program>"
        )

        # --- L2 ST presets (ModuleB_Area_L2) — always for transport areas ---
        main_l2 = [
            _rung_xml(0, "JSR(Conv_Speed,0);", "Conv_Speed"),
            _rung_xml(1, "JSR(FullTime,0);", "FullTime"),
            _rung_xml(2, "JSR(Merge,0);", "Merge"),
            _rung_xml(3, "JSR(PETime,0);", "PETime"),
        ]
        if not merge_st_lines:
            merge_st_lines = ["// No merges in this area"]
        programs_xml.append(
            f'<Program Name="{prog_l2}" TestEdits="false" MainRoutineName="Main_Routine" '
            f'Disabled="false" UseAsFolder="false">'
            f"<Tags/>"
            f"<Routines>"
            f'{routine("Main_Routine", main_l2)}'
            f"{st_routine('Conv_Speed', ['// Conv speed presets — site customize'])}"
            f"{st_routine('FullTime', ['// Full PE timers — site customize'])}"
            f"{st_routine('Merge', merge_st_lines)}"
            f"{st_routine('PETime', ['// PE timers — site customize'])}"
            f"</Routines></Program>"
        )

    # --- Sys_Comm (gold System under P11_Slow_200ms) — Devices_Comm_Logic from tar EIP/RIO ---
    if want_sys_comm:
        programs_xml.append(_build_sys_comm_program_xml(
            inp=inp,
            library_text=library_text,
            _add_tag_block=_add_tag_block,
            _rung_xml=_rung_xml,
            routine=routine,
            seen_tag_names=seen_tag_names,
        ))

    # --- IO_MAP from RUN/tar.gz (default): Conveyor.asc Bank.Bit → AENTR:I/O.Data[slot] ---
    # Gold Excel IO_MAP_Program.L5X is optional (include_io_map_gold) and replaces this scaffold.
    word_map = dict(getattr(inp, "io_word_map", None) or {})
    bank_index = _build_eip_bank_index(list(getattr(inp, "eip_topology", None) or []))
    configio_map = dict(getattr(inp, "configio_octal_map", None) or {})

    def _word_info(word: str, *, want_dir: str = "", bit: str = "") -> dict | None:
        """Resolve Fortna word → module. Prefer Configio, then heuristics, then EIPCSV."""
        w = str(word or "").strip()
        if not w:
            return None
        # 1) Configio + bank heuristics (Reno: EIPCSV empty)
        if want_dir:
            hit = _resolve_fortna_bank(
                w,
                want_dir=want_dir,
                bank_index=bank_index,
                bit=bit,
                configio_map=configio_map,
            )
            if hit:
                return hit
        # 2) Legacy word_map from EIPCSV / InputBank index
        info = word_map.get(w)
        if info:
            return info
        try:
            info = word_map.get(str(int(float(w))))
        except Exception:
            info = None
        if info:
            return info
        if w.isdigit() and int(w) % 2 == 1:
            return word_map.get(str(int(w) - 1))
        return None

    def _vfd_ms_member(tname: str, direction: str) -> str | None:
        """Map every VFD###_* point → P###_VFD Motor_Starter_UDT members.

        Applies to ALL VFDs on the project (VFD118, VFD500, VFD816A, …), not one example.
        Gold: AUX → .I.Auxiliary_Forward; EN out → .O.Run (no bare BOOL VFD tags).
        """
        m = (
            re.match(r"^VFD(\d+[A-Z]?)(?:_(.+))?$", tname, re.I)
            or re.match(r"^T_VFD(\d+[A-Z]?)(?:_(.+))?$", tname, re.I)
        )
        if not m:
            return None
        num = m.group(1)
        suffix = (m.group(2) or "").upper()
        base = f"P{num}_VFD"
        d = (direction or "").upper()
        if suffix in ("AUX", "AUXILIARY", "AUX_FWD", "AF", "FB", "FEEDBACK"):
            return f"{base}.I.Auxiliary_Forward"
        if suffix in ("OK", "CONTACTOR", "CONTACTOR_OK", "C_OK"):
            return f"{base}.I.Contactor_OK"
        if suffix in ("MS_OK", "OL", "OVERLOAD", "OVL"):
            return f"{base}.I.MS_OK"
        if suffix in ("EN", "ENABLE", "RUN", "CMD", "START"):
            # Output enable → UDT.O.Run; input-style EN rare → Contactor_OK
            if d in ("O", "OUT", "OUTPUT"):
                return f"{base}.O.Run"
            return f"{base}.I.Contactor_OK"
        if not suffix:
            if d in ("O", "OUT", "OUTPUT"):
                return f"{base}.O.Run"
            return f"{base}.I.Auxiliary_Forward"
        return None

    def _device_member(device_type: str, tname: str, direction: str) -> str:
        """Logix tag member for OTE/XIC on the field-device side.

        Only use .I.ES_OK / .I.PE_Clear when the tag is (or will be) a UDT with
        those members. Plain BOOL tags must be referenced bare — Studio error:
        'Invalid member specifier' if you write BOOL.I.ES_OK.
        """
        dt = (device_type or "").lower()
        vfd_m = _vfd_ms_member(tname, direction)
        if vfd_m:
            return vfd_m
        if dt == "photoeye" or re.match(r"^(?:EZ)?PE\d", tname, re.I):
            return f"{tname}.I.PE_Clear"
        if (
            dt in ("estop", "e-stop", "e_stop", "es")
            or re.match(r"^ES\d", tname, re.I)
            or re.match(r"^ESLS", tname, re.I)
            or re.match(r"^T_\d*ES\d", tname, re.I)
            or re.search(r"(?:^|_)(?:\d*)?(?:MCR|ESR)\d", tname, re.I)
        ):
            return f"{tname}.I.ES_OK"
        # BOOL / simple devices (PB, motor aux, digital, beacon force bit, …)
        return tname

    def _is_output_point(p: IoPoint, info: dict | None) -> bool:
        d = (getattr(p, "direction", None) or "").upper()
        if d in ("O", "OUT", "OUTPUT"):
            return True
        if info and (info.get("direction") or "").upper() == "O":
            return True
        dt = (getattr(p, "device_type", None) or "").lower()
        if dt in ("beacon",):
            return True
        return False

    cp_i_rungs: list[str] = [
        _rung_xml(0, "NOP();", "CP_I — RUN/tar.gz inputs → device tags"),
    ]
    cp_o_rungs: list[str] = [
        _rung_xml(0, "NOP();", "CP_O — device tags → RUN/tar.gz outputs"),
    ]
    io_map_mapped = 0
    io_map_unmapped = 0
    io_map_skipped_spare = 0
    io_map_skipped_dir = 0

    def _is_spare_io_point(p: IoPoint) -> bool:
        n = (p.device_name or "").strip().upper()
        dt = (p.device_type or "").strip().lower()
        desc = (getattr(p, "description", None) or "").strip().upper()
        if dt in ("spare", "invalid"):
            return True
        if not n or n in ("SPARE", "INVALID", "N/A", "NONE"):
            return True
        if n.startswith("SPARE") or "_SPARE" in n or n.endswith("SPARE"):
            return True
        if "SPARE" in desc and "NOT SPARE" not in desc:
            return True
        return False

    io_map_skipped_spare = sum(
        1 for p in (inp.io_points or []) if _is_spare_io_point(p)
    )
    map_points = [
        p for p in (inp.io_points or [])
        if (p.device_name or "").strip()
        and (p.fortna_bank or p.fortna_bit)
        and not _is_spare_io_point(p)
    ]
    # Resolve first, then emit in numerical adapter/slot order (AENTR3…AENTR14)
    resolved_rows: list[dict] = []
    for p in map_points:
        tname = _safe(p.device_name)
        if not tname:
            continue
        word = str(p.fortna_bank or "").strip()
        fbit = str(p.fortna_bit or "").strip()
        want = _io_point_want_dir(p.device_name or "", p.device_type or "", p.direction or "")
        info = _word_info(word, want_dir=want, bit=fbit)
        # No EIP card for this bank → omit (spares / unfinished can be added later)
        if not info:
            io_map_unmapped += 1
            continue
        mod_type = (info.get("type") or "")
        family = (info.get("family") or (
            "1734" if "1734" in mod_type or "1738" in mod_type else "1794"
        ))
        max_bit = _point_card_max_bit(mod_type) if family == "1734" else 15
        # Configio splits one Fortna word across two cards (Low/High). High-half
        # bits (octal 10-17) become Data[0..] on the High card — remap before clamp.
        bit_for_card = fbit
        if _fortna_bit_is_high(fbit) and (info.get("resolve_how") or "") == "configio":
            try:
                hv = int(str(fbit).strip(), 8)
            except ValueError:
                try:
                    hv = int(str(fbit).strip(), 10)
                except ValueError:
                    hv = -1
            if hv >= 8:
                bit_for_card = str(hv - 8)
        data_bit = _fortna_bit_to_data_bit(bit_for_card, max_bit=max_bit)
        if data_bit is None:
            io_map_unmapped += 1
            continue
        member = _device_member(p.device_type or "", tname, p.direction or "")
        how = info.get("resolve_how") or "map"
        comment = (
            f"{tname} · Bank{word}.{fbit}"
            + (f" · EIP{info.get('resolved_bank')}" if info.get("resolved_bank") else "")
            + (f" · {info.get('type')}" if info else "")
            + (f" · via {how}" if how not in ("direct", "map") else "")
        )
        rio = info["rio_name"]
        slot = int(info["flex_slot"])
        # PHYSICAL card direction wins (IB8 has no :O — never OTE to an input card).
        mod_dir = (info.get("direction") or want or "").upper()
        if mod_dir not in ("I", "O"):
            io_map_skipped_dir += 1
            continue
        channel = f"{rio}:{mod_dir}.Data[{slot}].{data_bit}"
        resolved_rows.append({
            "rio": rio,
            "slot": slot,
            "data_bit": data_bit,
            "mod_dir": mod_dir,
            "member": member,
            "channel": channel,
            "comment": comment,
            "tname": tname,
        })
        io_map_mapped += 1

    resolved_rows.sort(
        key=lambda r: (
            _rio_numeric_key(r["rio"]),
            r["slot"],
            r["data_bit"],
            r["tname"],
        )
    )
    last_rio_i = ""
    last_rio_o = ""
    for row in resolved_rows:
        rio = row["rio"]
        if row["mod_dir"] == "O":
            if rio != last_rio_o:
                cp_o_rungs.append(_rung_xml(0, "NOP();", rio))
                last_rio_o = rio
            text = f"XIC({row['member']})OTE({row['channel']});"
            cp_o_rungs.append(_rung_xml(0, text, row["comment"]))
        else:
            if rio != last_rio_i:
                cp_i_rungs.append(_rung_xml(0, "NOP();", rio))
                last_rio_i = rio
            text = f"XIC({row['channel']})OTE({row['member']});"
            cp_i_rungs.append(_rung_xml(0, text, row["comment"]))

    # --- Gold program exports: Sys (default). Gold Excel IO_MAP is CLI-only. ---
    # UI "IO_MAP" checkbox → include_io_map (RUN banks). Gold Excel is separate.
    want_io_map = bool(getattr(inp, "include_io_map", True))
    want_gold_io = bool(getattr(inp, "include_io_map_gold", False))
    site_rios = {
        str((info or {}).get("rio_name") or "").strip().upper()
        for info in (getattr(inp, "io_word_map", None) or {}).values()
        if isinstance(info, dict) and (info or {}).get("rio_name")
    }
    gold_io_blocked = False
    if want_gold_io and site_rios:
        site_is_cp5_family = any(r.startswith(("CP5", "CP6", "CP7")) for r in site_rios)
        site_is_cp1_4 = any(r.startswith(("CP1", "CP2", "CP3", "CP4")) for r in site_rios)
        if site_is_cp1_4 and not site_is_cp5_family:
            want_gold_io = False
            gold_io_blocked = True
            _emit_progress(
                "Gold Excel IO_MAP blocked — this site uses CP1–CP4 (use RUN IO_MAP)",
                35,
            )
    # Gold Excel replaces RUN map when allowed; otherwise RUN map if include_io_map
    if want_gold_io:
        want_io_map = True  # gold is an IO_MAP

    gold_programs = resolve_program_exports(
        list(getattr(inp, "include_programs", None) or []),
        include_sys=bool(getattr(inp, "include_sys", True)),
        include_io_map_gold=want_gold_io,
    )
    gold_program_names: list[str] = []
    gold_io_map_used = False
    extra_aoi_chunks: list[str] = []
    extra_dt_chunks: list[str] = []

    def _tag_datatype(block: str) -> str:
        m = re.search(r'\bDataType="([^"]+)"', block)
        return (m.group(1) if m else "").strip()

    def _is_atomic_dtype(dt: str) -> bool:
        return (dt or "").upper() in {
            "BOOL", "SINT", "INT", "DINT", "LINT", "USINT", "UINT", "UDINT", "ULINT",
            "REAL", "LREAL", "STRING", "TIMER", "COUNTER", "CONTROL",
        }

    def _upsert_tag_block(block: str, *, prefer: bool = False) -> None:
        """Add tag, or replace an existing atomic stub when gold has a richer UDT.

        RUN equipment scan creates many BOOL stubs (ES500, ES610C, …). Gold IO_MAP
        exports the same names as ES_UDT / CS_UDT / DINT. If the BOOL wins, Studio
        reports 'Invalid member specifier' on OTE(ES500.I.ES_OK).
        """
        m = re.search(r'<Tag Name="([^"]+)"', block)
        if not m:
            return
        tname = m.group(1)
        new_dt = _tag_datatype(block)
        if tname not in seen_tag_names:
            seen_tag_names.add(tname)
            all_tags.append(block)
            return
        if not prefer:
            return
        # Replace atomic stub with gold structured type
        for i, existing in enumerate(all_tags):
            em = re.search(r'<Tag Name="([^"]+)"', existing)
            if not em or em.group(1) != tname:
                continue
            old_dt = _tag_datatype(existing)
            if old_dt == new_dt:
                return
            # Always prefer non-atomic (UDT) over atomic; also BOOL → DINT etc.
            if _is_atomic_dtype(old_dt) and (not _is_atomic_dtype(new_dt) or new_dt.upper() != "BOOL"):
                all_tags[i] = block
            return

    site_stem = _safe(inp.project_name) or proj
    for gp in gold_programs:
        gname = gp["name"]
        # Prefer gold IO_MAP over RUN scaffold
        if gname == "IO_MAP":
            gold_io_map_used = True
        # Retarget Greensboro gold names → this site (MSCRENO_MSCRENOPACK, etc.)
        # Merge controller tags from the program export — gold wins over BOOL stubs
        for block in gp.get("tags") or []:
            _upsert_tag_block(
                _retarget_gold_site_names(block, site_stem), prefer=True
            )
        if gp.get("aois_xml"):
            extra_aoi_chunks.append(
                _retarget_gold_site_names(gp["aois_xml"], site_stem)
            )
        if gp.get("datatypes_xml"):
            extra_dt_chunks.append(
                _retarget_gold_site_names(gp["datatypes_xml"], site_stem)
            )
        programs_xml.append(
            _retarget_gold_site_names(gp["program_xml"], site_stem)
        )
        gold_program_names.append(gname)

    # --- Equipment plan from tar → auto-hint packs (Sorter Track, merges note) ---
    equip = dict(getattr(inp, "equipment_plan", None) or {})
    plan = dict(equip.get("plan") or {})
    if plan.get("features", {}).get("sorter_track_pack"):
        inc = list(getattr(inp, "include_programs", None) or [])
        if "Sorter_Track" not in inc:
            # Auto-include when tar shows sorter/ENC evidence (can still be empty config)
            inc.append("Sorter_Track")
            inp.include_programs = inc
            _emit_progress("Tar equipment plan → auto-include Sorter_Track", 38)

    # --- Live Sorter_Track: ONLY when Program pack includes Sorter_Track (checkbox) ---
    # Double safety: sorter_build UI alone does not emit the program.
    sorter_cfg = dict(getattr(inp, "sorter_build", None) or {})
    sorter_report: dict = {}
    want_sorter = any(
        (x or "").strip().lower().replace(" ", "_") in (
            "sorter_track", "sortertrack",
        )
        for x in (getattr(inp, "include_programs", None) or [])
    )
    try:
        from fortna_sorter_build import (
            build_sorter_track,
            sorter_build_is_configured,
        )
        # Automate: workbook sorter_build with real data ⇒ include pack even if
        # UI forgot the checkbox (good not perfect).
        if not want_sorter and sorter_build_is_configured(sorter_cfg):
            want_sorter = True
            inc = list(getattr(inp, "include_programs", None) or [])
            if "Sorter_Track" not in inc:
                inc.append("Sorter_Track")
                inp.include_programs = inc
            _emit_progress(
                "Sorter build data present → auto-including Sorter_Track pack",
                41,
            )
        if want_sorter:
            # Sorter_Track_Program.L5X (gold pack) configured from Sorter build UI
            live = build_sorter_track(
                sorter_cfg if sorter_build_is_configured(sorter_cfg) else {
                    "divert_count": 0,
                    "tracking": [],
                },
                library_text,
                io_points=list(inp.io_points or []),
                word_map=dict(getattr(inp, "io_word_map", None) or {}),
            )
            sorter_report = live.get("report") or {}
            sorter_report["pack_checkbox"] = True
            for block in live.get("tags") or []:
                _upsert_tag_block(block, prefer=True)
            if live.get("aoi_xml"):
                extra_aoi_chunks.append(live["aoi_xml"])
            if live.get("datatypes_xml"):
                extra_dt_chunks.append(live["datatypes_xml"])
            programs_xml.append(live["program_xml"])
            gold_program_names.append("Sorter_Track")
            mode = sorter_report.get("mode") or "configured_pack"
            _emit_progress(
                f"Sorter_Track ({mode}): "
                f"diverts kept={sorter_report.get('wave_rungs_kept', sorter_report.get('divert_count'))}/"
                f"{sorter_report.get('wave_rungs_in_pack', '?')}, "
                f"encoders={sorter_report.get('encoder_count', 0)}",
                42,
            )
        elif sorter_build_is_configured(sorter_cfg):
            sorter_report = {
                "mode": "skipped",
                "reason": "Sorter build has data but Program pack Sorter Track is off",
            }
            _emit_progress(
                "Sorter build saved but pack OFF — not emitting Sorter_Track",
                42,
            )
    except Exception as ex:
        sorter_report = {"mode": "error", "error": str(ex)}
        _emit_progress(f"Sorter_Track build failed: {ex}", 42)

    if not gold_io_map_used and want_io_map:
        # RUN/tar.gz map: CP_I (inputs) + CP_O (outputs) from Conveyor.asc + EIP word_map
        main_io_rungs = [
            _rung_xml(0, "JSR(CP_I,0);", "Inputs: CPxRIOn:I.Data → device tags (from RUN)"),
            _rung_xml(0, "JSR(CP_O,0);", "Outputs: device tags → CPxRIOn:O.Data (from RUN)"),
        ]
        programs_xml.append(
            f'<Program Name="IO_MAP" TestEdits="false" MainRoutineName="Main_Routine" '
            f'Disabled="false" UseAsFolder="false">'
            f"<Tags/>"
            f"<Routines>"
            f'{routine("Main_Routine", main_io_rungs)}'
            f'{routine("CP_I", cp_i_rungs)}'
            f'{routine("CP_O", cp_o_rungs)}'
            f"</Routines></Program>"
        )
    elif not want_io_map and not gold_io_map_used:
        _emit_progress("IO_MAP omitted (checkbox off)", 40)

    # Modules section — lightweight aliases from IO sheet (full module XML from lib is complex)
    def _extract_module_xml(lib: str, mod_name: str) -> str | None:
        """Pull one complete <Module>…</Module> (or true self-close) from library.

        IMPORTANT: opening-tag attrs must use [^>]* not [^/]*. The latter matches
        newlines and wrongly ends at the first nested '/>' (e.g. <EKey …/>),
        producing unclosed <Module> that Studio rejects (Xml_TagMismatchEx).
        """
        esc = re.escape(mod_name)
        # True self-closing Module (entire element is one tag)
        m = re.search(rf'<Module\s+Name="{esc}"(?:\s[^>]*)?/>', lib)
        if m:
            return m.group(0)
        # Full element — non-greedy to first </Module> (Modules are siblings, not nested)
        m = re.search(rf'<Module\s+Name="{esc}"(?:\s[^>]*)?>.*?</Module>', lib, re.S)
        if m:
            block = m.group(0)
            # Must start with Module and end with </Module> (not a nested self-close)
            if not re.match(r"<Module\b", block) or not re.search(r"</Module>\s*$", block):
                return None
            # Exactly one Module open (no nested Module siblings swallowed)
            if len(re.findall(r"<Module\b", block)) != 1:
                return None
            return block
        return None

    # Build a clean Modules tree (matches gold Excel Autogen shape):
    #   Local + EN2T backbone + CPxRIOn AENT/children only.
    # Do NOT dump the whole library Modules section (demo PointIO / scales / VFDs
    # cause Studio import noise and Local→Local1 rename when merging into an ACD).
    # Always open L5X as a NEW project (File→Open), not Import into an existing ACD.
    #
    # Parent Bus Size MUST match the AENT datatype in the library template:
    #   1794-AENT → AB:1794_AEN_8SLOT → Bus Size 8
    #   1734-AENT → AB:1734_40SLOT   → Bus Size 40
    # Wrong size → Studio "Data type mismatch" and children ParentModule not found.
    eip_module_names: list[str] = []
    eip_child_names: list[str] = []
    parent_tmpls = {
        fam: _extract_module_xml(library_text, name)
        for fam, name in EIP_PARENT_TEMPLATE.items()
    }
    child_tmpls = {
        key: _extract_module_xml(library_text, tmpl_name)
        for key, tmpl_name in EIP_CHILD_TEMPLATE.items()
    }
    enet_parent = "CPXXENET1"  # library EN2T under Local
    topology = list(getattr(inp, "eip_topology", None) or [])
    # Greensboro multi-panel tars: optionally trim to word_map RIO only.
    # Reno / SHIP: word_map can be nearly empty (empty EIPCSV) — NEVER drop
    # AENTR* racks from EIPAdapters (SHIP previously collapsed to AENTR_4 only).
    used_rios = {
        str((info or {}).get("rio_name") or "").strip()
        for info in (getattr(inp, "io_word_map", None) or {}).values()
        if isinstance(info, dict) and (info or {}).get("rio_name")
    }
    aentr_in_topo = [
        ad for ad in topology
        if str(ad.get("rio_name") or "").upper().startswith("AENTR")
    ]
    if used_rios and len(used_rios) >= 2 and not (
        # Keep all AENTR* when tar lists multiple print racks
        len(aentr_in_topo) > len(used_rios)
    ):
        filtered = [ad for ad in topology if (ad.get("rio_name") or "") in used_rios]
        if filtered:
            topology = filtered
    # Always prefer emitting every AENTR* adapter present in topology
    elif aentr_in_topo:
        topology = aentr_in_topo

    # --- Local controller module (fixed Ports like gold EDITED L5X) ---
    # Library Local is often L81E at chassis slot 4 — wrong for L83E open/import.
    local_mod = _extract_module_xml(library_text, "Local") or ""
    proc = (processor or "1756-L83E").strip()
    # ProductCode for common ControlLogix Ethernet processors (Studio EKey)
    _proc_product = {
        "1756-L81E": "164",
        "1756-L82E": "165",
        "1756-L83E": "166",
        "1756-L84E": "167",
        "1756-L85E": "168",
    }
    if local_mod:
        local_mod = re.sub(
            r'CatalogNumber="[^"]*"',
            f'CatalogNumber="{_xml_escape(proc)}"',
            local_mod,
            count=1,
        )
        if proc in _proc_product:
            local_mod = re.sub(
                r'ProductCode="[^"]*"',
                f'ProductCode="{_proc_product[proc]}"',
                local_mod,
                count=1,
            )
        # Chassis backplane: controller is always Address 0 on ICP bus
        local_mod = re.sub(
            r'(<Port Id="1" Address=")[^"]*(" Type="ICP")',
            r'\g<1>0\2',
            local_mod,
            count=1,
        )
        # Room for EN2T + future local cards (gold EDITED uses 17)
        local_mod = re.sub(
            r'(<Port Id="1"[^>]*>\s*<Bus Size=")[^"]*("/>)',
            r'\g<1>17\2',
            local_mod,
            count=1,
            flags=re.S,
        )
        # Ensure Ethernet port is present (empty Bus is valid for onboard ENET)
        if not re.search(r'<Port Id="2"[^>]*Type="Ethernet"', local_mod):
            local_mod = local_mod.replace(
                "</Ports>",
                '<Port Id="2" Type="Ethernet" Upstream="false">\n<Bus/>\n</Port>\n</Ports>',
                1,
            )

    enet_mod = _extract_module_xml(library_text, enet_parent) or ""
    if enet_mod:
        # Slot 1 on local chassis (gold PLC5ENET1 style)
        enet_mod = re.sub(
            r'(<Port Id="1" Address=")[^"]*(" Type="ICP")',
            r'\g<1>1\2',
            enet_mod,
            count=1,
        )
        enet_mod = re.sub(
            r'ParentModule="[^"]*"',
            'ParentModule="Local"',
            enet_mod,
            count=1,
        )

    extra_mods: list[str] = []
    if topology:
        for ad in topology:
            rio_name = ad.get("rio_name") or ""
            ip = (ad.get("ip") or "192.168.1.50").strip()
            if not rio_name:
                continue
            kids = list(ad.get("children") or [])
            family = (ad.get("family") or "").strip() or _eip_family_from_types(
                [c.get("type") or "" for c in kids]
            )
            aent_tmpl = parent_tmpls.get(family) or parent_tmpls.get("1794")
            if not aent_tmpl:
                continue
            parent_lib_name = EIP_PARENT_TEMPLATE.get(family, "IO_1N90")
            bus_size = EIP_PARENT_BUS_SIZE.get(family, 8)
            parent_cat = EIP_PARENT_CATALOG.get(family, "1794-AENT")
            # Parent AENT — clone library template (1794 Flex or 1734 POINT)
            block = aent_tmpl
            block = block.replace(f'Name="{parent_lib_name}"', f'Name="{rio_name}"', 1)
            block = re.sub(
                r'ParentModule="[^"]*"',
                f'ParentModule="{enet_parent}"',
                block,
                count=1,
            )
            block = re.sub(
                r'(Address=")[\d.]+(" Type="Ethernet")',
                rf'\g<1>{_xml_escape(ip)}\2',
                block,
                count=1,
            )
            block = re.sub(
                r'CatalogNumber="[^"]*"',
                f'CatalogNumber="{_xml_escape(parent_cat)}"',
                block,
                count=1,
            )
            block = re.sub(
                r'(<Bus Size=")[^"]*("/>)',
                rf'\g<1>{bus_size}\2',
                block,
                count=1,
            )
            extra_mods.append(block)
            eip_module_names.append(rio_name)

            for c in kids:
                mt = (c.get("type") or "").strip()
                tmpl = child_tmpls.get(mt)
                if not tmpl:
                    # unknown card — skip (Studio needs a real template for AB: types)
                    continue
                cname = c.get("name") or f"{rio_name}_{c.get('flex_slot')}"
                flex = int(c.get("flex_slot") or 0)
                cblock = tmpl
                # Rename module + parent + port address (Flex or PointIO)
                cblock = re.sub(
                    r'Name="IO_1N(?:90|80)_\d+"',
                    f'Name="{cname}"',
                    cblock,
                    count=1,
                )
                cblock = re.sub(
                    r'ParentModule="[^"]*"',
                    f'ParentModule="{rio_name}"',
                    cblock,
                    count=1,
                )
                cblock = re.sub(
                    r'(<Port Id="1" Address=")[^"]*(" Type="(?:Flex|PointIO)")',
                    rf'\g<1>{flex}\2',
                    cblock,
                    count=1,
                )
                catalog = c.get("catalog") or EIP_CATALOG.get(mt, mt)
                cblock = re.sub(
                    r'CatalogNumber="[^"]*"',
                    f'CatalogNumber="{_xml_escape(catalog)}"',
                    cblock,
                    count=1,
                )
                extra_mods.append(cblock)
                eip_child_names.append(cname)

    # Assemble Modules block: Local + EN2T + RIO tree only
    mod_parts: list[str] = []
    if local_mod:
        mod_parts.append(local_mod)
    if enet_mod:
        mod_parts.append(enet_mod)
    mod_parts.extend(extra_mods)
    if mod_parts:
        modules_block = "<Modules>\n" + "\n".join(mod_parts) + "\n</Modules>"
    else:
        # Fallback: library Modules wholesale (last resort)
        lib_modules = re.search(r"<Modules>.*?</Modules>", library_text, re.S)
        modules_block = lib_modules.group(0) if lib_modules else "<Modules/>"
        if processor:
            modules_block = re.sub(
                r'(<Module Name="Local"[^>]*CatalogNumber=")[^"]*"',
                rf'\g<1>{_xml_escape(processor)}"',
                modules_block,
                count=1,
            )

    dt_xml = datatypes.group(0) if datatypes else "<DataTypes/>"
    # Start from full library AOIs; later prune to only AOIs actually called so
    # unused sealed TRK_* etc. don't fail Studio verify (Invalid signature ID).
    aoi_xml = aois.group(0) if aois else "<AddOnInstructionDefinitions/>"
    # Prefer side-loaded exports (Slow_Flt_AOI.L5X from source-key re-export)
    aoi_xml = _overlay_aoi_exports(aoi_xml)

    # AOI prune happens after programs/tags are known (see below).

    # Merge extra DataTypes / AOIs from gold program exports (skip names already present)
    def _merge_named_blocks(host: str, chunks: list[str], wrapper: str, item_tag: str) -> str:
        if not chunks:
            return host
        existing = set(re.findall(rf'<{item_tag}\s+Name="([^"]+)"', host))
        extras: list[str] = []
        for ch in chunks:
            for bm in re.finditer(
                rf'<{item_tag}\s+Name="([^"]+)"[^>]*>.*?</{item_tag}>', ch, re.S
            ):
                if bm.group(1) in existing:
                    continue
                existing.add(bm.group(1))
                extras.append(bm.group(0))
            # self-closing rare
            for bm in re.finditer(rf'<{item_tag}\s+Name="([^"]+)"[^>]*/>', ch):
                if bm.group(1) in existing:
                    continue
                existing.add(bm.group(1))
                extras.append(bm.group(0))
        if not extras:
            return host
        if host.rstrip().endswith(f"</{wrapper}>"):
            return host.rstrip()[: -len(f"</{wrapper}>")] + "\n" + "\n".join(extras) + f"\n</{wrapper}>"
        return host

    dt_xml = _merge_named_blocks(dt_xml, extra_dt_chunks, "DataTypes", "DataType")
    aoi_xml = _merge_named_blocks(
        aoi_xml, extra_aoi_chunks, "AddOnInstructionDefinitions", "AddOnInstructionDefinition"
    )

    # CommDiag_UDT lives in gold PLC2, not always in OReilly_Library — inject when System pack used
    if "CommDiag_UDT" not in dt_xml:
        gold_txt = _load_gold_plc2_text()
        m_cd = re.search(
            r'<DataType Name="CommDiag_UDT"[^>]*>.*?</DataType>',
            gold_txt or "",
            re.S,
        )
        if m_cd and dt_xml.rstrip().endswith("</DataTypes>"):
            dt_xml = dt_xml.rstrip()[: -len("</DataTypes>")] + "\n" + m_cd.group(0) + "\n</DataTypes>"

    # Shorten only unsealed AOI Description text. Never rewrite EncodedData.
    aoi_xml = _shorten_aoi_descriptions(aoi_xml)

    # Keep only AOIs actually used in rungs/tags — drops unused sealed TRK_* etc.
    # (Studio verifies every AOI def; unused sealed ones still throw Invalid signature.)
    _aoi_lib_names = set(
        re.findall(
            r'(?:EncodedData EncodedType="AddOnInstructionDefinition"|AddOnInstructionDefinition)'
            r'[^>]*Name="([^"]+)"',
            aoi_xml,
        )
    )
    _used = set()
    for _chunk in list(programs_xml) + list(all_tags):
        _used.update(re.findall(r"\b([A-Za-z][A-Za-z0-9_]{2,60})\s*\(", _chunk))
        _used.update(re.findall(r'DataType="([^"]+)"', _chunk))
    _keep = _aoi_lib_names & _used
    # Always keep core transport AOIs if present (safety net)
    for _core in ("Fast_Conv", "Slow_Flt", "Slow_Jam", "PE_Logic", "Full_PE", "Merge_2to1"):
        if _core in _aoi_lib_names:
            _keep.add(_core)
    if _keep:
        before_n = len(_aoi_lib_names)
        aoi_xml = _filter_aois_to_used(aoi_xml, _keep)
        _emit_progress(
            f"AOIs kept {len(_keep)}/{before_n} (dropped unused sealed defs)",
            88,
        )
        # Drop UDTs that embed removed AOIs (e.g. Track_Divert_AOI → TRK_Divert)
        def _strip_udts_for_missing_aois(dt_block: str, keep_aois: set[str]) -> str:
            out_parts: list[str] = []
            # Keep opening wrapper
            m_wrap = re.match(r"(<DataTypes\b[^>]*>)", dt_block)
            head = m_wrap.group(1) if m_wrap else "<DataTypes>"
            for dm in re.finditer(
                r'<DataType\s+Name="([^"]+)"[^>]*>.*?</DataType>', dt_block, re.S
            ):
                body = dm.group(0)
                # Member DataType references to AOIs not in keep → drop whole UDT
                member_types = set(re.findall(r'DataType="([^"]+)"', body))
                # Skip the DataType's own Name attribute match by checking Members only
                mem_section = re.search(r"<Members>(.*?)</Members>", body, re.S)
                if mem_section:
                    member_types = set(
                        re.findall(r'DataType="([^"]+)"', mem_section.group(1))
                    )
                else:
                    member_types = set()
                bad = [
                    mt for mt in member_types
                    if mt in _aoi_lib_names and mt not in keep_aois
                ]
                if bad:
                    continue
                # Also drop obvious sorter track UDTs when no TRK AOIs kept
                name = dm.group(1)
                if not any(a.startswith("TRK_") for a in keep_aois):
                    if name.startswith("Track_") or name.startswith("TRK"):
                        continue
                out_parts.append(body)
            return head + "\n" + "\n".join(out_parts) + "\n</DataTypes>"

        dt_xml = _strip_udts_for_missing_aois(dt_xml, _keep)

    prog_names = []
    for p in programs_xml:
        m = re.search(r'<Program\s+Name="([^"]+)"', p)
        if m:
            prog_names.append(m.group(1))

    # Own tasks — match finished gold: P02_Track = IO_MAP + Sorter_Track; Sys alone.
    # Sys_Comm (gold System) rides P11_Slow_200ms with area Slow programs.
    # WCS / ShippingSorter L3 stay on P11 Slow until area packs are site-wired.
    own_task_programs = {"IO_MAP", "Sys", "Sorter_Track"}
    optional_slow = {
        "WCS_Interface_TCP_IP", "ShippingSorter_Area_L3", "System", "Sys_Comm",
    }
    # Also catch any gold program names that aren't Sys/IO_MAP/Sorter_Track
    for gn in gold_program_names:
        if gn not in own_task_programs:
            optional_slow.add(gn)

    slow_names = [
        n for n in prog_names
        if n not in own_task_programs
        and (n.endswith("_Slow") or n in optional_slow)
    ]
    fast_names = [n for n in prog_names if n.endswith("_Fast")]
    l1_names = [n for n in prog_names if n.endswith("_L1")]
    l2_names = [n for n in prog_names if n.endswith("_L2")]
    if not slow_names:
        slow_names = [
            n for n in prog_names
            if n not in fast_names
            and n not in own_task_programs
            and n not in l1_names
            and n not in l2_names
        ]
    sched_slow = "".join(f'<ScheduledProgram Name="{_xml_escape(n)}"/>' for n in slow_names)
    sched_fast = "".join(f'<ScheduledProgram Name="{_xml_escape(n)}"/>' for n in fast_names)
    sched_l1 = "".join(f'<ScheduledProgram Name="{_xml_escape(n)}"/>' for n in l1_names)
    sched_l2 = "".join(f'<ScheduledProgram Name="{_xml_escape(n)}"/>' for n in l2_names)

    def _task_xml(name: str, rate: int, priority: int, watchdog: int, programs: list[str]) -> str:
        if not programs:
            return ""
        sched = "".join(f'<ScheduledProgram Name="{_xml_escape(n)}"/>' for n in programs)
        return (
            f'<Task Name="{_xml_escape(name)}" Type="PERIODIC" Rate="{rate}" '
            f'Priority="{priority}" Watchdog="{watchdog}" '
            f'DisableUpdateOutputs="false" InhibitTask="false">\n'
            f'<ScheduledPrograms>\n{sched}\n</ScheduledPrograms>\n'
            f'</Task>\n'
        )

    # Controller SFC attrs + element order match library / Rockwell schema.
    # Open this L5X as a NEW Studio project (not Import into existing .acd).
    minor_i = int(minor) if str(minor).isdigit() else 0
    tags_block = "<Tags>\n" + "".join(all_tags) + "\n</Tags>"
    programs_block = "<Programs>\n" + "".join(programs_xml) + "\n</Programs>"
    tasks_block = '<Tasks>\n'
    # P02_Track (gold finished): IO_MAP + optional Sorter_Track on same task
    p02_progs = [n for n in ("IO_MAP", "Sorter_Track") if n in prog_names]
    if p02_progs:
        tasks_block += _task_xml(
            "P02_Track_10ms", rate=10, priority=2, watchdog=50, programs=p02_progs
        )
    elif "IO_MAP" in prog_names:
        tasks_block += _task_xml("IO_MAP", rate=20, priority=12, watchdog=100, programs=["IO_MAP"])
    if "Sys" in prog_names:
        # Gold: P15_Config_Sys_Event — keep periodic Sys task for constants pack
        tasks_block += _task_xml("Sys", rate=100, priority=13, watchdog=250, programs=["Sys"])
    if sched_fast:
        tasks_block += (
            '<Task Name="P10_Fast_50ms" Type="PERIODIC" Rate="50" Priority="10" '
            'Watchdog="500" DisableUpdateOutputs="false" InhibitTask="false">\n'
            f'<ScheduledPrograms>\n{sched_fast}\n</ScheduledPrograms>\n'
            '</Task>\n'
        )
    if sched_slow:
        # Includes area *_Slow + Sys_Comm (gold System)
        tasks_block += (
            '<Task Name="P11_Slow_200ms" Type="PERIODIC" Rate="200" Priority="11" '
            'Watchdog="500" DisableUpdateOutputs="false" InhibitTask="false">\n'
            f'<ScheduledPrograms>\n{sched_slow}\n</ScheduledPrograms>\n'
            '</Task>\n'
        )
    if sched_l1:
        tasks_block += (
            '<Task Name="P15_Config_L1_Event" Type="EVENT" Rate="60000" Priority="15" '
            'Watchdog="2000" DisableUpdateOutputs="true" InhibitTask="false">\n'
            '<EventInfo EventTrigger="EVENT Instruction Only" EnableTimeout="false"/>\n'
            f'<ScheduledPrograms>\n{sched_l1}\n</ScheduledPrograms>\n'
            '</Task>\n'
        )
    if sched_l2:
        tasks_block += (
            '<Task Name="P15_Config_L2_Event" Type="EVENT" Rate="60000" Priority="15" '
            'Watchdog="2000" DisableUpdateOutputs="true" InhibitTask="false">\n'
            '<EventInfo EventTrigger="EVENT Instruction Only" EnableTimeout="false"/>\n'
            f'<ScheduledPrograms>\n{sched_l2}\n</ScheduledPrograms>\n'
            '</Task>\n'
        )
    tasks_block += '</Tasks>'
    # Prefer library SoftwareRevision when present (helps sealed AOI host match)
    lib_sw = re.search(r'SoftwareRevision="([^"]+)"', library_text)
    sw_rev = lib_sw.group(1) if lib_sw else f"{major}.{minor}"
    l5x = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<RSLogix5000Content SchemaRevision="1.0" SoftwareRevision="{sw_rev}" TargetName="{_xml_escape(proj)}" TargetType="Controller" ContainsContext="false" Owner="SiteForge" ExportDate="{stamp}" ExportOptions="NoRawData L5KData DecoratedData ForceProtectedEncoding AllProjDocTrans">
<Controller Use="Target" Name="{_xml_escape(proj)}" ProcessorType="{_xml_escape(processor)}" MajorRev="{major}" MinorRev="{minor_i}" ProjectCreationDate="{stamp}" LastModifiedDate="{stamp}" SFCExecutionControl="CurrentActive" SFCRestartPosition="MostRecent" SFCLastScan="DontScan" ProjectSN="16#0000_0000" MatchProjectToController="false" CanUseRPIFromProducer="false" InhibitAutomaticFirmwareUpdate="0" PassThroughConfiguration="EnabledWithAppend" DownloadProjectDocumentationAndExtendedProperties="true" DownloadProjectCustomProperties="true" ReportMinorOverflow="false">
<RedundancyInfo Enabled="false" KeepTestEditsOnSwitchOver="false"/>
<Security Code="0" ChangesToDetect="16#ffff_ffff_ffff_ffff"/>
<SafetyInfo/>
{dt_xml}
{modules_block}
{aoi_xml}
{tags_block}
{programs_block}
{tasks_block}
<CST/>
<WallClockTime/>
<Trends/>
<DataLogs/>
<TimeSynchronize/>
<EthernetPorts/>
</Controller>
</RSLogix5000Content>
'''

    # Slim report for Electron IPC (full conveyor lists live in autogen_input.json on disk)
    pe_with_real = sum(1 for c in cloned if any(p != "NO_PE" for p in (c.get("jam_pes") or [])))
    report = {
        "project": proj,
        "processor": processor,
        "revision": f"{major}.{minor}",
        "library": str(library_path),
        "conveyor_count": len(cloned),
        "area_count": len(by_area),
        "tag_count": len(all_tags),
        "program_count": len(programs_xml),
        "programs": prog_names,
        "areas_summary": {a: len(items) for a, items in by_area.items()},
        "conveyor_sample": [x["conveyor"] for x in cloned[:25]],
        "template_usage": {},
        "missing_excel_templates_in_library": sorted(missing_templates),
        "io_module_count": len(inp.modules),
        "io_point_count": len(inp.io_points),
        "io_tags_in_l5x": len(io_tag_rows),
        "pe_device_count": len(getattr(inp, "pe_devices", None) or []),
        "pe_logic_rungs": pe_wired_count,
        "slow_flt_rungs": flt_count,
        "conveyors_with_real_pe": pe_with_real,
        "eip_adapter_count": len(getattr(inp, "eip_topology", None) or []),
        "eip_rio_modules": eip_module_names,
        "eip_child_modules": eip_child_names,
        "eip_child_count": len(eip_child_names),
        "eip_interface_ip": getattr(inp, "eip_interface_ip", "") or "",
        "io_word_map_count": len(getattr(inp, "io_word_map", None) or {}),
        "io_map_rungs": max(0, len(cp_i_rungs) + len(cp_o_rungs) - 2),
        "io_map_mapped": io_map_mapped,
        "io_map_unmapped": io_map_unmapped,
        "io_map_source": (
            "gold_program_excel"
            if gold_io_map_used
            else ("run_tar_gz_banks_eip" if want_io_map else "omitted")
        ),
        "io_map_note": (
            "Gold Excel IO_MAP (library) — Greensboro CP5/CP6/CP7 only"
            if gold_io_map_used
            else (
                "Omitted — IO_MAP checkbox was off"
                if not want_io_map
                else (
                    "Built from RUN Conveyor.asc banks + EIP word_map → CPxRIOn modules"
                    + (
                        " (gold Excel blocked — site is CP1–CP4)"
                        if gold_io_blocked
                        else ""
                    )
                )
            )
        ),
        "include_io_map": want_io_map or gold_io_map_used,
        "gold_io_map_blocked": gold_io_blocked,
        "eip_modules_filtered_to_word_map": sorted(used_rios) if used_rios else [],
        "gold_programs": gold_program_names,
        "sorter_build": sorter_report,
        "equipment_plan": getattr(inp, "equipment_plan", None) or {},
        "optional_programs_available": list(OPTIONAL_PROGRAMS.keys()),
        "task_schedule": {
            "P02_Track_10ms": "IO_MAP + Sorter_Track (live from Sorter build UI)",
            "Sys": "Task Sys @ 100ms (own task)",
            "P10_Fast_20ms": fast_names,
            "P11_Slow_200ms": slow_names,
        },
        "encoded_aois_stripped": False,
        "logic": "Fast_Conv + Slow_Jam + PE_Logic/Full_PE + Slow_Flt + IO_MAP (module:I.Data)",
        "note": (
            "Site Forge Python autogen from RUN: real PE tags (NO_PE), PE_Logic/Full_PE, Slow_Flt, "
            "full Flex I/O tree CPxRIOn + CPxRIOn_k children with AB: data types "
            "(DI_Delay16/DO8/IB16/DO16), IO_MAP XIC(CPxRIOn:I.Data[s].b)OTE(PE.I.PE_Clear) "
            "via EIPCSV Word map + octal bits. "
            "OPEN as NEW Studio project (File→Open L5X)."
        ),
        "io_tag_rows": io_tag_rows,
    }
    for item in cloned:
        t = item["template"]
        report["template_usage"][t] = report["template_usage"].get(t, 0) + 1

    # Final pass: unsealed AOI descriptions only (never EncodedData / sealed bodies)
    l5x = _shorten_aoi_descriptions(l5x)
    # Last chance: any leftover gold Greensboro names → this site
    l5x = _retarget_gold_site_names(l5x, site_stem)
    return l5x, report


def _emit_progress(message: str, pct: int = 0, **extra) -> None:
    """Progress for Electron (stderr) — never pollute stdout JSON."""
    payload = {"phase": "autogen", "message": message, "pct": pct, **extra}
    try:
        print(f"FORTNA_PROGRESS {json.dumps(payload, separators=(',', ':'))}", file=sys.stderr, flush=True)
    except Exception:
        pass


def _default_twin_peers(site: str) -> list[dict]:
    """Known peer sites for controller-upgrade similarity search."""
    s = (site or "").upper()
    peers = [
        {
            "site": "ORLY_Greensboro_NC_PLC2",
            "role": "gold_transport",
            "note": "Greensboro PLC2 — Merge_2to1 + ModuleB Fast/Slow pattern",
        },
        {"site": "MSCRENO_MSCRENOSHIP", "role": "sibling", "note": "Reno Ship"},
        {"site": "MSCRENO_MSCRENOPACK", "role": "sibling", "note": "Reno Pack"},
        {"site": "MSCRENO_MSCRENOPICK", "role": "sibling", "note": "Reno Pick"},
    ]
    # Don't list self as a peer
    return [p for p in peers if p["site"].upper() not in s and s not in p["site"].upper()]


def build_twin_gaps(inp: AutogenInput, report: dict | None = None) -> dict:
    """Structured gaps Autogen left for PRISM / AI gap-fill (Phase 0)."""
    report = report or {}
    pe_known: set[str] = set()
    for pe in getattr(inp, "pe_devices", None) or []:
        if isinstance(pe, dict) and pe.get("name"):
            pe_known.add(_safe(pe["name"]))
        elif isinstance(pe, dict) and pe.get("fortna_name"):
            pe_known.add(_safe(pe["fortna_name"]))
    for p in getattr(inp, "io_points", None) or []:
        if (getattr(p, "device_type", "") or "").lower() == "photoeye":
            pe_known.add(_safe(getattr(p, "device_name", "") or ""))
    pe_known.discard("")
    pe_known.discard("NO_PE")

    stubs = {
        _safe(t)
        for t in (getattr(inp, "transport_stub_tags", None) or [])
        if str(t).strip()
    }
    gaps: list[dict] = []
    for tag in sorted(stubs):
        gaps.append(
            {
                "type": "transport_stub_not_in_run",
                "severity": "info",
                "conveyor": tag,
                "message": f"{tag} bound in Transport Build but not in active RUN — stub Fast/Slow emitted",
            }
        )

    def _pe_name(raw: str) -> str:
        t = (raw or "").strip()
        if not t or t.upper() in ("NO_PE", "NONE", "-", "N/A"):
            return ""
        return _safe(t)

    for c in inp.conveyors or []:
        cname = _safe(c.clean_name or c.conveyor)
        area = (c.main_area or "").strip()
        exit_pe = _pe_name(getattr(c, "exit_pe_tag", "") or "")
        add_pe = _pe_name(getattr(c, "add_pe_tag", "") or "")
        if not exit_pe:
            gaps.append(
                {
                    "type": "missing_exit_pe",
                    "severity": "warn",
                    "conveyor": cname,
                    "area": area,
                    "message": f"{cname} has no Exit PE (Fast_Conv uses NO_PE)",
                }
            )
        pe_roles = [(exit_pe, "exit"), (add_pe, "add")]
        for t in getattr(c, "jam_pe_tags", None) or []:
            pe_roles.append((_pe_name(str(t)), "jam"))
        for t in getattr(c, "full_pe_tags", None) or []:
            pe_roles.append((_pe_name(str(t)), "full"))
        for pe, role in pe_roles:
            if not pe:
                continue
            if pe not in pe_known:
                gaps.append(
                    {
                        "type": "pe_not_in_run_io",
                        "severity": "warn",
                        "conveyor": cname,
                        "area": area,
                        "pe": pe,
                        "role": role,
                        "message": f"{pe} ({role}) on {cname} has no RUN photoeye/IO bank — no IO_MAP bit",
                    }
                )

    for m in getattr(inp, "merges_2to1", None) or []:
        if not isinstance(m, dict):
            continue
        name = _safe(m.get("name") or m.get("discharge") or "")
        for field in ("pe_a", "pe_b", "pe_c", "jam_pe"):
            raw = (m.get(field) or "").strip()
            if raw and raw.upper() not in ("NO_PE", "NONE", "-"):
                pe = _safe(raw)
                if pe and pe not in pe_known:
                    gaps.append(
                        {
                            "type": "merge_pe_not_in_run_io",
                            "severity": "warn",
                            "merge": name,
                            "field": field,
                            "pe": pe,
                            "message": f"Merge {name} {field}={pe} not in RUN IO",
                        }
                    )
            elif field in ("pe_a", "pe_b") and not raw:
                gaps.append(
                    {
                        "type": "merge_pe_blank",
                        "severity": "info",
                        "merge": name,
                        "field": field,
                        "message": f"Merge {name} {field} blank → NO_PE at emit",
                    }
                )

    by_type: dict[str, int] = {}
    for g in gaps:
        by_type[g["type"]] = by_type.get(g["type"], 0) + 1

    return {
        "kind": "fortna_twin_gaps",
        "version": 1,
        "generated_utc": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "project": getattr(inp, "project_name", "") or "",
        "areas": list(getattr(inp, "areas", None) or []),
        "conveyor_count": len(inp.conveyors or []),
        "merge_count": len(getattr(inp, "merges_2to1", None) or []),
        "gap_count": len(gaps),
        "by_type": by_type,
        "gaps": gaps,
        "note": (
            "Gaps Autogen left for PRISM site-twin / AI propose-patches. "
            "Does not rewrite L5X — engineer Apply only."
        ),
        "report_hints": {
            "areas_summary": (report or {}).get("areas_summary"),
            "io_map_unmapped": (report or {}).get("io_map_unmapped"),
            "conveyors_with_real_pe": (report or {}).get("conveyors_with_real_pe"),
        },
    }


def generate(
    inp: AutogenInput,
    library: Path,
    out_dir: Path | None = None,
) -> dict:
    if not library.is_file():
        raise FileNotFoundError(f"Library L5X not found: {library}")
    # Folder: timestamp + archive label (track history).
    # L5X file + Controller name: site + panel ONLY (no date) — Studio rejects long dated names.
    try:
        from fortna_source_id import (
            export_label_from_meta,
            safe_fs_name,
            studio_project_stem,
        )
        export_label = export_label_from_meta()
        folder_stem = safe_fs_name(export_label) if export_label else safe_fs_name(inp.project_name)
        # OReillyDC27_ORDENCP4 — never 20260803_0815_…
        file_stem = studio_project_stem(inp.project_name, getattr(inp, "machine", "") or "")
        if not file_stem or file_stem == "Autogen_Project":
            file_stem = studio_project_stem(inp.project_name, "")
    except Exception:
        export_label = ""
        folder_stem = _safe(inp.project_name) or "Autogen_Project"
        file_stem = _safe(inp.project_name) or "Autogen_Project"
        # strip date if present
        file_stem = re.sub(r"^\d{8}_?\d{0,6}_?", "", file_stem).strip("_") or "Autogen_Project"
    if not file_stem:
        file_stem = _safe(inp.project_name) or "Autogen_Project"
    if not folder_stem:
        folder_stem = file_stem
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    # Always unique folder per run (stamp + site) so exports don't overwrite.
    folder = f"{stamp}-{folder_stem}"
    out = out_dir or (REPO_ROOT / "exports" / "autogen" / folder)
    out.mkdir(parents=True, exist_ok=True)

    # Force short controller name inside L5X (matches file stem)
    try:
        inp.project_name = file_stem
    except Exception:
        pass

    _emit_progress(f"Building L5X for {len(inp.conveyors)} conveyors…", 20, conveyor_count=len(inp.conveyors))
    l5x, report = build_l5x(inp, library)
    l5x_path = out / f"{file_stem}.L5X"
    _emit_progress("Writing L5X file…", 70)
    l5x_path.write_text(l5x, encoding="utf-8")

    # Studio 5000 Tools → Import tag CSV (controller scope) — separate from L5X
    studio_csv_path = out / f"{file_stem}_Controller_Tags.csv"
    csv_count = 0
    try:
        from fortna_plc_export import write_studio_tags_csv

        io_rows = report.pop("io_tag_rows", []) or []
        # Also list conveyor tags for review
        for c in inp.conveyors:
            n = _safe(c.clean_name)
            if not n:
                continue
            io_rows.append({
                "tag": f"{n}_Conv",
                "fortna_name": c.conveyor,
                "fortna_address": c.main_area or "",
                "description": f"Conveyor UDT placeholder {c.type}",
                "type": "BOOL",  # CSV helper is BOOL-oriented; L5X has real Conv_UDT
                "device_class": "conveyor",
            })
        csv_count = write_studio_tags_csv(
            io_rows,
            studio_csv_path,
            controller_context=file_stem,
            software_version=f"Studio 5000 v{inp.major_rev}.{inp.minor_rev}",
        )
        report["studio_tags_csv"] = str(studio_csv_path)
        report["studio_tags_csv_count"] = csv_count
    except Exception as exc:
        report.pop("io_tag_rows", None)
        report["studio_tags_csv_error"] = str(exc)

    # Save input snapshot for audit
    snap = {
        "project_name": inp.project_name,
        "processor": inp.processor,
        "major_rev": inp.major_rev,
        "minor_rev": inp.minor_rev,
        "areas": inp.areas,
        "safety_zones": inp.safety_zones,
        "conveyors": [asdict(c) for c in inp.conveyors],
        "modules": [asdict(m) for m in inp.modules],
        "io_points": [asdict(p) for p in inp.io_points],
        "pe_devices": list(getattr(inp, "pe_devices", None) or []),
        "eip_adapters": list(getattr(inp, "eip_adapters", None) or []),
        "eip_topology": list(getattr(inp, "eip_topology", None) or []),
        "io_word_map": dict(getattr(inp, "io_word_map", None) or {}),
        "eip_interface_ip": getattr(inp, "eip_interface_ip", "") or "",
        "include_programs": list(getattr(inp, "include_programs", None) or []),
        "include_sys": bool(getattr(inp, "include_sys", True)),
        "include_io_map_gold": bool(getattr(inp, "include_io_map_gold", False)),
        "merges_2to1": list(getattr(inp, "merges_2to1", None) or []),
        "engine": "fortna_autogen.py (Python — RUN/tar.gz primary; Excel gold optional)",
    }
    (out / "autogen_input.json").write_text(json.dumps(snap, indent=2), encoding="utf-8")
    # Tar equipment → build plan (AOIs/packs correlated to gold PLC2/4/5)
    try:
        ep = getattr(inp, "equipment_plan", None) or {}
        if ep:
            (out / "equipment_plan.json").write_text(
                json.dumps(ep, indent=2), encoding="utf-8"
            )
    except Exception:
        pass

    # Physical I/O verification: Fortna Word.Bit → CPxRIOn:I/O.Data[s].b
    # Sources: Conveyor.asc banks + EIPCSV/EIPModules word_map (same as eipcfg/ASC in RUN).
    try:
        word_map = dict(getattr(inp, "io_word_map", None) or {})
        topo = list(getattr(inp, "eip_topology", None) or [])

        # RIO inventory for Studio module tree check
        rio_inv = {
            "note": (
                "Remote I/O from RUN PROJECT/EIPAdapters + EIPModules + EIPCSV "
                "(and FORTNA/*eipcfg.xml) — 1734 POINT or 1794 Flex per tar. "
                "One RUN = one controller's network. "
                "CP1/CP4 panels need their own RUN if they are separate PLCs."
            ),
            "interface_ip": getattr(inp, "eip_interface_ip", "") or "",
            "adapters": topo,
            "word_map_count": len(word_map),
            "rio_names": [t.get("rio_name") for t in topo],
            "module_count": sum(len(t.get("children") or []) for t in topo),
        }
        (out / "rio_inventory.json").write_text(
            json.dumps(rio_inv, indent=2), encoding="utf-8"
        )
        report["rio_inventory"] = str(out / "rio_inventory.json")
        report["rio_module_count"] = rio_inv["module_count"]
        report["rio_names"] = rio_inv["rio_names"]

        bank_index_phys = _build_eip_bank_index(topo)
        configio_phys = dict(getattr(inp, "configio_octal_map", None) or {})

        def _resolve_physical(bank: str, bit: str, *, want_dir: str = "") -> tuple[str, str]:
            """Return (module_data_ref, note). Prefer Configio when EIPCSV empty."""
            w = str(bank or "").strip()
            b = str(bit or "").strip()
            info = None
            if want_dir:
                info = _resolve_fortna_bank(
                    w,
                    want_dir=want_dir,
                    bank_index=bank_index_phys,
                    bit=b,
                    configio_map=configio_phys,
                )
            if not info:
                info = word_map.get(w)
            if not info:
                try:
                    info = word_map.get(str(int(float(w))))
                except Exception:
                    info = None
            if not info and w.isdigit() and int(w) % 2 == 1:
                info = word_map.get(str(int(w) - 1))
            if not info:
                return "", (
                    f"UNMAPPED word {w} — not in Configio/EIPCSV "
                    f"(other panel/PLC bank, or missing module)"
                )
            mod_type = (info.get("type") or "")
            family = info.get("family") or (
                "1734" if "1734" in mod_type or "1738" in mod_type else "1794"
            )
            max_bit = _point_card_max_bit(mod_type) if family == "1734" else 15
            bit_for_card = b
            if _fortna_bit_is_high(b) and (info.get("resolve_how") or "") == "configio":
                try:
                    hv = int(str(b).strip(), 8)
                except ValueError:
                    try:
                        hv = int(str(b).strip(), 10)
                    except ValueError:
                        hv = -1
                if hv >= 8:
                    bit_for_card = str(hv - 8)
            data_bit = _fortna_bit_to_data_bit(bit_for_card, max_bit=max_bit)
            if data_bit is None or data_bit < 0:
                return "", f"bad bit {b} for word {w}"
            rio = info.get("rio_name") or ""
            slot = int(info.get("flex_slot") or 0)
            direction = (info.get("direction") or want_dir or "I").upper()
            ref = f"{rio}:{slot}:{direction}.Data.{data_bit}"
            alt = f"{rio}:{direction}.Data[{slot}].{data_bit}"
            how = info.get("resolve_how") or "map"
            note = f"alt={alt}; type={info.get('type')}; word={w}"
            if info.get("resolved_bank"):
                note += f"; eip={info.get('resolved_bank')}"
            if how not in ("direct", "map"):
                note += f"; via={how}"
            return ref, note

        map_lines = [
            "fortna_name,device_type,direction,fortna_bank,fortna_bit,"
            "module_data_ref,mapped,notes"
        ]
        mapped_n = 0
        unmapped_n = 0
        # PE devices
        for p in getattr(inp, "pe_devices", None) or []:
            ref, note = _resolve_physical(
                str(p.get("bank") or ""), str(p.get("bit") or ""), want_dir="I"
            )
            ok = "Y" if ref else "N"
            if ref:
                mapped_n += 1
            else:
                unmapped_n += 1
            map_lines.append(
                ",".join(
                    [
                        json.dumps(p.get("fortna_name") or p.get("name") or ""),
                        json.dumps("photoeye"),
                        json.dumps("I"),
                        json.dumps(str(p.get("bank") or "")),
                        json.dumps(str(p.get("bit") or "")),
                        json.dumps(ref),
                        ok,
                        json.dumps(note),
                    ]
                )
            )
        # Other IO points from io_points (beacons, encoders, motors, …)
        seen_names = {
            (p.get("fortna_name") or p.get("name") or "").upper()
            for p in (getattr(inp, "pe_devices", None) or [])
        }
        for p in getattr(inp, "io_points", None) or []:
            dname = (getattr(p, "device_name", None) or "").strip()
            if not dname or dname.upper() in seen_names:
                continue
            dtype = (getattr(p, "device_type", None) or getattr(p, "io_type", None) or "").strip()
            bank = str(getattr(p, "fortna_bank", None) or "")
            bit = str(getattr(p, "fortna_bit", None) or "")
            if not bank and not bit:
                continue
            # Prefer name/type rules (_AUX→I, ENC→I, beacon→O) over stale ASC io_type
            direction = _io_point_want_dir(
                dname, dtype, getattr(p, "direction", None) or ""
            )
            ref, note = _resolve_physical(bank, bit, want_dir=direction)
            ok = "Y" if ref else "N"
            if ref:
                mapped_n += 1
            else:
                unmapped_n += 1
            map_lines.append(
                ",".join(
                    [
                        json.dumps(dname),
                        json.dumps(dtype or "io"),
                        json.dumps(direction),
                        json.dumps(bank),
                        json.dumps(bit),
                        json.dumps(ref),
                        ok,
                        json.dumps(note),
                    ]
                )
            )
        (out / "physical_io_map.csv").write_text("\n".join(map_lines), encoding="utf-8")
        # Keep legacy filename for older dashboards
        (out / "io_map_pending.csv").write_text("\n".join(map_lines), encoding="utf-8")
        report["physical_io_map_csv"] = str(out / "physical_io_map.csv")
        report["physical_io_mapped"] = mapped_n
        report["physical_io_unmapped"] = unmapped_n
        report["io_map_pending_csv"] = str(out / "io_map_pending.csv")
    except Exception as exc:
        report["physical_io_map_error"] = str(exc)
        report["io_map_pending_csv_error"] = str(exc)

    # Drop bulky rows from on-disk report if still present
    report.pop("io_tag_rows", None)
    (out / "autogen_report.json").write_text(json.dumps(report, indent=2), encoding="utf-8")

    # Human report
    lines = [
        f"Site Forge PLC Autogen (Python) — {stamp}",
        f"Project: {inp.project_name}",
        f"Processor: {inp.processor}  v{inp.major_rev}.{inp.minor_rev}",
        f"Library: {library.name}",
        f"Conveyors: {report['conveyor_count']}  (with real PE: {report.get('conveyors_with_real_pe', 0)}) "
        f"— scoped to this master PLC only",
        f"Areas: {report['area_count']}",
        f"Tags: {report['tag_count']}",
        f"PE devices: {report.get('pe_device_count', 0)}  PE_Logic/Full_PE rungs: {report.get('pe_logic_rungs', 0)}",
        f"Slow_Flt rungs: {report.get('slow_flt_rungs', 0)}",
        f"EIP adapters: {report.get('eip_adapter_count', 0)}  RIO: {report.get('eip_rio_modules', [])}",
        f"EIP child cards: {report.get('eip_child_count', 0)} (AB: typed C/I/O tags)",
        f"IO_MAP source: {report.get('io_map_source')} — {report.get('io_map_note', '')}",
        f"IO_MAP: {report.get('io_map_mapped', 0)} mapped, {report.get('io_map_unmapped', 0)} unmapped "
        f"(word map entries: {report.get('io_word_map_count', 0)})",
        f"RIO modules: {report.get('rio_module_count', 0)}  "
        f"names: {', '.join(report.get('rio_names') or [])}",
        f"Physical map: {report.get('physical_io_mapped', 0)} mapped / "
        f"{report.get('physical_io_unmapped', 0)} unmapped → physical_io_map.csv",
        f"Studio tag CSV: {studio_csv_path.name if csv_count else '(none)'} ({csv_count} rows)",
        f"Programs: {', '.join(report['programs'])}",
        f"Templates: {report['template_usage']}",
        f"Output: {l5x_path}",
        "",
        report["note"],
        "Engine: fortna_autogen.py — IO_MAP from RUN/tar.gz banks + EIP modules by default.",
        "Gold Excel IO_MAP is optional (--io-map-gold / UI checkbox); it replaces RUN mapping.",
        "Logic: Fast_Conv + Slow_Jam + PE_Logic/Full_PE + Slow_Flt + IO_MAP CP_I/CP_O.",
        "I/O tree: RUN EIPAdapters/EIPModules/EIPCSV + eipcfg.xml → CPxRIOn (1734 POINT or 1794 Flex).",
        "Verify: open physical_io_map.csv + rio_inventory.json next to the L5X.",
        "OPEN as NEW project in Studio (File→Open). Do not Import into existing .acd.",
    ]
    if report["missing_excel_templates_in_library"]:
        lines.append(
            "Excel templates not in library (used fallbacks): "
            + ", ".join(report["missing_excel_templates_in_library"])
        )
    (out / "autogen_report.txt").write_text("\n".join(lines), encoding="utf-8")

    # PRISM site-twin gaps (Phase 0) — what Autogen could not fully wire
    twin_gaps = build_twin_gaps(inp, report)
    try:
        (out / "twin_gaps.json").write_text(
            json.dumps(twin_gaps, indent=2), encoding="utf-8"
        )
    except Exception:
        pass
    report["twin_gaps"] = {
        "count": twin_gaps.get("gap_count") or 0,
        "by_type": twin_gaps.get("by_type") or {},
    }

    # Copy library alongside for Studio import reference (best-effort; OneDrive can lock)
    try:
        shutil.copy2(library, out / library.name)
    except Exception:
        pass

    # Push export into PRISM (same site as tar.gz; skips re-index noise via upsert)
    prism_info: dict = {}
    try:
        from fortna_prism_ingest import after_export, stage_twin

        prism_info = after_export(export_dir=out, kind="autogen", site=file_stem)
        twin_info = stage_twin(
            site=file_stem,
            gaps=twin_gaps,
            peers=_default_twin_peers(file_stem),
        )
        prism_info = {**prism_info, "twin": twin_info}
    except Exception as exc:
        prism_info = {"ok": False, "error": str(exc)}

    result = {
        "ok": True,
        "engine": "python",
        "export_name": file_stem,
        "source_label": file_stem,
        "out_dir": str(out),
        "l5x": str(l5x_path),
        "studio_tags_csv": str(studio_csv_path) if csv_count else "",
        "report": report,
        "report_txt": str(out / "autogen_report.txt"),
        "l5x_bytes": l5x_path.stat().st_size if l5x_path.is_file() else 0,
        "twin_gaps": twin_gaps,
        "prism": prism_info,
    }
    # Persist full result for Electron recovery if stdout/IPC fails
    try:
        (out / "autogen_result.json").write_text(
            json.dumps(result, separators=(",", ":")), encoding="utf-8"
        )
    except Exception:
        pass
    _emit_progress(
        f"Done — {report.get('conveyor_count', 0)} conveyors, {report.get('tag_count', 0)} tags",
        100,
    )
    return result


def main() -> int:
    ap = argparse.ArgumentParser(description="Site Forge PLC Autogen (Excel → Python)")
    sub = ap.add_subparsers(dest="cmd", required=True)

    p_ins = sub.add_parser("inspect-excel", help="Explain sheets / why Excel feels locked")
    p_ins.add_argument("excel")

    p_ex = sub.add_parser("from-excel", help="Generate L5X from autogen .xlsm/.xlsx")
    p_ex.add_argument("excel")
    p_ex.add_argument("--library", default=str(DEFAULT_LIBRARY))
    p_ex.add_argument("--out-dir", default="")

    p_js = sub.add_parser("from-json", help="Generate L5X from JSON input snapshot")
    p_js.add_argument("json_path")
    p_js.add_argument("--library", default=str(DEFAULT_LIBRARY))
    p_js.add_argument("--out-dir", default="")

    p_demo = sub.add_parser("demo", help="Run against bundled sample Excel + library")
    p_demo.add_argument("--library", default=str(DEFAULT_LIBRARY))
    p_demo.add_argument("--excel", default=str(DEFAULT_SAMPLE_XLS))
    p_demo.add_argument("--out-dir", default="")

    p_run = sub.add_parser(
        "from-run",
        help="Build autogen input from Fortna RUN (tar.gz extract) and generate L5X — no Excel typing",
    )
    p_run.add_argument("--run-dir", default=str(DEFAULT_RUN), help="Path to RUN folder (default workspace/active/RUN)")
    p_run.add_argument("--library", default=str(DEFAULT_LIBRARY))
    p_run.add_argument("--processor", default="1756-L83E")
    p_run.add_argument("--out-dir", default="")
    p_run.add_argument(
        "--preview-only",
        action="store_true",
        help="Only dump autogen_input JSON (no L5X) to inspect mapping",
    )
    p_run.add_argument(
        "--include-programs",
        default="",
        help=(
            "Comma-separated optional programs to merge: "
            "ShippingSorter_Area_L3,WCS_Interface_TCP_IP,Sorter_Track,Sawtooth_Merge"
        ),
    )
    p_run.add_argument(
        "--no-sys",
        action="store_true",
        help="Do not merge gold Sys_Program.L5X",
    )
    p_run.add_argument(
        "--with-io-map",
        action="store_true",
        help="Include RUN/tar.gz IO_MAP program (default when --no-io-map not set)",
    )
    p_run.add_argument(
        "--no-io-map",
        action="store_true",
        help="Omit IO_MAP program from the L5X entirely",
    )
    p_run.add_argument(
        "--no-io-map-gold",
        action="store_true",
        help="Do not merge gold Excel IO_MAP_Program.L5X (default)",
    )
    p_run.add_argument(
        "--io-map-gold",
        action="store_true",
        help="OPTIONAL CLI: merge gold Excel IO_MAP_Program.L5X (Greensboro CP5–CP7)",
    )
    p_run.add_argument(
        "--workbook",
        default="",
        help="Path to Site Forge autogen_workbook.json (dashboard edits). Applied over RUN before L5X.",
    )

    args = ap.parse_args()
    try:
        def _out(obj: dict, pretty: bool = False) -> None:
            if pretty:
                print(json.dumps(obj, indent=2))
            else:
                print(json.dumps(obj, separators=(",", ":")))

        if args.cmd == "inspect-excel":
            _out(inspect_excel(Path(args.excel)), pretty=True)
            return 0
        if args.cmd == "from-excel":
            inp = load_from_excel(Path(args.excel))
            result = generate(inp, Path(args.library), Path(args.out_dir) if args.out_dir else None)
            _out(result)
            return 0 if result.get("ok") else 1
        if args.cmd == "from-json":
            inp = load_from_json(Path(args.json_path))
            result = generate(inp, Path(args.library), Path(args.out_dir) if args.out_dir else None)
            _out(result)
            return 0 if result.get("ok") else 1
        if args.cmd == "from-run":
            _emit_progress(f"Loading RUN from {args.run_dir}…", 5)
            inp = load_from_run(Path(args.run_dir), processor=args.processor)
            # Dashboard workbook (Inputdata replacement) — human edits over RUN
            wb_path = (getattr(args, "workbook", "") or "").strip()
            if wb_path:
                try:
                    from fortna_workbook import apply_workbook_to_input, load_workbook
                    wb = load_workbook(Path(wb_path))
                    if wb:
                        run_names = {
                            (c.conveyor or "").strip().upper()
                            for c in (inp.conveyors or [])
                            if (c.conveyor or "").strip()
                        }
                        inp = apply_workbook_to_input(inp, wb)
                        sb = wb.get("sorter_build")
                        if isinstance(sb, dict) and sb:
                            inp.sorter_build = sb
                        m2 = wb.get("merges_2to1")
                        if isinstance(m2, list) and m2:
                            inp.merges_2to1 = m2
                            _emit_progress(
                                f"2:1 merges from workbook: {len(m2)}",
                                12,
                            )
                        area_counts: dict[str, int] = {}
                        stub_names: list[str] = []
                        for c in inp.conveyors or []:
                            a = (c.main_area or "").strip() or "(none)"
                            area_counts[a] = area_counts.get(a, 0) + 1
                            nm = (c.conveyor or "").strip()
                            if nm and nm.upper() not in run_names:
                                stub_names.append(nm)
                        area_bits = ", ".join(f"{a}×{n}" for a, n in sorted(area_counts.items()))
                        _emit_progress(
                            f"Applied workbook ({len(wb.get('conveyors') or [])} rows) — areas: {area_bits}",
                            12,
                        )
                        if stub_names:
                            shown = ", ".join(stub_names[:16])
                            more = f" (+{len(stub_names) - 16} more)" if len(stub_names) > 16 else ""
                            _emit_progress(
                                f"Transport stubs not in active RUN (still emitted): {shown}{more}",
                                13,
                            )
                            # For PRISM twin gaps.json (Phase 0)
                            inp.transport_stub_tags = list(stub_names)
                except Exception as wb_exc:
                    _emit_progress(f"Workbook apply skipped: {wb_exc}", 12)
            # Optional gold programs (ShippingSorter / WCS) + live Sorter_Track
            raw_inc = getattr(args, "include_programs", "") or ""
            inp.include_programs = [
                x.strip() for x in str(raw_inc).replace(";", ",").split(",") if x.strip()
            ]
            inp.include_sys = not bool(getattr(args, "no_sys", False))
            # IO_MAP on/off (RUN banks). Default ON unless --no-io-map.
            if bool(getattr(args, "no_io_map", False)):
                inp.include_io_map = False
            else:
                inp.include_io_map = True
            # Gold Excel only with --io-map-gold (CLI advanced)
            if bool(getattr(args, "io_map_gold", False)):
                inp.include_io_map_gold = True
            else:
                inp.include_io_map_gold = False
            if args.preview_only:
                vfd_n = sum(1 for c in inp.conveyors if "vfd" in (c.type or "").lower())
                with_pe = sum(1 for c in inp.conveyors if c.all_pe_tags or c.exit_pe_tag)
                snap = {
                    "ok": True,
                    "engine": "python",
                    "project_name": inp.project_name,
                    "processor": inp.processor,
                    "areas": inp.areas,
                    "conveyor_count": len(inp.conveyors),
                    "vfd_conveyor_count": vfd_n,
                    "ms_conveyor_count": len(inp.conveyors) - vfd_n,
                    "conveyors_with_pe": with_pe,
                    "pe_device_count": len(inp.pe_devices or []),
                    "eip_adapter_count": len(inp.eip_adapters or []),
                    "eip_interface_ip": inp.eip_interface_ip or "",
                    "conveyors": [asdict(c) for c in inp.conveyors[:40]],
                    "pe_sample": (inp.pe_devices or [])[:15],
                    "io_point_count": len(inp.io_points),
                    "include_programs": inp.include_programs,
                    "include_sys": inp.include_sys,
                    "include_io_map_gold": inp.include_io_map_gold,
                    "optional_programs_available": list(OPTIONAL_PROGRAMS.keys()),
                    "run_dir": str(Path(args.run_dir).resolve()),
                    "note": (
                        "Python autogen preview — real PE tags + EIP from RUN. "
                        "Full list written on Generate → autogen_input.json"
                    ),
                }
                _out(snap)
                return 0
            result = generate(inp, Path(args.library), Path(args.out_dir) if args.out_dir else None)
            _out(result)
            return 0 if result.get("ok") else 1
        if args.cmd == "demo":
            xl = Path(args.excel)
            if not xl.is_file():
                _out({"ok": False, "error": f"Sample Excel not found: {xl}"})
                return 1
            inp = load_from_excel(xl)
            result = generate(inp, Path(args.library), Path(args.out_dir) if args.out_dir else None)
            _out(result)
            return 0 if result.get("ok") else 1
    except Exception as exc:
        print(json.dumps({"ok": False, "error": str(exc), "engine": "python"}, separators=(",", ":")))
        return 1
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
