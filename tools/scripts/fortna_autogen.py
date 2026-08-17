#!/usr/bin/env python3
"""
fortna_autogen.py — Python port of the Excel VBA PLC Autogen (autogen_VBS_test.xlsm).

Reads conveyor/IO tables (from Excel Inputdata/IO sheets or JSON) + O'Reilly L5X library,
then generates a Studio 5000 .L5X project by cloning library template tags/rungs and
renaming them to site conveyor names (same approach as the Excel tool).

Why Excel felt "locked":
  - VBA macros + ActiveX buttons require Trust Center macro enable
  - Dict / IO Dict / Template sheets are formula engines (not for manual edit)
  - Only Inputdata + IO are meant as inputs
  - OneDrive + huge formula sheets make Excel sluggish

Usage:
  py tools/scripts/fortna_autogen.py from-excel path/to/autogen.xlsm --library tools/libraries/OReilly_Library_v3.L5X
  py tools/scripts/fortna_autogen.py from-json path/to/input.json --library tools/libraries/OReilly_Library_v3.L5X
  py tools/scripts/fortna_autogen.py inspect-excel path/to/autogen.xlsm
"""
from __future__ import annotations

import argparse
import json
import re
import shutil
import sys
from copy import deepcopy
from dataclasses import dataclass, field, asdict
from datetime import datetime, timezone
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parent.parent
sys.path.insert(0, str(SCRIPT_DIR))
DEFAULT_LIBRARY = REPO_ROOT / "tools" / "libraries" / "OReilly_Library_v3.L5X"
DEFAULT_SAMPLE_XLS = REPO_ROOT / "tools" / "libraries" / "autogen_VBS_test.xlsm"
DEFAULT_RUN = REPO_ROOT / "workspace" / "active" / "RUN"
PROGRAM_LIBRARY_DIR = REPO_ROOT / "tools" / "libraries" / "programs"

# Gold program exports from Excel/Studio (Desktop/Autogen) — always or optional
ALWAYS_PROGRAMS: dict[str, str] = {
    "Sys": "Sys_Program.L5X",
    "IO_MAP": "IO_MAP_Program.L5X",  # full gold map; replaces RUN scaffold when present
}
OPTIONAL_PROGRAMS: dict[str, str] = {
    "ShippingSorter_Area_L3": "ShippingSorter_Area_L3_Program.L5X",
    "WCS_Interface_TCP_IP": "WCS_Interface_TCP_IP_Program.L5X",
    "Sorter_Track": "Sorter_Track_Program.L5X",
}

# Fortna ASC mechanical types → Excel autogen TYPE strings
FORTNA_TYPE_TO_AUTOGEN = {
    "STRAIGHT": "Transport with MS",
    "BELT": "Transport with MS",
    "CURVE": "Transport with MS",
    "MERGE": "Transport with MS",
    "SKEW": "Transport with MS",
    "SPUR": "Transport with MS",
    "TRIANG": "Transport with MS",
    "ACCUM": "Accumulation with MS",
    "ZEROPRESSURE": "Accumulation with MS",
}
FORTNA_TYPE_TO_AUTOGEN_VFD = {
    "STRAIGHT": "Transport with VFD",
    "BELT": "Transport with VFD",
    "CURVE": "Transport with VFD",
    "MERGE": "Transport with VFD",
    "SKEW": "Transport with VFD",
    "SPUR": "Transport with VFD",
    "TRIANG": "Transport with VFD",
    "ACCUM": "Accumulation with VFD",
    "ZEROPRESSURE": "Accumulation with VFD",
}
CONVEYOR_ASC_TYPES = frozenset(FORTNA_TYPE_TO_AUTOGEN.keys())

# Excel Config maps TYPE -> template base name. Library uses P1000–P4000 (not P7000/P8000).
# We prefer library-real templates; Config names remapped when missing.
TYPE_TO_TEMPLATE = {
    "transport with vfd": "P1000_Conv",
    "accumulation with vfd": "P2000_Conv",
    "transport with ms": "P3000_Conv",  # library: MS = P3000 (Excel said P7000)
    "accumulation with ms": "P4000_Conv",  # library: MS accum = P4000 (Excel said P8000)
    "transport with mdr": "P4000_Conv",
    "accumulation with mdr": "P3000_Conv",
    "gravity": "P3000_Conv",
}

# Excel Config TYPE string -> preferred template (from Config sheet)
EXCEL_CONFIG_OVERRIDES = {
    "transport with vfd": "P1000_Conv",
    "accumulation with vfd": "P2000_Conv",
    "transport with ms": "P7000_Conv",  # may not exist in lib — fallback applied
    "accumulation with ms": "P8000_Conv",
    "transport with mdr": "P4000_Conv",
    "accumulation with mdr": "P3000_Conv",
    "gravity": "P5000_Conv",
}


@dataclass
class ConveyorRow:
    number: int
    system: str = ""
    main_area: str = ""
    safety_zone: str = ""
    conveyor: str = ""
    type: str = ""
    downstream: str = ""
    track: str = ""
    exit_pe: str = ""
    entry_pe: str = ""
    induct: str = ""
    jam: str = ""
    height: str = ""
    full: str = ""
    add_pe: str = ""
    motor_starter: str = ""
    espb: str = ""
    espc: str = ""
    control_station: str = ""
    power_supply: str = ""
    # Real PE tag names from RUN (Excel-parity wiring)
    exit_pe_tag: str = ""          # product/discharge PE for Fast_Conv
    add_pe_tag: str = ""           # second PE for Fast_Conv add slot
    jam_pe_tags: list = field(default_factory=list)   # up to 5 for Slow_Jam
    full_pe_tags: list = field(default_factory=list)  # Full_PE AOIs
    product_pe_tags: list = field(default_factory=list)  # PE_Logic product eyes
    all_pe_tags: list = field(default_factory=list)   # every PE on this conv

    @property
    def clean_name(self) -> str:
        return re.sub(r"\s+", "", (self.conveyor or "").strip())

    @property
    def template_key(self) -> str:
        return (self.type or "").strip().lower()


@dataclass
class IoModule:
    name: str
    type: str
    slot: str = ""
    ip: str = ""
    parent: str = ""
    rack: str = ""
    connection: str = ""


@dataclass
class IoPoint:
    device_name: str
    device_type: str = ""
    source_module: str = ""
    port: str = ""
    direction: str = ""  # I / O
    slot_or_port: str = ""
    fortna_bank: str = ""
    fortna_bit: str = ""
    conveyor: str = ""
    description: str = ""
    pe_role: str = ""  # product | full | jam | other


@dataclass
class AutogenInput:
    project_name: str = "Autogen_Project"
    processor: str = "1756-L83E"
    major_rev: str = "35"
    minor_rev: str = "00"
    areas: list[str] = field(default_factory=list)
    safety_zones: list[str] = field(default_factory=list)
    conveyors: list[ConveyorRow] = field(default_factory=list)
    modules: list[IoModule] = field(default_factory=list)
    io_points: list[IoPoint] = field(default_factory=list)
    # EIP adapters from eipcfg (for Modules tree)
    eip_adapters: list = field(default_factory=list)
    eip_interface_ip: str = ""
    # Named topology: [{rio_name, rack, ip, children:[{flex_slot, type, word, ...}]}]
    eip_topology: list = field(default_factory=list)
    # Fortna octal Word → {rio_name, flex_slot, catalog, direction}
    io_word_map: dict = field(default_factory=dict)
    # All photoeyes (for PE_UDT tags + IO_MAP)
    pe_devices: list = field(default_factory=list)
    # Optional gold programs to merge (keys from OPTIONAL_PROGRAMS)
    include_programs: list = field(default_factory=list)
    # Sys gold is OK (timers/nulls).
    # include_io_map: emit RUN/tar.gz IO_MAP program (any site). OFF = no IO_MAP in L5X.
    # include_io_map_gold: optional Greensboro Excel merge (CLI only; blocked if site is CP1–CP4).
    include_sys: bool = True
    include_io_map: bool = True
    include_io_map_gold: bool = False
    # Sorter build UI config (induct / tracking / encoders / divert count)
    sorter_build: dict = field(default_factory=dict)


def load_program_export(path: Path) -> dict | None:
    """
    Load a Studio Program-export L5X (TargetType=Program) from tools/libraries/programs.

    Returns {name, program_xml, tags: [Tag blocks], datatypes_xml?, aois_xml?}.
    """
    path = Path(path)
    if not path.is_file():
        return None
    text = path.read_text(encoding="utf-8", errors="replace").lstrip("\ufeff")
    m = re.search(r'<Program\s+Use="Target"\s+([^>]+)>(.*?)</Program>', text, re.S)
    if not m:
        # Some exports may omit Use="Target"
        m = re.search(r'<Program\s+Name="([^"]+)"([^>]*)>(.*?)</Program>', text, re.S)
        if not m:
            return None
        name = m.group(1)
        attrs = f'Name="{name}"' + m.group(2)
        body = m.group(3)
    else:
        attrs = m.group(1)
        body = m.group(2)
        nm = re.search(r'Name="([^"]+)"', attrs)
        name = nm.group(1) if nm else path.stem.replace("_Program", "")

    # Controller L5X programs are not Use=Target
    program_xml = f"<Program {attrs}>{body}</Program>"
    program_xml = re.sub(r'\s*Use="Target"', "", program_xml, count=1)

    tags: list[str] = []
    # Context tags (full Base tags with Data) sit under Controller Use=Context
    ctx = re.search(r'<Tags\s+Use="Context"[^>]*>(.*?)</Tags>', text, re.S)
    if ctx:
        for tm in re.finditer(r"<Tag\b[^>]*>.*?</Tag>", ctx.group(1), re.S):
            tags.append(tm.group(0))
    # Also program-scoped tags if present
    prog_tags = re.search(r"<Program[^>]*>\s*<Tags>(.*?)</Tags>", text, re.S)
    if prog_tags and prog_tags.group(1).strip():
        for tm in re.finditer(r"<Tag\b[^>]*>.*?</Tag>", prog_tags.group(1), re.S):
            tags.append(tm.group(0))

    # Program exports use <DataTypes Use="Context"> / <AddOnInstructionDefinitions Use="Context">
    # — must allow attributes or WCS/Sorter packs merge ZERO types/AOIs (Studio: Data type not found).
    dt = re.search(r"<DataTypes\b[^>]*>.*?</DataTypes>", text, re.S)
    aoi = re.search(
        r"<AddOnInstructionDefinitions\b[^>]*>.*?</AddOnInstructionDefinitions>",
        text,
        re.S,
    )
    # Normalize wrappers to plain tags so controller merge accepts them
    dt_xml = ""
    if dt:
        dt_xml = re.sub(
            r"<DataTypes\b[^>]*>",
            "<DataTypes>",
            dt.group(0),
            count=1,
        )
    aoi_xml = ""
    if aoi:
        aoi_xml = re.sub(
            r"<AddOnInstructionDefinitions\b[^>]*>",
            "<AddOnInstructionDefinitions>",
            aoi.group(0),
            count=1,
        )
    return {
        "name": name,
        "program_xml": program_xml,
        "tags": tags,
        "datatypes_xml": dt_xml,
        "aois_xml": aoi_xml,
        "source": str(path),
        "tag_count": len(tags),
    }


def resolve_program_exports(
    include_optional: list[str] | None = None,
    *,
    include_sys: bool = True,
    include_io_map_gold: bool = False,
    programs_dir: Path | None = None,
) -> list[dict]:
    """Load Sys / optional gold programs. Gold IO_MAP only if explicitly requested."""
    pdir = Path(programs_dir) if programs_dir else PROGRAM_LIBRARY_DIR
    wanted: list[tuple[str, str]] = []
    if include_sys and "Sys" in ALWAYS_PROGRAMS:
        wanted.append(("Sys", ALWAYS_PROGRAMS["Sys"]))
    if include_io_map_gold and "IO_MAP" in ALWAYS_PROGRAMS:
        wanted.append(("IO_MAP", ALWAYS_PROGRAMS["IO_MAP"]))
    for key in include_optional or []:
        k = (key or "").strip()
        if not k:
            continue
        # Accept short aliases
        aliases = {
            "shippingsorter": "ShippingSorter_Area_L3",
            "shipping_sorter": "ShippingSorter_Area_L3",
            "shippingsorter_shoe": "ShippingSorter_Area_L3",
            "shippingsorter_shoesorter": "ShippingSorter_Area_L3",
            # PopUp Divert — no gold L5X yet; skip silently (UI option only)
            "shippingsorter_popup_divert": "",
            "shippingsorter_popupdivert": "",
            "wcs": "WCS_Interface_TCP_IP",
            "wcs_interface": "WCS_Interface_TCP_IP",
            "sorter_track": "Sorter_Track",
            "sortertrack": "Sorter_Track",
        }
        k2 = aliases.get(k.lower().replace(" ", "_").replace("-", "_"), k)
        if k2 == "" or k in (
            "ShippingSorter_PopUp_Divert",
            "ShippingSorter_PopUpDivert",
        ):
            # Placeholder pack — no program file to merge yet
            continue
        # Sorter_Track gold pack is Greensboro-fixed (~15 diverts). Live build
        # replaces it when sorter_build config is present (handled in build_l5x).
        if k2 == "Sorter_Track":
            continue
        fname = OPTIONAL_PROGRAMS.get(k2) or OPTIONAL_PROGRAMS.get(k)
        if not fname:
            # allow exact filename
            if (pdir / k).is_file():
                loaded = load_program_export(pdir / k)
                if loaded:
                    wanted.append((loaded["name"], k))
            continue
        wanted.append((k2, fname))

    out: list[dict] = []
    seen_names: set[str] = set()
    for _key, fname in wanted:
        path = pdir / fname
        loaded = load_program_export(path)
        if not loaded:
            continue
        if loaded["name"] in seen_names:
            continue
        seen_names.add(loaded["name"])
        out.append(loaded)
    return out


def _safe(s: str) -> str:
    """Studio/Logix tag name: letters, digits, underscore; must not start with a digit."""
    t = re.sub(r"[^A-Za-z0-9_]", "_", (s or "").strip())
    t = re.sub(r"_+", "_", t).strip("_")
    if not t:
        return "Tag"
    # 1PBSTART / 7ES are valid Fortna names but illegal Logix identifiers
    if t[0].isdigit():
        t = f"T_{t}"
    return t[:40]


def _xml_escape(s: str) -> str:
    return (
        (s or "")
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


# ---------------------------------------------------------------------------
# Excel reader
# ---------------------------------------------------------------------------

def inspect_excel(path: Path) -> dict:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=False, keep_vba=True)
    info = {
        "path": str(path),
        "sheets": wb.sheetnames,
        "has_vba": "vbaProject.bin" in path.read_bytes()[:100] or True,
        "notes": [
            "Edit only Inputdata (conveyors) and IO (modules/points).",
            "Config / Dict / IO Dict / Template are formula engines — look locked.",
            "Enable macros (Trust Center) for VBA Generate buttons.",
            "ActiveX controls require Excel desktop (not Excel Online).",
        ],
    }
    if "Inputdata" in wb.sheetnames:
        ws = wb["Inputdata"]
        info["project_name"] = ws["B2"].value
        info["processor"] = ws["B3"].value
        info["version"] = f"{ws['B4'].value}.{ws['C4'].value}"
        n = 0
        for r in range(17, 5000):
            if ws.cell(r, 5).value:
                n += 1
            elif n and r > 30:
                break
        info["conveyor_rows"] = n
    return info


def load_from_excel(path: Path) -> AutogenInput:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, keep_vba=False)
    if "Inputdata" not in wb.sheetnames:
        raise ValueError("Excel missing Inputdata sheet")

    ws = wb["Inputdata"]
    inp = AutogenInput(
        project_name=str(ws["B2"].value or "Autogen_Project").strip(),
        processor=str(ws["B3"].value or "1756-L83E").strip(),
        major_rev=str(ws["B4"].value or "35").strip(),
        minor_rev=str(ws["C4"].value or "00").strip().zfill(2),
    )

    # Areas row 7 C..V
    seen_areas: set[str] = set()
    for c in range(3, 23):
        v = ws.cell(7, c).value
        if not v:
            continue
        s = str(v).strip()
        if not s or "Enter" in s or "Area From" in s:
            continue
        if s not in seen_areas:
            seen_areas.add(s)
            inp.areas.append(s)

    # Safety zones row 10 C..
    for c in range(3, 25):
        v = ws.cell(10, c).value
        if v and str(v).strip():
            s = str(v).strip()
            if s not in inp.safety_zones and "Safe Zone" not in s:
                inp.safety_zones.append(s)

    # Conveyor table from row 17
    for r in range(17, 9000):
        conv = ws.cell(r, 5).value
        if not conv:
            if r > 50 and not any(ws.cell(r, c).value for c in range(1, 8)):
                # allow sparse but stop after long empty streak
                empty = all(not ws.cell(rr, 5).value for rr in range(r, min(r + 20, 9000)))
                if empty:
                    break
            continue
        row = ConveyorRow(
            number=int(ws.cell(r, 1).value or (r - 16)),
            system=str(ws.cell(r, 2).value or "").strip(),
            main_area=str(ws.cell(r, 3).value or "").strip(),
            safety_zone=str(ws.cell(r, 4).value or "").strip(),
            conveyor=str(conv).strip(),
            type=str(ws.cell(r, 6).value or "").strip(),
            downstream=str(ws.cell(r, 7).value or "").strip(),
            track=str(ws.cell(r, 8).value or "").strip(),
            exit_pe=str(ws.cell(r, 9).value or "").strip(),
            entry_pe=str(ws.cell(r, 10).value or "").strip(),
            induct=str(ws.cell(r, 11).value or "").strip(),
            jam=str(ws.cell(r, 12).value or "").strip(),
            height=str(ws.cell(r, 13).value or "").strip(),
            full=str(ws.cell(r, 14).value or "").strip(),
            add_pe=str(ws.cell(r, 15).value or "").strip(),
            motor_starter=str(ws.cell(r, 16).value or "").strip(),
            espb=str(ws.cell(r, 17).value or "").strip(),
            espc=str(ws.cell(r, 18).value or "").strip(),
            control_station=str(ws.cell(r, 19).value or "").strip(),
            power_supply=str(ws.cell(r, 20).value or "").strip(),
        )
        if row.clean_name:
            inp.conveyors.append(row)
            if row.main_area and row.main_area not in inp.areas:
                inp.areas.append(row.main_area)

    if "IO" in wb.sheetnames:
        iows = wb["IO"]
        for r in range(6, 2000):
            mod = iows.cell(r, 2).value
            if mod:
                inp.modules.append(
                    IoModule(
                        name=str(mod).strip(),
                        type=str(iows.cell(r, 3).value or "").strip(),
                        slot=str(iows.cell(r, 4).value or "").strip(),
                        ip=str(iows.cell(r, 5).value or "").strip(),
                        parent=str(iows.cell(r, 6).value or "").strip(),
                    )
                )
            dev = iows.cell(r, 8).value
            if dev:
                inp.io_points.append(
                    IoPoint(
                        device_name=str(dev).strip(),
                        device_type=str(iows.cell(r, 9).value or "").strip(),
                        source_module=str(iows.cell(r, 10).value or "").strip(),
                        port=str(iows.cell(r, 11).value or "").strip(),
                        direction=str(iows.cell(r, 13).value or "").strip().upper()[:1],
                        slot_or_port=str(iows.cell(r, 14).value or "").strip(),
                    )
                )

    return inp


def load_from_json(path: Path) -> AutogenInput:
    data = json.loads(path.read_text(encoding="utf-8"))
    convs = [ConveyorRow(**c) if isinstance(c, dict) else c for c in data.get("conveyors", [])]
    mods = [IoModule(**m) if isinstance(m, dict) else m for m in data.get("modules", [])]
    pts = [IoPoint(**p) if isinstance(p, dict) else p for p in data.get("io_points", [])]
    return AutogenInput(
        project_name=data.get("project_name", "Autogen_Project"),
        processor=data.get("processor", "1756-L83E"),
        major_rev=str(data.get("major_rev", "35")),
        minor_rev=str(data.get("minor_rev", "00")).zfill(2),
        areas=list(data.get("areas") or []),
        safety_zones=list(data.get("safety_zones") or []),
        conveyors=convs,
        modules=mods,
        io_points=pts,
    )


def _area_from_conveyor_name(name: str, fallback: str = "Main_Area") -> str:
    """P309 → Zone3_Area style bucket (autogen wants Main_Area names)."""
    m = re.match(r"^P(\d)", (name or "").upper())
    if m:
        return f"Zone{m.group(1)}_Area"
    return fallback


def _classify_pe_role(name: str, desc: str = "") -> str:
    """Classify Fortna PE into product | full | jam | other (Excel-style roles)."""
    n = (name or "").upper()
    d = (desc or "").upper()
    # Full eyes first (EZPE116_F, PE704A_F, EZPE530_F1)
    if re.search(r"_F\d*$", n) or n.endswith("_F") or "FULL EYE" in d or "FULL DETECTION" in d:
        if "FULL" in d or re.search(r"_F\d*$", n) or n.endswith("_F"):
            # jam-full hybrid suffixes like _JF stay jam
            if "_JF" in n or "_FDJ" in n:
                return "jam"
            return "full"
    if "_JF" in n or "_FDJ" in n or re.search(r"_J\d*$", n) or n.endswith("_J") or "JAM" in d:
        return "jam"
    if (
        re.search(r"_P\d*$", n)
        or n.endswith("_P")
        or "PRODUCT" in d
        or "DISCHARGE" in d
        or "PRESENT" in d
    ):
        return "product"
    return "other"


def _link_pe_to_conveyor(p: dict) -> str:
    """Resolve PE → P### conveyor name from RUN description / tag."""
    link = (p.get("conveyor") or "").upper()
    if link and re.match(r"^P\d", link):
        return link
    desc = (p.get("description") or "").upper()
    name = (p.get("fortna_name") or p.get("io_name") or "").upper()
    m = re.search(r"\bP(\d{2,4}[A-Z]?)\b", desc)
    if m:
        return f"P{m.group(1)}"
    m = re.search(r"(?:EZ)?PE(\d{2,4}[A-Z]?)", name)
    if m:
        return f"P{m.group(1)}"
    return ""


def _pe_wiring_for_conv(pe_rows: list[dict]) -> dict:
    """
    Build Excel-parity PE wiring for one conveyor from RUN photoeye rows.

    Returns tag names (not Yes/No options) for Fast_Conv / Slow_Jam / Full_PE.
    """
    product: list[str] = []
    full: list[str] = []
    jam: list[str] = []
    other: list[str] = []
    for p in pe_rows:
        raw = (p.get("fortna_name") or p.get("io_name") or "").strip()
        if not raw:
            continue
        tag = _safe(raw)
        role = _classify_pe_role(raw, p.get("description") or "")
        if role == "full":
            full.append(tag)
        elif role == "jam":
            jam.append(tag)
        elif role == "product":
            product.append(tag)
        else:
            other.append(tag)

    # Slow_Jam slots: product eyes first, then jam, then other (max 5); full eyes are Full_PE
    jam_slots: list[str] = []
    for src in (product, jam, other):
        for t in src:
            if t not in jam_slots:
                jam_slots.append(t)
            if len(jam_slots) >= 5:
                break
        if len(jam_slots) >= 5:
            break

    exit_tag = product[0] if product else (jam_slots[0] if jam_slots else "")
    add_tag = product[1] if len(product) > 1 else ""
    all_tags = list(dict.fromkeys(product + full + jam + other))

    return {
        "exit_pe_tag": exit_tag,
        "add_pe_tag": add_tag,
        "jam_pe_tags": jam_slots[:5],
        "full_pe_tags": full,
        "product_pe_tags": product,
        "all_pe_tags": all_tags,
        "full_opt": "Yes Standard Logic" if full else "",
        "exit_opt": "Yes Standard when Jam reset and PE clear" if exit_tag else "",
        "jam_opt": "Yes" if jam_slots else "",
    }


# Library templates → Rockwell catalog + AB: config datatype (from OReilly_Library_v3)
EIP_CHILD_TEMPLATE = {
    "1794-IA16": "IO_1N90_1",   # AB:1794_DI_Delay16:C:0, catalog 1794-IA16/A
    "1794-OA8I": "IO_1N90_2",   # AB:1794_DO8:C:0
    "1794-OW8": "IO_1N90_3",    # AB:1794_DO8:C:0
    "1794-IB16": "IO_1N90_4",   # AB:1794_IB16:C:0
    "1794-OB16P": "IO_1N90_5",  # AB:1794_DO16:C:0
}
EIP_CATALOG = {
    "1794-AENT": "1794-AENT",
    "1794-IA16": "1794-IA16/A",
    "1794-OA8I": "1794-OA8I/A",
    "1794-OW8": "1794-OW8/A",
    "1794-IB16": "1794-IB16/A",
    "1794-OB16P": "1794-OB16P/A",
}


def _fortna_bit_to_data_bit(bit: str | int) -> int | None:
    """Fortna PE bits are PLC-5 octal style (0-7, 10-17) → Logix Data bit 0-15."""
    s = str(bit or "").strip()
    if not s:
        return None
    try:
        # Prefer octal (matches gold IO_MAP 93/94 cases)
        return int(s, 8)
    except ValueError:
        try:
            return int(s, 10)
        except ValueError:
            return None


def _infer_rack_from_ip(ip: str, fallback: str = "CP5") -> str:
    ip = (ip or "").strip()
    m = re.search(r"\.(\d+)$", ip)
    if not m:
        return fallback
    last = int(m.group(1))
    if 51 <= last <= 53:
        return "CP5"
    if 54 <= last <= 56:
        return "CP6"
    if 57 <= last <= 58:
        return "CP7"
    return fallback


def load_eip_topology(run_dir: Path) -> dict:
    """
    Build named Flex I/O tree + Fortna Word→module map from RUN.

    Naming matches edited gold: CP5RIO0, CP5RIO0_0 (OA8I), CP5RIO0_1 (IA16), …
    Parent AENT uses AB:1794_AEN_8SLOT:I:0 / :O:0; children carry correct AB: C:0 types.

    Fortna Conveyor.asc IO_Address_Word is the EIPCSV octal Word (e.g. 510 → IA16 on CP5RIO1).
    """
    from fortna_asc import read_asc

    run_dir = Path(run_dir)
    if (run_dir / "RUN" / "project.cfg").is_file():
        run_dir = run_dir / "RUN"

    result: dict = {
        "interface_ip": "",
        "adapters_raw": [],
        "topology": [],  # named CPxRIOn with children
        "word_map": {},  # str(word) -> mapping
        "modules_flat": [],
    }

    # --- Adapters + modules from ASC (most complete for slot/bank) ---
    adapters: list[dict] = []
    proj = run_dir / "PROJECT"
    ad_path = next(iter(sorted(proj.glob("EIPAdapters.asc*"))), None)
    mod_path = next(iter(sorted(proj.glob("EIPModules.asc*"))), None)
    csv_path = next(iter(sorted(proj.glob("EIPCSV.asc*"))), None)

    # Prefer eipcfg for IPs
    try:
        from fortna_ignition_build import load_eip_modules

        eip = load_eip_modules(run_dir)
        result["interface_ip"] = eip.get("interface_ip") or ""
        if eip.get("adapters"):
            for a in eip["adapters"]:
                adapters.append({
                    "name": a.get("name") or "",
                    "ip": a.get("ip") or "",
                    "rack": (a.get("rack") or "").strip(),
                    "input_address": a.get("input_address") or "",
                    "modules": list(a.get("modules") or []),
                })
    except Exception:
        pass

    if ad_path and ad_path.is_file():
        _, rows = read_asc(ad_path)
        by_name = {a["name"]: a for a in adapters if a.get("name")}
        for r in rows:
            name = (r.get("Name") or "").strip()
            if not name or name.upper() in ("N/A", "INVALID"):
                continue
            ip = (r.get("TargetIP") or "").strip()
            rack = (r.get("Rack") or "").strip()
            if name in by_name:
                if ip:
                    by_name[name]["ip"] = ip
                if rack:
                    by_name[name]["rack"] = rack
                by_name[name]["input_address"] = r.get("InputAddress") or by_name[name].get("input_address")
            else:
                adapters.append({
                    "name": name,
                    "ip": ip,
                    "rack": rack,
                    "input_address": r.get("InputAddress") or "",
                    "modules": [],
                })
                by_name[name] = adapters[-1]

    if mod_path and mod_path.is_file():
        _, rows = read_asc(mod_path)
        by_name = {a["name"]: a for a in adapters if a.get("name")}
        for r in rows:
            name = (r.get("Name") or "").strip()
            if not name or name.upper() in ("N/A", "INVALID"):
                continue
            mt = (r.get("Type") or "").strip()
            ad = (r.get("Adapter") or "").strip()
            conn = (r.get("Connection") or "").strip()
            try:
                slot = int(float(r.get("Slot") or 0))
            except Exception:
                slot = 0
            try:
                ib = int(float(r.get("InputBank") or 0))
            except Exception:
                ib = 0
            try:
                ob = int(float(r.get("OutputBank") or 0))
            except Exception:
                ob = 0
            rec = {
                "name": name,
                "type": mt,
                "slot": slot,
                "connection": conn,
                "input_bank": ib,
                "output_bank": ob,
                "word": "",  # filled from EIPCSV
            }
            if ad in by_name:
                # replace empty modules list from xml with ASC detail
                existing = by_name[ad].setdefault("modules", [])
                # avoid dupes by slot
                if not any(int(m.get("slot") or -1) == slot for m in existing):
                    existing.append(rec)
                else:
                    for m in existing:
                        if int(m.get("slot") or -1) == slot:
                            m.update({k: v for k, v in rec.items() if v not in ("", None)})
                            break
            elif conn == "HEADNODE":
                pass

    # Attach words from EIPCSV (Word is what Conveyor.asc IO_Address_Word uses)
    word_to_mod: dict[str, dict] = {}
    if csv_path and csv_path.is_file():
        _, rows = read_asc(csv_path)
        # track last non-empty type/name for continuation rows
        last_type = ""
        last_name = ""
        last_rack = ""
        last_ip = ""
        for r in rows:
            name = (r.get("Name") or "").strip()
            mt = (r.get("Type") or "").strip()
            rack = (r.get("Rack") or "").strip()
            ip = (r.get("IP") or "").strip()
            word = str(r.get("Word") or "").strip()
            io_dir = (r.get("IO") or "").strip().upper()
            bit_rng = (r.get("Bit") or "").strip()
            if name and name.upper() not in ("N/A", "INVALID"):
                last_name = name
            else:
                name = last_name
            if mt:
                last_type = mt
            else:
                mt = last_type
            if rack:
                last_rack = rack
            else:
                rack = last_rack
            if ip:
                last_ip = ip
            else:
                ip = last_ip
            if not word or word in ("0", ""):
                continue
            # Only keep first (low) half for mapping word → card; both halves share word
            key = word
            if key not in word_to_mod or (io_dir == "I" and "00" in bit_rng):
                word_to_mod[key] = {
                    "word": word,
                    "type": mt,
                    "name": name,
                    "rack": rack,
                    "ip": ip,
                    "io": io_dir,
                    "bit_range": bit_rng,
                    "bank": str(r.get("Bank") or "").strip(),
                }

    # Stamp words onto adapter modules when type+rack match
    for ad in adapters:
        for m in ad.get("modules") or []:
            mt = (m.get("type") or "").strip()
            for w, info in word_to_mod.items():
                if info.get("type") == mt and info.get("rack") == (ad.get("rack") or info.get("rack")):
                    # match by input bank if available
                    try:
                        ib = int(info.get("bank") or -1)
                    except Exception:
                        ib = -1
                    if ib >= 0 and int(m.get("input_bank") or -2) == ib:
                        m["word"] = w
                        break
            if not m.get("word"):
                # fallback: first word of same type on same rack not yet assigned
                for w, info in word_to_mod.items():
                    if info.get("type") == mt and (
                        not ad.get("rack") or info.get("rack") == ad.get("rack")
                    ):
                        if not any(
                            (om.get("word") == w)
                            for om in (ad.get("modules") or [])
                        ):
                            m["word"] = w
                            break

    # Name adapters CPxRIOn and children CPxRIOn_k
    by_rack: dict[str, list] = {}
    for ad in adapters:
        rack = (ad.get("rack") or "").strip().upper()
        if not rack:
            rack = _infer_rack_from_ip(ad.get("ip") or "")
            ad["rack"] = rack
        by_rack.setdefault(rack or "CP5", []).append(ad)

    topology: list[dict] = []
    word_map: dict[str, dict] = {}
    modules_flat: list[IoModule] = []

    for rack in sorted(by_rack.keys()):
        # Gold Excel / IO_MAP names CP7 heads CP7RIO1 + CP7RIO2 (no CP7RIO0).
        # CP5/CP6 use RIO0..n. Mismatch left gold IO_MAP OTE(CP7RIO2:…) undefined.
        rio_start = 1 if str(rack).upper() in ("CP7", "PLC7") else 0
        for idx, ad in enumerate(by_rack[rack]):
            rio = f"{rack}RIO{idx + rio_start}"
            ip = (ad.get("ip") or "").strip()
            children = []
            # Bridged modules only; Flex address = EIP slot - 1 when slot0 is AENT headnode
            bridged = [
                m for m in sorted(ad.get("modules") or [], key=lambda x: int(x.get("slot") or 0))
                if (m.get("connection") or "").upper() != "HEADNODE"
                and (m.get("type") or "").upper() != "1794-AENT"
            ]
            # If modules list has no connection flags, skip pure AENT types
            if not bridged:
                bridged = [
                    m for m in sorted(ad.get("modules") or [], key=lambda x: int(x.get("slot") or 0))
                    if "AENT" not in (m.get("type") or "").upper()
                ]
            for m in bridged:
                eip_slot = int(m.get("slot") or 0)
                # Gold: first I/O card is flex address 0 (EIP often uses slot 1 after AENT@0)
                flex = eip_slot - 1 if eip_slot >= 1 else eip_slot
                # If slots already 0-based without headnode, detect: min slot == 0 and type not AENT
                mt = (m.get("type") or "").strip()
                child_name = f"{rio}_{flex}"
                catalog = EIP_CATALOG.get(mt, mt)
                word = str(m.get("word") or "").strip()
                # Prefer EIPCSV word for this type@rack@bank
                if not word:
                    for w, info in word_to_mod.items():
                        if info.get("type") == mt and info.get("rack") == rack:
                            try:
                                if int(info.get("bank") or -1) == int(m.get("input_bank") or -2):
                                    word = w
                                    break
                            except Exception:
                                pass
                child = {
                    "name": child_name,
                    "type": mt,
                    "catalog": catalog,
                    "template": EIP_CHILD_TEMPLATE.get(mt, ""),
                    "eip_slot": eip_slot,
                    "flex_slot": flex,
                    "word": word,
                    "input_bank": m.get("input_bank"),
                    "output_bank": m.get("output_bank"),
                    "direction": "I" if any(x in mt for x in ("IA", "IB", "IM")) else (
                        "O" if any(x in mt for x in ("OA", "OB", "OW")) else ""
                    ),
                }
                children.append(child)
                if word:
                    word_map[word] = {
                        "rio_name": rio,
                        "child_name": child_name,
                        "flex_slot": flex,
                        "type": mt,
                        "catalog": catalog,
                        "rack": rack,
                        "ip": ip,
                        "direction": child["direction"],
                    }
                modules_flat.append(
                    IoModule(
                        name=child_name,
                        type=mt,
                        slot=str(flex),
                        ip="",
                        parent=rio,
                        rack=rack,
                        connection="BRIDGED",
                    )
                )

            # Fix flex slots if EIP used 0-based without headnode (min flex became -1)
            if children and min(c["flex_slot"] for c in children) < 0:
                for c in children:
                    c["flex_slot"] = c["eip_slot"]
                    c["name"] = f"{rio}_{c['flex_slot']}"

            # Also map any EIPCSV words for this rack/type that match child types
            for c in children:
                if c.get("word"):
                    continue
                for w, info in word_to_mod.items():
                    if info.get("rack") == rack and info.get("type") == c["type"] and w not in word_map:
                        c["word"] = w
                        word_map[w] = {
                            "rio_name": rio,
                            "child_name": c["name"],
                            "flex_slot": c["flex_slot"],
                            "type": c["type"],
                            "catalog": c["catalog"],
                            "rack": rack,
                            "ip": ip,
                            "direction": c["direction"],
                        }
                        break

            topology.append({
                "rio_name": rio,
                "adapter_name": ad.get("name") or "",
                "rack": rack,
                "ip": ip,
                "children": children,
            })
            modules_flat.insert(
                0,
                IoModule(
                    name=rio,
                    type="1794-AENT",
                    slot="0",
                    ip=ip,
                    parent="",
                    rack=rack,
                    connection="HEADNODE",
                ),
            )

    # Final word_map pass: map every EIPCSV input word of known type on known rack
    # to the matching child by order of input cards
    for rack, ads in by_rack.items():
        for idx, ad in enumerate(ads):
            rio = f"{rack}RIO{idx}"
            topo = next((t for t in topology if t["rio_name"] == rio), None)
            if not topo:
                continue
            in_children = [c for c in topo["children"] if c["direction"] == "I"]
            # words for this rack that are input modules
            rack_words = [
                (w, info) for w, info in word_to_mod.items()
                if info.get("rack") == rack and info.get("io") == "I"
                and any(x in (info.get("type") or "") for x in ("IA", "IB"))
            ]
            rack_words.sort(key=lambda x: int(x[0]) if str(x[0]).isdigit() else 0)
            # assign by sequence if not already mapped
            used_flex = {word_map[w]["flex_slot"] for w in word_map if word_map[w].get("rio_name") == rio}
            for w, info in rack_words:
                if w in word_map:
                    continue
                mt = info.get("type") or ""
                for c in in_children:
                    if c["type"] == mt and c["flex_slot"] not in used_flex:
                        word_map[w] = {
                            "rio_name": rio,
                            "child_name": c["name"],
                            "flex_slot": c["flex_slot"],
                            "type": c["type"],
                            "catalog": c["catalog"],
                            "rack": rack,
                            "ip": topo["ip"],
                            "direction": "I",
                        }
                        c["word"] = w
                        used_flex.add(c["flex_slot"])
                        break

    result["adapters_raw"] = adapters
    result["topology"] = topology
    result["word_map"] = word_map
    result["modules_flat"] = modules_flat
    return result


def _load_eip_adapters(run_dir: Path) -> tuple[list[dict], str, list[IoModule], list, dict]:
    """Backward-compatible wrapper → topology + word map."""
    topo = load_eip_topology(run_dir)
    return (
        topo.get("adapters_raw") or [],
        topo.get("interface_ip") or "",
        list(topo.get("modules_flat") or []),
        list(topo.get("topology") or []),
        dict(topo.get("word_map") or {}),
    )


def load_from_run(run_dir: Path, *, processor: str = "1756-L83E") -> AutogenInput:
    """
    Build autogen input from a Fortna RUN package (tar.gz already extracted).

    This replaces the manual step: pull tags from prints → type into Excel.
    Conveyor.asc mechanical rows become Inputdata conveyor lines; PE rows attach
    real PE tag names for Fast_Conv/Slow_Jam/PE_Logic; EIP modules from eipcfg.
    """
    from fortna_asc import read_asc
    from fortna_io_extract import (
        equipment_kind,
        extract_io_points,
        normalize_io_name,
        read_project_meta,
    )

    run_dir = Path(run_dir)
    if (run_dir / "RUN" / "project.cfg").is_file():
        run_dir = run_dir / "RUN"
    if not (run_dir / "project.cfg").is_file():
        raise FileNotFoundError(f"No Fortna RUN at {run_dir}")

    meta = read_project_meta(run_dir)
    machine = meta.get("machine_name") or "Machine"
    project = meta.get("project_name") or machine

    conv_path = run_dir / "FORTNA" / "Conveyor.asc"
    if not conv_path.is_file():
        raise FileNotFoundError(f"Missing Conveyor.asc under {run_dir}")

    _, rows = read_asc(conv_path)
    points = extract_io_points(run_dir, include_spares=False)

    # EIP first — word_map scopes devices to this master PLC's RIO network
    eip_adapters, eip_ip, eip_modules, eip_topology, io_word_map = _load_eip_adapters(run_dir)

    from fortna_io_extract import belongs_to_controller, row_machine_matches

    # --- Scope I/O + PE to this controller only (Greensboro has 3 masters) ---
    pe_by_conv: dict[str, list[dict]] = {}
    io_points: list[IoPoint] = []
    pe_devices: list[dict] = []
    linked_conveyors: set[str] = set()  # P### owned via PE/IO on this machine

    # VFD tags on this controller only (for conveyor MS vs VFD classification)
    vfd_num_keys: set[str] = set()
    for row in rows:
        rn = (row.get("IO_Name") or "").strip()
        if not re.match(r"^VFD", rn, re.I):
            continue
        if not belongs_to_controller(
            machine_name=(row.get("Machine_Name") or "").strip(),
            io_word=str(row.get("IO_Address_Word") or "").strip(),
            controller=machine,
            word_map=io_word_map,
        ):
            continue
        core = normalize_io_name(rn)
        m = re.match(r"^VFD[\s\-_]*([A-Z0-9]*\d[A-Z0-9]{0,6})", core, re.I)
        if m:
            c = re.sub(r"(_EN|_AUX|_FLT|_RUN|_OK)$", "", m.group(1), flags=re.I).upper()
            if c:
                vfd_num_keys.add(c)
                dm = re.search(r"(\d{2,4})", c)
                if dm:
                    vfd_num_keys.add(dm.group(1))

    # Optional print OCR VFD ids (from last OCR run) → same number matching
    try:
        ocr_path = REPO_ROOT / "workspace" / "ocr-last-result.json"
        if ocr_path.is_file():
            ocr = json.loads(ocr_path.read_text(encoding="utf-8"))
            for p in ocr.get("print_vfd_params") or []:
                val = str(p.get("value") or p.get("device_id") or "")
                m = re.match(r"^VFD[\s\-_]*([A-Z0-9]*\d[A-Z0-9]{0,6})", val, re.I)
                if m:
                    core = m.group(1).upper()
                    vfd_num_keys.add(core)
                    dm = re.search(r"(\d{2,4})", core)
                    if dm:
                        vfd_num_keys.add(dm.group(1))
    except Exception:
        pass

    for p in points:
        if not belongs_to_controller(
            machine_name=str(p.get("machine_name") or ""),
            io_word=str(p.get("fortna_bank") or ""),
            controller=machine,
            word_map=io_word_map,
        ):
            continue
        kind = p.get("equipment_kind") or equipment_kind(
            p.get("fortna_name") or "", p.get("device_type") or "", p.get("description") or ""
        )
        name = p.get("fortna_name") or ""
        if kind == "photoeye":
            link = _link_pe_to_conveyor(p)
            role = _classify_pe_role(name, p.get("description") or "")
            pe_rec = {
                "name": _safe(name),
                "fortna_name": name,
                "conveyor": link,
                "role": role,
                "bank": str(p.get("fortna_bank") or ""),
                "bit": str(p.get("fortna_bit") or ""),
                "description": (p.get("description") or "")[:120],
                "machine_name": str(p.get("machine_name") or ""),
            }
            pe_devices.append(pe_rec)
            if link:
                pe_by_conv.setdefault(link.upper(), []).append(p)
                linked_conveyors.add(link.upper())
            io_points.append(
                IoPoint(
                    device_name=name,
                    device_type="photoeye",
                    source_module=p.get("module") or "",
                    direction="I",
                    fortna_bank=str(p.get("fortna_bank") or ""),
                    fortna_bit=str(p.get("fortna_bit") or ""),
                    conveyor=link,
                    description=(p.get("description") or "")[:120],
                    pe_role=role,
                )
            )
        elif p.get("fortna_bank") and kind not in ("conveyor", "spare", "invalid"):
            # Every bank-addressed field device from tar.gz (PB, ES, motor, beacon,
            # digital_in, scanner, power_supply, …) — not just PE.
            if kind == "vfd":
                dm = re.search(r"(\d{2,4})", name)
                if dm:
                    linked_conveyors.add(f"P{dm.group(1)}")
            io_dir = str(p.get("io_type") or "").upper()
            direction = "O" if io_dir in ("OUT", "O", "OUTPUT") else "I"
            # Beacons / horns are almost always outputs even if Type mislabeled
            if kind in ("beacon",) or re.search(r"WH\d|HORN|BEACON|LAMP", name, re.I):
                direction = "O"
            io_points.append(
                IoPoint(
                    device_name=name,
                    device_type=kind or str(p.get("device_class") or "io"),
                    source_module=p.get("module") or "",
                    direction=direction,
                    fortna_bank=str(p.get("fortna_bank") or ""),
                    fortna_bit=str(p.get("fortna_bit") or ""),
                    description=(p.get("description") or "")[:120],
                )
            )

    # Conveyors: explicit Machine_Name match OR linked from this controller's PE/VFD
    # (plant-wide ASC often leaves conveyors as Machine_Name=N/A)
    conveyors: list[ConveyorRow] = []
    areas: list[str] = []
    n = 0
    for row in rows:
        typ = (row.get("Type") or "").strip().upper()
        if typ not in CONVEYOR_ASC_TYPES:
            continue
        raw_name = (row.get("IO_Name") or "").strip()
        if not raw_name or raw_name.upper() in ("INVALID", "N/A", "SPARE", "ALWAYSON", "NEVERON"):
            continue
        kind = equipment_kind(raw_name, typ, row.get("General_Description") or "")
        if kind in ("photoeye", "pushbutton", "estop", "beacon", "motor", "vfd"):
            continue
        name = normalize_io_name(raw_name)
        if not re.match(r"^P\d{2,4}[A-Z0-9_]*$", name, re.I):
            continue
        if re.search(r"_AUX$|_FLT$|_OK$|_RUN$", name, re.I):
            continue
        if not name:
            continue

        row_mach = (row.get("Machine_Name") or "").strip()
        name_u = name.upper()
        # Exclude conveyors tagged to a different master
        if row_mach and row_mach.upper() not in ("N/A", "INVALID", "", "NONE", "ALL"):
            if not row_machine_matches(row_mach, machine):
                continue
        else:
            # Untagged: only if PE/VFD on this controller linked to this belt
            # also allow base P### match for P600C-style PE linking to P600
            base_m = re.match(r"^(P\d{2,4})", name_u)
            base = base_m.group(1) if base_m else name_u
            if name_u not in linked_conveyors and base not in linked_conveyors:
                # any linked name starts with this conveyor (P600 ← P600C PE)
                if not any(
                    lc == name_u or lc.startswith(name_u) or name_u.startswith(lc)
                    for lc in linked_conveyors
                ):
                    continue

        desc = (row.get("General_Description") or "").strip()
        drive = (row.get("Drive") or "").strip()
        is_vfd = bool(re.search(r"\bVFD\b", f"{raw_name} {desc} {drive}", re.I))
        if drive in ("0", "1", " ", "N/A", "N", "~", ""):
            pass
        elif any(c.isalpha() for c in drive) or len(drive) > 2:
            is_vfd = True
        if not is_vfd:
            pm = re.match(r"^P(\d{2,4}[A-Z]?)", name, re.I)
            if pm:
                pkey = pm.group(1).upper()
                pdig = re.search(r"(\d{2,4})", pkey)
                if pkey in vfd_num_keys or (pdig and pdig.group(1) in vfd_num_keys):
                    is_vfd = True

        type_map = FORTNA_TYPE_TO_AUTOGEN_VFD if is_vfd else FORTNA_TYPE_TO_AUTOGEN
        ag_type = type_map.get(typ, "Transport with MS")

        area = _area_from_conveyor_name(name, f"{_safe(machine)}_Area")
        if area not in areas:
            areas.append(area)

        pe = _pe_wiring_for_conv(
            pe_by_conv.get(name_u, []) or pe_by_conv.get(name, [])
        )
        n += 1
        conveyors.append(
            ConveyorRow(
                number=n,
                system=machine,
                main_area=area,
                safety_zone=f"{area.replace('_Area', '')}_ESZone1",
                conveyor=name,
                type=ag_type,
                downstream="",  # topology not reliable in ASC
                exit_pe=pe["exit_opt"],
                full=pe["full_opt"],
                jam=pe["jam_opt"] or pe["exit_opt"],
                motor_starter="" if is_vfd else "Yes",
                exit_pe_tag=pe["exit_pe_tag"],
                add_pe_tag=pe["add_pe_tag"],
                jam_pe_tags=pe["jam_pe_tags"],
                full_pe_tags=pe["full_pe_tags"],
                product_pe_tags=pe["product_pe_tags"],
                all_pe_tags=pe["all_pe_tags"],
            )
        )

    conveyors.sort(key=lambda c: c.conveyor)

    return AutogenInput(
        project_name=f"{project}_{machine}",
        processor=processor,
        major_rev="35",
        minor_rev="00",
        areas=areas,
        safety_zones=[f"{a.replace('_Area', '')}_ESZone1" for a in areas],
        conveyors=conveyors,
        modules=eip_modules,
        io_points=io_points,
        eip_adapters=eip_adapters,
        eip_interface_ip=eip_ip,
        eip_topology=eip_topology,
        io_word_map=io_word_map,
        pe_devices=pe_devices,
    )


# ---------------------------------------------------------------------------
# Library template resolution
# ---------------------------------------------------------------------------

def resolve_template(conv_type: str, library_text: str) -> str:
    key = (conv_type or "").strip().lower()
    preferred = EXCEL_CONFIG_OVERRIDES.get(key) or TYPE_TO_TEMPLATE.get(key) or "P1000_Conv"
    fallbacks = [
        preferred,
        TYPE_TO_TEMPLATE.get(key, "P1000_Conv"),
        "P1000_Conv",
        "P3000_Conv",
        "P2000_Conv",
        "P4000_Conv",
    ]
    for t in fallbacks:
        if t and f'Tag Name="{t}"' in library_text:
            return t
    return "P1000_Conv"


def extract_tag_block(library_text: str, tag_name: str) -> str | None:
    """Extract full <Tag Name="...">...</Tag> block."""
    pat = rf'<Tag Name="{re.escape(tag_name)}"[^>]*>.*?</Tag>'
    m = re.search(pat, library_text, re.S)
    return m.group(0) if m else None


def extract_routine_block(library_text: str, routine_name: str) -> str | None:
    pat = rf'<Routine Name="{re.escape(routine_name)}"[^>]*>.*?</Routine>'
    m = re.search(pat, library_text, re.S)
    return m.group(0) if m else None


def _pad_pe_slots(tags: list[str], n: int = 5) -> list[str]:
    """Pad PE tag list to n slots with NO_PE (library PE_UDT null object)."""
    out = [t for t in (tags or []) if t][:n]
    while len(out) < n:
        out.append("NO_PE")
    return out


def clone_template_for_conveyor(
    library_text: str,
    template: str,
    conveyor: str,
    area: str,
    safety_zone: str,
    downstream: str,
    *,
    is_vfd: bool = False,
    exit_pe_tag: str = "",
    add_pe_tag: str = "",
    jam_pe_tags: list | None = None,
    full_pe_tags: list | None = None,
    product_pe_tags: list | None = None,
) -> dict:
    """
    Clone library template tags + Excel-style AOI rungs for one conveyor.

    Matches edited gold L5X patterns:
      Fast_Conv(…, Next_or_NO_Conv, ExitPE_or_NO_PE, AddPE_or_NO_PE, …)
      Slow_Jam(…, PE1..PE5 with NO_PE padding)
      PE_Logic / Full_PE per eye
      Slow_Flt for motor/VFD fault
    """
    conv = _safe(conveyor)
    area_s = _safe(area) or "Main_Area"
    safe_s = _safe(safety_zone) or f"{area_s}_Safe"
    if not (safe_s.endswith("_Safe") or "ESZone" in safe_s):
        safe_s = f"{safe_s}_Safe" if safe_s else f"{area_s}_Safe"

    # Next conveyor: real Pxxx_Conv or NO_Conv (never fake Next_Conv tag)
    next_tag = "NO_Conv"
    if downstream:
        dn = _safe(downstream)
        if dn and dn not in ("Next_Conv", "NO_Conv"):
            next_tag = f"{dn}_Conv" if not dn.endswith("_Conv") else dn

    base = template  # P1000_Conv / P3000_Conv / …
    aoi = f"{base}_AOI"
    new_base = f"{conv}_Conv"
    new_aoi = f"{conv}_Conv_AOI"
    # Gold finished PLC5 (ORLY … PLC5Finished): tags named P###_VFD are
    # Motor_Starter_UDT (aux / contactor I/O), NOT Ethernet VFD_UDT.
    # Slow_Flt(IO_VFD=NO_VFD, IO_MS=P###_VFD) and Fast_Conv uses NO_VFD.
    # True PowerFlex program params live in print OCR / docs, not this UDT.
    vfd_tag = "NO_VFD"
    ms_tag = f"{conv}_VFD" if is_vfd else f"{conv}_MS"

    exit_pe = _safe(exit_pe_tag) if exit_pe_tag else "NO_PE"
    add_pe = _safe(add_pe_tag) if add_pe_tag else "NO_PE"
    jam_slots = _pad_pe_slots([_safe(t) for t in (jam_pe_tags or []) if t], 5)

    replacements = [
        (base, new_base),
        (aoi, new_aoi),
        ("Conv_Area_Safe", safe_s),
        ("Conv_Area", area_s),
        ("Main_Area", area_s),
    ]

    tags = []
    # Conv_UDT + Conv_AOI backing tag + optional brake tags (same as Excel autogen)
    for name in (base, aoi, f"{base}_BrakeReleased", f"{base}_BrakeReleased_Aux"):
        block = extract_tag_block(library_text, name)
        if not block:
            continue
        cloned = block
        for old, new in sorted(replacements, key=lambda x: -len(x[0])):
            cloned = cloned.replace(old, new)
        tags.append(cloned)

    # Motor_Starter_UDT for Slow_Flt IO_MS — MS conveyors use P###_MS; VFD-fed
    # discrete I/O conveyors use P###_VFD (same UDT, gold naming).
    ms_src = extract_tag_block(library_text, "NO_MS")
    if ms_src:
        tags.append(ms_src.replace("NO_MS", ms_tag))

    # Real PE_UDT + PE_Logic / Full_PE AOI backing tags (cloned from library templates)
    pe_udt_src = (
        extract_tag_block(library_text, "PE1000_P")
        or extract_tag_block(library_text, "PEExit_P")
        or extract_tag_block(library_text, "NO_PE")
    )
    pe_logic_src = extract_tag_block(library_text, "PE1000_P_AOI")
    full_pe_src = extract_tag_block(library_text, "PE1000_F") or pe_udt_src
    full_aoi_src = extract_tag_block(library_text, "PE1000_F_AOI")

    pe_tags_needed = list(
        dict.fromkeys(
            [t for t in jam_slots if t != "NO_PE"]
            + [exit_pe, add_pe]
            + [_safe(t) for t in (full_pe_tags or []) if t]
            + [_safe(t) for t in (product_pe_tags or []) if t]
        )
    )
    pe_tags_needed = [t for t in pe_tags_needed if t and t != "NO_PE"]

    for pe_name in pe_tags_needed:
        role_full = pe_name in {_safe(t) for t in (full_pe_tags or [])} or re.search(
            r"_F\d*$", pe_name, re.I
        )
        if role_full and full_pe_src:
            tags.append(full_pe_src.replace("PE1000_F", pe_name).replace("PEExit_P", pe_name).replace("NO_PE", pe_name))
            if full_aoi_src:
                tags.append(
                    full_aoi_src.replace("PE1000_F_AOI", f"{pe_name}_AOI").replace(
                        "PE1000_F", pe_name
                    )
                )
        else:
            if pe_udt_src:
                tags.append(
                    pe_udt_src.replace("PE1000_P", pe_name)
                    .replace("PEExit_P", pe_name)
                    .replace("NO_PE", pe_name)
                )
            if pe_logic_src:
                tags.append(
                    pe_logic_src.replace("PE1000_P_AOI", f"{pe_name}_AOI").replace(
                        "PE1000_P", pe_name
                    )
                )

    # --- Rungs (Excel / edited gold mnemonics) ---
    fast_text = (
        f"Fast_Conv({new_aoi}.Fast,{new_base},{area_s},{safe_s},{next_tag},"
        f"{exit_pe},{add_pe},{new_base}.Type,{vfd_tag},0,1,0,0,HMIColor,HMI_StatsClear);"
    )
    jam_text = (
        f"Slow_Jam({new_aoi}.Jam,{new_base},{area_s},"
        f"{','.join(jam_slots)});"
    )
    # Slow_Flt — motor/VFD fault (Type2 = standard MS type code from library)
    flt_text = (
        f"Slow_Flt({new_aoi}.Flt,{new_base},{area_s},{vfd_tag},NO_Enc,Type2,"
        f"{ms_tag},NO_PS,NO_AirPress,NO_AdditionalFlt,{area_s}.MtrFlt_Reset);"
    )

    # Excel Autogen rung comments (tilde banner — clean in Studio ladder view)
    # "with VFD" = discrete VFD feeder (P###_VFD Motor_Starter_UDT), not Ethernet VFD_UDT
    conv_kind = "Accumulation Conv with VFD" if is_vfd else "Transport Conv with MS"
    if not is_vfd and any(
        x in (template or "").upper() for x in ("P3000", "P4000", "ACCUM", "ZERO")
    ):
        conv_kind = "Accumulation Conv with MS"
    elif is_vfd and any(x in (template or "").upper() for x in ("P3000", "P4000")):
        conv_kind = "Accumulation Conv with VFD"

    def _excel_comment(*lines: str) -> str:
        body = "\n".join(lines)
        return f"~~~~~~~~~~~\n{body}\n~~~~~~~~~~~"

    rungs = [
        {
            "label": "Fast",
            "text": fast_text,
            "comment": _excel_comment(f"{new_base} Fast logic", f"({conv_kind})"),
        },
        {
            "label": "Jam",
            "text": jam_text,
            "comment": _excel_comment(
                f"{new_base} Jam logic",
                "Jam Sum logic & Jam Silence logic",
            ),
        },
        {
            "label": "Flt",
            "text": flt_text,
            "comment": _excel_comment(
                f"{new_base} Motor/VFD Fault logic",
                "Standard Logic",
            ),
        },
    ]

    # Per-eye PE_Logic / Full_PE rungs
    for pe_name in pe_tags_needed:
        is_full = pe_name in {_safe(t) for t in (full_pe_tags or [])} or bool(
            re.search(r"_F\d*$", pe_name, re.I)
        )
        aoi_tag = f"{pe_name}_AOI"
        if is_full:
            rungs.append(
                {
                    "label": "PE",
                    "text": f"Full_PE({aoi_tag},{pe_name},{new_base},1,0,HMI_StatsClear);",
                    "comment": _excel_comment(
                        f"{pe_name} Full PELogic",
                        "Note: Some of the AOI values are coming from the Full_Init Routine from Slow Task",
                        "Standard Logic",
                    ),
                }
            )
        else:
            rungs.append(
                {
                    "label": "PE",
                    "text": (
                        f"PE_Logic({aoi_tag},{pe_name},{new_base},"
                        f"{area_s}.PI.JamRst,0,0,0,{aoi_tag}.I_AutoClearTmr,HMI_StatsClear);"
                    ),
                    "comment": _excel_comment(
                        f"{pe_name} Debounce and Jam Logic",
                        "Standard Logic -> Jam Reset only when PE is Clear",
                    ),
                }
            )

    if not pe_tags_needed:
        rungs.append(
            {
                "label": "PE",
                "text": "NOP();",
                "comment": f"{conv} PE",
            }
        )

    return {
        "conveyor": conv,
        "template": template,
        "area": area_s,
        "safety_zone": safe_s,
        "is_vfd": is_vfd,
        "exit_pe": exit_pe,
        "jam_pes": jam_slots,
        "pe_tags": pe_tags_needed,
        "ms_tag": ms_tag,
        "tags": tags,
        "rungs": rungs,
        "tag_names": [new_base, new_aoi] + pe_tags_needed,
    }


# ---------------------------------------------------------------------------
# L5X assembly
# ---------------------------------------------------------------------------

# Preferred short AOI titles (Studio Description field — one line only).
_AOI_SHORT_DESC: dict[str, str] = {
    "Fast_Conv": "Conv Fast Routine Logic AOI",
    "Slow_Jam": "Conveyor Jam Logic AOI",
    "Slow_Flt": "Conveyor Fault Logic AOI",
    "PE_Logic": "PE Clear and Jam Logic AOI",
    "Full_PE": "Full PE Logic AOI",
    "Full_PE_Init": "Full PE Init AOI",
    "Fast_Pulse": "Pulse Logic AOI",
    "Fast_TimeStamp": "Date Time Stamp AOI",
    "AB_VFD35": "PowerFlex 35 VFD Logic AOI",
    "AB_VFD525": "PowerFlex 525 VFD Logic AOI",
    "AB_VFD750": "PowerFlex 750 VFD Logic AOI",
    "ES_PI10": "E-Stop PI 10 Logic AOI",
    "ES_PI20": "E-Stop PI 20 Logic AOI",
    "ES_SIL1_Cat1": "E-Stop SIL1 Cat1 Logic AOI",
    "Enc_CounterCard": "Encoder Counter Card Logic AOI",
    "Enc_RIOCard": "Encoder RIO Card Logic AOI",
    "Enc_Virtual_DistBased": "Virtual Encoder Distance AOI",
    "Gapper_Basic": "Gapper Basic Logic AOI",
    "Merge_2to1": "2:1 Merge Logic AOI",
    "Merge_2to1_RAT": "2:1 RAT Merge Logic AOI",
    "AOI_CommDiag": "Comm Loss Diagnostic AOI",
    "AOI_TIME_ADD": "DateTime Add Time AOI",
    "AOI_TIME_DIFFERENCE": "DateTime Difference AOI",
    "AOI_SNTP_QUERY": "SNTP Clock Sync AOI",
}


def _aoi_short_description(name: str, current: str = "") -> str:
    """One-line AOI purpose for Studio. Prefer curated map, else clean/fallback."""
    if name in _AOI_SHORT_DESC:
        return _AOI_SHORT_DESC[name]
    cur = (current or "").strip()
    # Drop copyright / parameter-list essays
    junk = (
        "copyright" in cur.lower()
        or "unauthorized" in cur.lower()
        or "all rights reserved" in cur.lower()
        or cur.count(",") >= 3
        or len(cur) > 72
        or "\n" in cur
    )
    if cur and not junk:
        # Keep a short existing title (strip trailing punctuation noise)
        first = re.split(r"[\r\n:]", cur, maxsplit=1)[0].strip()
        if 3 <= len(first) <= 72 and "copyright" not in first.lower():
            if not first.upper().endswith("AOI"):
                return f"{first} AOI" if not first.endswith("AOI") else first
            return first
    # Name → readable title
    pretty = name.replace("_", " ").strip()
    if not pretty.upper().endswith("AOI"):
        pretty = f"{pretty} AOI"
    return pretty


def _shorten_aoi_descriptions(aoi_xml: str) -> str:
    """
    Rewrite unsealed AOI <Description> text for Studio.

    CRITICAL: never touch <EncodedData>…</EncodedData> sealed bodies.
    Mutating sealed AOI XML (even Description CDATA) invalidates SignatureID
    and produces Studio: "Invalid signature ID. Reseal instruction to resolve."
    """
    if not aoi_xml:
        return aoi_xml

    def _repl_desc(m: re.Match) -> str:
        open_tag = m.group(1)
        name = m.group(2)
        old = m.group(3) or ""
        short = _aoi_short_description(name, old)
        return f"{open_tag}{short}{m.group(4)}"

    # Plain (unsealed) AOI definitions only — sealed EncodedData left byte-for-byte
    aoi_xml = re.sub(
        r'(<AddOnInstructionDefinition\b[^>]*\bName="([^"]+)"[^>]*>\s*'
        r'<Description>\s*<!\[CDATA\[)(.*?)(\]\]>\s*</Description>)',
        _repl_desc,
        aoi_xml,
        flags=re.S,
    )
    return aoi_xml


def build_l5x(inp: AutogenInput, library_path: Path) -> tuple[str, dict]:
    library_text = library_path.read_text(encoding="utf-8", errors="replace")

    # Pull DataTypes + AOI definitions from library wholesale (allow attributes)
    datatypes = re.search(r"<DataTypes\b[^>]*>.*?</DataTypes>", library_text, re.S)
    aois = re.search(
        r"<AddOnInstructionDefinitions\b[^>]*>.*?</AddOnInstructionDefinitions>",
        library_text,
        re.S,
    )

    stamp = datetime.now(timezone.utc).strftime("%a %b %d %H:%M:%S %Y")
    proj = _safe(inp.project_name) or "Autogen_Project"
    major = str(inp.major_rev or "35")
    minor = str(inp.minor_rev or "00").zfill(2)
    processor = inp.processor or "1756-L83E"

    cloned = []
    by_area: dict[str, list] = {}
    missing_templates = set()
    for row in inp.conveyors:
        tmpl = resolve_template(row.type, library_text)
        if tmpl != (EXCEL_CONFIG_OVERRIDES.get(row.template_key) or tmpl):
            pass
        excel_pref = EXCEL_CONFIG_OVERRIDES.get(row.template_key)
        if excel_pref and f'Tag Name="{excel_pref}"' not in library_text:
            missing_templates.add(excel_pref)
        is_vfd = "vfd" in (row.type or "").lower()
        item = clone_template_for_conveyor(
            library_text,
            tmpl,
            row.clean_name,
            row.main_area or "Main_Area",
            row.safety_zone or "",
            row.downstream or "",
            is_vfd=is_vfd,
            exit_pe_tag=getattr(row, "exit_pe_tag", "") or "",
            add_pe_tag=getattr(row, "add_pe_tag", "") or "",
            jam_pe_tags=list(getattr(row, "jam_pe_tags", None) or []),
            full_pe_tags=list(getattr(row, "full_pe_tags", None) or []),
            product_pe_tags=list(getattr(row, "product_pe_tags", None) or []),
        )
        item["type"] = row.type
        item["number"] = row.number
        cloned.append(item)
        by_area.setdefault(item["area"], []).append(item)

    # Controller tags: all cloned tags + area helpers
    all_tags = []
    seen_tag_names = set()
    for item in cloned:
        for block in item["tags"]:
            m = re.search(r'<Tag Name="([^"]+)"', block)
            if m and m.group(1) not in seen_tag_names:
                seen_tag_names.add(m.group(1))
                all_tags.append(block)

    def _add_tag_block(block: str) -> None:
        if not block:
            return
        m = re.search(r'<Tag Name="([^"]+)"', block)
        if not m or m.group(1) in seen_tag_names:
            return
        seen_tag_names.add(m.group(1))
        all_tags.append(block)

    def _bool_tag(name: str, value: int = 0) -> str:
        return (
            f'<Tag Name="{_xml_escape(name)}" TagType="Base" DataType="BOOL" '
            f'Radix="Decimal" Constant="false" ExternalAccess="Read/Write">'
            f'<Data Format="L5K"><![CDATA[{value}]]></Data>'
            f'<Data Format="Decorated"><DataValue DataType="BOOL" Value="{value}"/></Data></Tag>'
        )

    def _ensure_library_tag(name: str, fallback_bool: bool = False) -> None:
        """Clone a library controller tag by name, or create a BOOL stub."""
        if name in seen_tag_names:
            return
        block = extract_tag_block(library_text, name)
        if block:
            _add_tag_block(block)
            return
        if fallback_bool:
            _add_tag_block(_bool_tag(name, 0))

    # Simple BOOL tags for areas / zones
    for area in sorted({i["area"] for i in cloned} | set(map(_safe, inp.areas))):
        if not area:
            continue
        for suffix, dtype in (("", "Area_UDT"), ("_Safe", "ES_Zone_UDT"), ("_HMI", "Area_HMI")):
            name = f"{area}{suffix}" if suffix else area
            if name in seen_tag_names:
                continue
            # only create simple BOOL if UDT not in library sample
            if dtype == "Area_UDT" and extract_tag_block(library_text, "Main_Area"):
                block = extract_tag_block(library_text, "Main_Area")
                if block:
                    all_tags.append(block.replace("Main_Area", area))
                    seen_tag_names.add(name)
                    continue
            if suffix == "_Safe" and extract_tag_block(library_text, "Main_Area_Safe"):
                block = extract_tag_block(library_text, "Main_Area_Safe")
                if block:
                    all_tags.append(block.replace("Main_Area_Safe", name).replace("Main_Area", area))
                    seen_tag_names.add(name)
                    continue
            all_tags.append(_bool_tag(name, 0))
            seen_tag_names.add(name)

    # --- Shared tags required by Fast_Conv / Slow_Jam / Slow_Flt (Excel gold wiring) ---
    _add_tag_block(_bool_tag("AlwaysOff", 0))
    _add_tag_block(_bool_tag("AlwaysOn", 1))
    # Null objects — critical for unused PE/Conv slots (edited gold uses these heavily)
    for lib_tag in (
        "NO_PE",
        "NO_Conv",
        "NO_VFD",
        "NO_Enc",
        "NO_MS",
        "NO_AirPress",
        "NO_AdditionalFlt",
        "PConv_VFD",
        "PEExit_P",
        "PEAdd_PE",
        "HMIColor",
        "HMI_StatsClear",
        "Type2",
    ):
        _ensure_library_tag(lib_tag, fallback_bool=(lib_tag in ("HMI_StatsClear",)))
    # NO_PS (PS_UDT) — used by Slow_Flt; not always in library, synthesize if missing
    if "NO_PS" not in seen_tag_names:
        ps_block = extract_tag_block(library_text, "NO_PS")
        if ps_block:
            _add_tag_block(ps_block)
        else:
            _add_tag_block(
                '<Tag Name="NO_PS" TagType="Base" DataType="PS_UDT" Constant="false" '
                'ExternalAccess="Read/Write">'
                '<Data Format="L5K"><![CDATA[[[0],[[0,0,0],0],0]]]></Data>'
                '<Data Format="Decorated"><Structure DataType="PS_UDT"/></Data></Tag>'
            )

    # Safety zones referenced by Fast_Conv after rename (e.g. Zone1_ESZone1)
    for item in cloned:
        sz = item.get("safety_zone") or ""
        if not sz or sz in seen_tag_names:
            continue
        if extract_tag_block(library_text, "Main_Area_Safe"):
            _add_tag_block(
                extract_tag_block(library_text, "Main_Area_Safe").replace("Main_Area_Safe", sz)
            )
        elif extract_tag_block(library_text, "Conv_Area_Safe"):
            _add_tag_block(
                extract_tag_block(library_text, "Conv_Area_Safe").replace("Conv_Area_Safe", sz)
            )
        else:
            _add_tag_block(_bool_tag(sz, 0))

    # --- Ensure every RUN photoeye has PE_UDT (+ AOI) even if not on a conveyor ---
    pe_udt_src = (
        extract_tag_block(library_text, "PE1000_P")
        or extract_tag_block(library_text, "PEExit_P")
        or extract_tag_block(library_text, "NO_PE")
    )
    pe_logic_src = extract_tag_block(library_text, "PE1000_P_AOI")
    full_pe_src = extract_tag_block(library_text, "PE1000_F") or pe_udt_src
    full_aoi_src = extract_tag_block(library_text, "PE1000_F_AOI")
    for pe in getattr(inp, "pe_devices", None) or []:
        pe_name = _safe(pe.get("name") or pe.get("fortna_name") or "")
        if not pe_name or pe_name in seen_tag_names:
            continue
        role = pe.get("role") or _classify_pe_role(pe_name, pe.get("description") or "")
        if role == "full" and full_pe_src:
            _add_tag_block(
                full_pe_src.replace("PE1000_F", pe_name)
                .replace("PEExit_P", pe_name)
                .replace("NO_PE", pe_name)
            )
            if full_aoi_src and f"{pe_name}_AOI" not in seen_tag_names:
                _add_tag_block(
                    full_aoi_src.replace("PE1000_F_AOI", f"{pe_name}_AOI").replace(
                        "PE1000_F", pe_name
                    )
                )
        elif pe_udt_src:
            _add_tag_block(
                pe_udt_src.replace("PE1000_P", pe_name)
                .replace("PEExit_P", pe_name)
                .replace("NO_PE", pe_name)
            )
            if pe_logic_src and f"{pe_name}_AOI" not in seen_tag_names:
                _add_tag_block(
                    pe_logic_src.replace("PE1000_P_AOI", f"{pe_name}_AOI").replace(
                        "PE1000_P", pe_name
                    )
                )

    # --- Non-PE I/O as BOOL tags (PEs are PE_UDT already cloned above) ---
    pe_name_set = {
        _safe(p.device_name)
        for p in (inp.io_points or [])
        if (p.device_type or "") == "photoeye"
    }
    for pe in getattr(inp, "pe_devices", None) or []:
        if pe.get("name"):
            pe_name_set.add(_safe(pe["name"]))

    io_tag_rows: list[dict] = []
    for p in inp.io_points or []:
        raw = (p.device_name or "").strip()
        if not raw:
            continue
        tname = _safe(raw)[:40]
        if not tname or tname in seen_tag_names:
            # still record PE rows for CSV / IO_MAP even if tag already exists as PE_UDT
            if tname in pe_name_set or (p.device_type or "") == "photoeye":
                io_tag_rows.append({
                    "tag": tname,
                    "fortna_name": raw,
                    "fortna_address": f"Bank{p.fortna_bank}.{p.fortna_bit}" if p.fortna_bank else (p.source_module or ""),
                    "description": (p.description or f"PE {raw}")[:120],
                    "type": "PE_UDT",
                    "device_class": "photoeye",
                })
            continue
        if tname in pe_name_set or (p.device_type or "") == "photoeye":
            # PE_UDT created in clone_template; skip BOOL
            io_tag_rows.append({
                "tag": tname,
                "fortna_name": raw,
                "fortna_address": f"Bank{p.fortna_bank}.{p.fortna_bit}" if p.fortna_bank else "",
                "description": (p.description or raw)[:120],
                "type": "PE_UDT",
                "device_class": "photoeye",
            })
            continue
        desc = f"Fortna {raw} ({p.device_type or 'io'}) {p.direction or ''}".strip()
        desc_c = (desc[:120]).replace("]]>", "]] >")
        dtype = "BOOL"
        dtype_u = (p.device_type or "").lower()
        # Prefer Fortna UDT stubs for device classes that IO_MAP addresses with .I./.O.
        # (BOOL stubs cause Studio 'Invalid member specifier' on ES500.I.ES_OK etc.)
        udt_block = None
        # ALL project VFD discrete points (VFD500_AUX, VFD500_EN, VFD816A_AUX, …)
        # → one Motor_Starter_UDT P###_VFD per drive number (gold style). No BOOL stubs.
        vfd_m = (
            re.match(r"^VFD(\d+[A-Z]?)(?:_.*)?$", tname, re.I)
            or re.match(r"^VFD(\d+[A-Z]?)(?:_.*)?$", raw, re.I)
            or re.match(r"^T_VFD(\d+[A-Z]?)(?:_.*)?$", tname, re.I)
        )
        if not vfd_m and dtype_u in ("vfd", "drive", "powerflex") and re.search(
            r"VFD(\d+[A-Z]?)", raw, re.I
        ):
            vfd_m = re.search(r"VFD(\d+[A-Z]?)", raw, re.I)
        if vfd_m:
            ms_name = f"P{vfd_m.group(1)}_VFD"
            if ms_name not in seen_tag_names:
                ms_src = extract_tag_block(library_text, "NO_MS")
                if ms_src:
                    _add_tag_block(ms_src.replace("NO_MS", ms_name))
                else:
                    _add_tag_block(
                        f'<Tag Name="{_xml_escape(ms_name)}" TagType="Base" '
                        f'DataType="Motor_Starter_UDT" Constant="false" '
                        f'ExternalAccess="Read/Write">'
                        f'<Data Format="Decorated">'
                        f'<Structure DataType="Motor_Starter_UDT"/></Data></Tag>'
                    )
            # Skip BOOL VFD500_AUX — IO_MAP addresses P###_VFD.I/O members
            io_tag_rows.append({
                "tag": ms_name,
                "fortna_name": raw,
                "fortna_address": (
                    f"Bank{p.fortna_bank}.{p.fortna_bit}" if p.fortna_bank else ""
                ),
                "description": f"VFD discrete → {ms_name} (Motor_Starter_UDT)",
                "type": "Motor_Starter_UDT",
                "device_class": "vfd_ms",
            })
            continue
        # ES_UDT for e-stops AND for MCR/ESR aux contacts that IO_MAP addresses as .I.ES_OK.
        # Names may be T_1MCR1_AUX (digit-leading Fortna tags get T_ prefix).
        needs_es_udt = (
            dtype_u in ("estop", "e-stop", "e_stop", "es")
            or re.match(r"^ES\d", tname, re.I)
            or re.match(r"^ESLS", tname, re.I)
            or re.match(r"^T_\d*ES\d", tname, re.I)  # T_1ES, T_4ES…
            or re.search(r"(?:^|_)(?:\d*)?(?:MCR|ESR)\d", tname, re.I)  # T_1MCR1_AUX, T_4ESR3_AUX
            or re.search(r"(?:^|_)(?:\d*)?(?:MCR|ESR)\d", raw, re.I)
        )
        if needs_es_udt:
            src = (
                extract_tag_block(library_text, "NO_ES")
                or extract_tag_block(library_text, "CP5_ES")
            )
            if src:
                # Rename template tag to this device
                udt_block = re.sub(
                    r'Tag Name="[^"]+"',
                    f'Tag Name="{_xml_escape(tname)}"',
                    src,
                    count=1,
                )
                dtype = "ES_UDT"
        if udt_block:
            _add_tag_block(udt_block)
        else:
            _add_tag_block(
                f'<Tag Name="{_xml_escape(tname)}" TagType="Base" DataType="BOOL" '
                f'Radix="Decimal" Constant="false" ExternalAccess="Read/Write">'
                f'<Description><![CDATA[{desc_c}]]></Description>'
                f'<Data Format="L5K"><![CDATA[0]]></Data>'
                f'<Data Format="Decorated"><DataValue DataType="BOOL" Value="0"/></Data></Tag>'
            )
        bank_addr = (
            f"Bank{p.fortna_bank}.{p.fortna_bit}"
            if p.fortna_bank or p.fortna_bit
            else (p.source_module or "")
        )
        io_tag_rows.append({
            "tag": tname,
            "fortna_name": raw,
            "fortna_address": bank_addr,
            "description": desc,
            "type": dtype,
            "device_class": p.device_type or "",
        })

    def _rung_xml(num: int, text: str, comment: str = "") -> str:
        c = f"<Comment><![CDATA[{comment}]]></Comment>" if comment else ""
        return (
            f'<Rung Number="{num}" Type="N">{c}'
            f"<Text><![CDATA[{text}]]></Text></Rung>"
        )

    def routine(name: str, rungs: list[str]) -> str:
        if not rungs:
            rungs = ['<Rung Number="0" Type="N"><Text><![CDATA[NOP();]]></Text></Rung>']
        fixed = []
        for i, r in enumerate(rungs):
            fixed.append(re.sub(r'Number="\d+"', f'Number="{i}"', r, count=1))
        return (
            f'<Routine Name="{name}" Type="RLL"><RLLContent>'
            + "".join(fixed)
            + "</RLLContent></Routine>"
        )

    # Build programs per area
    programs_xml = []
    pe_wired_count = 0
    flt_count = 0
    for area, items in sorted(by_area.items()):
        rungs_fast: list[str] = []
        rungs_jam: list[str] = []
        rungs_pe: list[str] = []
        rungs_flt: list[str] = []
        for item in items:
            for r in item["rungs"]:
                rx = _rung_xml(0, r["text"], r.get("comment") or "")
                if r["label"] == "Fast":
                    rungs_fast.append(rx)
                elif r["label"] == "Jam":
                    rungs_jam.append(rx)
                elif r["label"] == "Flt":
                    rungs_flt.append(rx)
                    flt_count += 1
                elif r["label"] in ("PE", "Full"):
                    rungs_pe.append(rx)
                    if "PE_Logic" in r["text"] or "Full_PE" in r["text"]:
                        pe_wired_count += 1

        # Slow = Jam + PE + Flt only. Fast = Fast_Conv only.
        # Putting Fast_Conv in BOTH Slow and Fast caused Studio:
        #   "Duplicate AOI Backing Tag Reference …_Conv_AOI.Fast"
        main_slow = [
            _rung_xml(0, "JSR(Conv_Jam,0);", "Conv_Jam"),
            _rung_xml(1, "JSR(Conv_PE,0);", "Conv_PE"),
            _rung_xml(2, "JSR(Conv_Flt,0);", "Conv_Flt"),
        ]
        main_fast = [
            _rung_xml(0, "JSR(Conv_Fast,0);", "Conv_Fast"),
        ]

        prog_slow = f"{area}_Slow" if area.endswith("_Area") else f"{area}_Area_Slow"
        prog_fast = f"{area}_Fast" if area.endswith("_Area") else f"{area}_Area_Fast"
        prog_slow = _safe(prog_slow)[:40]
        prog_fast = _safe(prog_fast)[:40]

        programs_xml.append(
            f'<Program Name="{prog_slow}" TestEdits="false" MainRoutineName="Main_Routine" '
            f'Disabled="false" UseAsFolder="false">'
            f"<Tags/>"
            f"<Routines>"
            f'{routine("Main_Routine", main_slow)}'
            f'{routine("Conv_Jam", rungs_jam)}'
            f'{routine("Conv_PE", rungs_pe)}'
            f'{routine("Conv_Flt", rungs_flt)}'
            f"</Routines></Program>"
        )
        programs_xml.append(
            f'<Program Name="{prog_fast}" TestEdits="false" MainRoutineName="Main_Routine" '
            f'Disabled="false" UseAsFolder="false">'
            f"<Tags/>"
            f"<Routines>"
            f'{routine("Main_Routine", main_fast)}'
            f'{routine("Conv_Fast", rungs_fast)}'
            f"</Routines></Program>"
        )

    # --- IO_MAP from RUN/tar.gz (default): Conveyor.asc Bank.Bit → CPxRIOn:I/O.Data via EIP ---
    # Gold Excel IO_MAP_Program.L5X is optional (include_io_map_gold) and replaces this scaffold.
    word_map = dict(getattr(inp, "io_word_map", None) or {})

    def _word_info(word: str) -> dict | None:
        w = str(word or "").strip()
        if not w:
            return None
        info = word_map.get(w)
        if info:
            return info
        try:
            info = word_map.get(str(int(float(w))))
        except Exception:
            info = None
        if info:
            return info
        # 16-bit modules sometimes list only the low word (510); high half is 511
        if w.isdigit() and int(w) % 2 == 1:
            return word_map.get(str(int(w) - 1))
        return None

    def _vfd_ms_member(tname: str, direction: str) -> str | None:
        """Map every VFD###_* point → P###_VFD Motor_Starter_UDT members.

        Applies to ALL VFDs on the project (VFD118, VFD500, VFD816A, …), not one example.
        Gold: AUX → .I.Auxiliary_Forward; EN out → .O.Run (no bare BOOL VFD tags).
        """
        m = (
            re.match(r"^VFD(\d+[A-Z]?)(?:_(.+))?$", tname, re.I)
            or re.match(r"^T_VFD(\d+[A-Z]?)(?:_(.+))?$", tname, re.I)
        )
        if not m:
            return None
        num = m.group(1)
        suffix = (m.group(2) or "").upper()
        base = f"P{num}_VFD"
        d = (direction or "").upper()
        if suffix in ("AUX", "AUXILIARY", "AUX_FWD", "AF", "FB", "FEEDBACK"):
            return f"{base}.I.Auxiliary_Forward"
        if suffix in ("OK", "CONTACTOR", "CONTACTOR_OK", "C_OK"):
            return f"{base}.I.Contactor_OK"
        if suffix in ("MS_OK", "OL", "OVERLOAD", "OVL"):
            return f"{base}.I.MS_OK"
        if suffix in ("EN", "ENABLE", "RUN", "CMD", "START"):
            # Output enable → UDT.O.Run; input-style EN rare → Contactor_OK
            if d in ("O", "OUT", "OUTPUT"):
                return f"{base}.O.Run"
            return f"{base}.I.Contactor_OK"
        if not suffix:
            if d in ("O", "OUT", "OUTPUT"):
                return f"{base}.O.Run"
            return f"{base}.I.Auxiliary_Forward"
        return None

    def _device_member(device_type: str, tname: str, direction: str) -> str:
        """Logix tag member for OTE/XIC on the field-device side.

        Only use .I.ES_OK / .I.PE_Clear when the tag is (or will be) a UDT with
        those members. Plain BOOL tags must be referenced bare — Studio error:
        'Invalid member specifier' if you write BOOL.I.ES_OK.
        """
        dt = (device_type or "").lower()
        vfd_m = _vfd_ms_member(tname, direction)
        if vfd_m:
            return vfd_m
        if dt == "photoeye" or re.match(r"^(?:EZ)?PE\d", tname, re.I):
            return f"{tname}.I.PE_Clear"
        if (
            dt in ("estop", "e-stop", "e_stop", "es")
            or re.match(r"^ES\d", tname, re.I)
            or re.match(r"^ESLS", tname, re.I)
            or re.match(r"^T_\d*ES\d", tname, re.I)
            or re.search(r"(?:^|_)(?:\d*)?(?:MCR|ESR)\d", tname, re.I)
        ):
            return f"{tname}.I.ES_OK"
        # BOOL / simple devices (PB, motor aux, digital, beacon force bit, …)
        return tname

    def _is_output_point(p: IoPoint, info: dict | None) -> bool:
        d = (getattr(p, "direction", None) or "").upper()
        if d in ("O", "OUT", "OUTPUT"):
            return True
        if info and (info.get("direction") or "").upper() == "O":
            return True
        dt = (getattr(p, "device_type", None) or "").lower()
        if dt in ("beacon",):
            return True
        return False

    cp_i_rungs: list[str] = [
        _rung_xml(0, "NOP();", "CP_I — RUN/tar.gz inputs → device tags"),
    ]
    cp_o_rungs: list[str] = [
        _rung_xml(0, "NOP();", "CP_O — device tags → RUN/tar.gz outputs"),
    ]
    io_map_mapped = 0
    io_map_unmapped = 0
    map_points = [
        p for p in (inp.io_points or [])
        if (p.device_name or "").strip() and (p.fortna_bank or p.fortna_bit)
    ]
    map_points.sort(
        key=lambda p: (
            str(p.fortna_bank or ""),
            str(p.fortna_bit or ""),
            str(p.device_name or ""),
        )
    )
    last_rio_i = ""
    last_rio_o = ""
    for p in map_points:
        tname = _safe(p.device_name)
        if not tname:
            continue
        word = str(p.fortna_bank or "").strip()
        fbit = str(p.fortna_bit or "").strip()
        data_bit = _fortna_bit_to_data_bit(fbit)
        info = _word_info(word)
        member = _device_member(p.device_type or "", tname, p.direction or "")
        is_out = _is_output_point(p, info)
        comment = (
            f"{tname} · Bank{word}.{fbit}"
            + (f" · {info.get('type')}" if info else "")
        )
        if info and data_bit is not None and 0 <= data_bit <= 15:
            rio = info["rio_name"]
            slot = int(info["flex_slot"])
            # Prefer module direction from EIP card type; fall back to point direction
            mod_dir = (info.get("direction") or ("O" if is_out else "I")).upper()
            if mod_dir == "O" or is_out:
                # Output: field tag → module output bit
                if rio != last_rio_o:
                    cp_o_rungs.append(_rung_xml(0, "NOP();", rio))
                    last_rio_o = rio
                text = f"XIC({member})OTE({rio}:O.Data[{slot}].{data_bit});"
                cp_o_rungs.append(_rung_xml(0, text, comment))
            else:
                # Input: module input bit → field tag
                if rio != last_rio_i:
                    cp_i_rungs.append(_rung_xml(0, "NOP();", rio))
                    last_rio_i = rio
                text = f"XIC({rio}:I.Data[{slot}].{data_bit})OTE({member});"
                cp_i_rungs.append(_rung_xml(0, text, comment))
            io_map_mapped += 1
        else:
            # Keep a visible placeholder so unmapped banks are obvious in Studio
            if is_out:
                text = f"XIC({member})OTE(AlwaysOff);"
                cp_o_rungs.append(
                    _rung_xml(0, text, comment + " · UNMAPPED bank (not in EIPCSV)")
                )
            else:
                text = f"XIC(AlwaysOff)OTE({member});"
                cp_i_rungs.append(
                    _rung_xml(0, text, comment + " · UNMAPPED bank (not in EIPCSV)")
                )
            io_map_unmapped += 1

    # --- Gold program exports: Sys (default). Gold Excel IO_MAP is CLI-only. ---
    # UI "IO_MAP" checkbox → include_io_map (RUN banks). Gold Excel is separate.
    want_io_map = bool(getattr(inp, "include_io_map", True))
    want_gold_io = bool(getattr(inp, "include_io_map_gold", False))
    site_rios = {
        str((info or {}).get("rio_name") or "").strip().upper()
        for info in (getattr(inp, "io_word_map", None) or {}).values()
        if isinstance(info, dict) and (info or {}).get("rio_name")
    }
    gold_io_blocked = False
    if want_gold_io and site_rios:
        site_is_cp5_family = any(r.startswith(("CP5", "CP6", "CP7")) for r in site_rios)
        site_is_cp1_4 = any(r.startswith(("CP1", "CP2", "CP3", "CP4")) for r in site_rios)
        if site_is_cp1_4 and not site_is_cp5_family:
            want_gold_io = False
            gold_io_blocked = True
            _emit_progress(
                "Gold Excel IO_MAP blocked — this site uses CP1–CP4 (use RUN IO_MAP)",
                35,
            )
    # Gold Excel replaces RUN map when allowed; otherwise RUN map if include_io_map
    if want_gold_io:
        want_io_map = True  # gold is an IO_MAP

    gold_programs = resolve_program_exports(
        list(getattr(inp, "include_programs", None) or []),
        include_sys=bool(getattr(inp, "include_sys", True)),
        include_io_map_gold=want_gold_io,
    )
    gold_program_names: list[str] = []
    gold_io_map_used = False
    extra_aoi_chunks: list[str] = []
    extra_dt_chunks: list[str] = []

    def _tag_datatype(block: str) -> str:
        m = re.search(r'\bDataType="([^"]+)"', block)
        return (m.group(1) if m else "").strip()

    def _is_atomic_dtype(dt: str) -> bool:
        return (dt or "").upper() in {
            "BOOL", "SINT", "INT", "DINT", "LINT", "USINT", "UINT", "UDINT", "ULINT",
            "REAL", "LREAL", "STRING", "TIMER", "COUNTER", "CONTROL",
        }

    def _upsert_tag_block(block: str, *, prefer: bool = False) -> None:
        """Add tag, or replace an existing atomic stub when gold has a richer UDT.

        RUN equipment scan creates many BOOL stubs (ES500, ES610C, …). Gold IO_MAP
        exports the same names as ES_UDT / CS_UDT / DINT. If the BOOL wins, Studio
        reports 'Invalid member specifier' on OTE(ES500.I.ES_OK).
        """
        m = re.search(r'<Tag Name="([^"]+)"', block)
        if not m:
            return
        tname = m.group(1)
        new_dt = _tag_datatype(block)
        if tname not in seen_tag_names:
            seen_tag_names.add(tname)
            all_tags.append(block)
            return
        if not prefer:
            return
        # Replace atomic stub with gold structured type
        for i, existing in enumerate(all_tags):
            em = re.search(r'<Tag Name="([^"]+)"', existing)
            if not em or em.group(1) != tname:
                continue
            old_dt = _tag_datatype(existing)
            if old_dt == new_dt:
                return
            # Always prefer non-atomic (UDT) over atomic; also BOOL → DINT etc.
            if _is_atomic_dtype(old_dt) and (not _is_atomic_dtype(new_dt) or new_dt.upper() != "BOOL"):
                all_tags[i] = block
            return

    for gp in gold_programs:
        gname = gp["name"]
        # Prefer gold IO_MAP over RUN scaffold
        if gname == "IO_MAP":
            gold_io_map_used = True
        # Merge controller tags from the program export — gold wins over BOOL stubs
        for block in gp.get("tags") or []:
            _upsert_tag_block(block, prefer=True)
        if gp.get("aois_xml"):
            extra_aoi_chunks.append(gp["aois_xml"])
        if gp.get("datatypes_xml"):
            extra_dt_chunks.append(gp["datatypes_xml"])
        programs_xml.append(gp["program_xml"])
        gold_program_names.append(gname)

    # --- Live Sorter_Track: ONLY when Program pack includes Sorter_Track (checkbox) ---
    # Double safety: sorter_build UI alone does not emit the program.
    sorter_cfg = dict(getattr(inp, "sorter_build", None) or {})
    sorter_report: dict = {}
    want_sorter = any(
        (x or "").strip().lower().replace(" ", "_") in (
            "sorter_track", "sortertrack",
        )
        for x in (getattr(inp, "include_programs", None) or [])
    )
    try:
        from fortna_sorter_build import (
            build_sorter_track,
            sorter_build_is_configured,
        )
        # Automate: workbook sorter_build with real data ⇒ include pack even if
        # UI forgot the checkbox (good not perfect).
        if not want_sorter and sorter_build_is_configured(sorter_cfg):
            want_sorter = True
            inc = list(getattr(inp, "include_programs", None) or [])
            if "Sorter_Track" not in inc:
                inc.append("Sorter_Track")
                inp.include_programs = inc
            _emit_progress(
                "Sorter build data present → auto-including Sorter_Track pack",
                41,
            )
        if want_sorter:
            # Sorter_Track_Program.L5X (gold pack) configured from Sorter build UI
            live = build_sorter_track(
                sorter_cfg if sorter_build_is_configured(sorter_cfg) else {
                    "divert_count": 0,
                    "tracking": [],
                },
                library_text,
                io_points=list(inp.io_points or []),
                word_map=dict(getattr(inp, "io_word_map", None) or {}),
            )
            sorter_report = live.get("report") or {}
            sorter_report["pack_checkbox"] = True
            for block in live.get("tags") or []:
                _upsert_tag_block(block, prefer=True)
            if live.get("aoi_xml"):
                extra_aoi_chunks.append(live["aoi_xml"])
            if live.get("datatypes_xml"):
                extra_dt_chunks.append(live["datatypes_xml"])
            programs_xml.append(live["program_xml"])
            gold_program_names.append("Sorter_Track")
            mode = sorter_report.get("mode") or "configured_pack"
            _emit_progress(
                f"Sorter_Track ({mode}): "
                f"diverts kept={sorter_report.get('wave_rungs_kept', sorter_report.get('divert_count'))}/"
                f"{sorter_report.get('wave_rungs_in_pack', '?')}, "
                f"encoders={sorter_report.get('encoder_count', 0)}",
                42,
            )
        elif sorter_build_is_configured(sorter_cfg):
            sorter_report = {
                "mode": "skipped",
                "reason": "Sorter build has data but Program pack Sorter Track is off",
            }
            _emit_progress(
                "Sorter build saved but pack OFF — not emitting Sorter_Track",
                42,
            )
    except Exception as ex:
        sorter_report = {"mode": "error", "error": str(ex)}
        _emit_progress(f"Sorter_Track build failed: {ex}", 42)

    if not gold_io_map_used and want_io_map:
        # RUN/tar.gz map: CP_I (inputs) + CP_O (outputs) from Conveyor.asc + EIP word_map
        main_io_rungs = [
            _rung_xml(0, "JSR(CP_I,0);", "Inputs: CPxRIOn:I.Data → device tags (from RUN)"),
            _rung_xml(0, "JSR(CP_O,0);", "Outputs: device tags → CPxRIOn:O.Data (from RUN)"),
        ]
        programs_xml.append(
            f'<Program Name="IO_MAP" TestEdits="false" MainRoutineName="Main_Routine" '
            f'Disabled="false" UseAsFolder="false">'
            f"<Tags/>"
            f"<Routines>"
            f'{routine("Main_Routine", main_io_rungs)}'
            f'{routine("CP_I", cp_i_rungs)}'
            f'{routine("CP_O", cp_o_rungs)}'
            f"</Routines></Program>"
        )
    elif not want_io_map and not gold_io_map_used:
        _emit_progress("IO_MAP omitted (checkbox off)", 40)

    # Modules section — lightweight aliases from IO sheet (full module XML from lib is complex)
    def _extract_module_xml(lib: str, mod_name: str) -> str | None:
        """Pull one complete <Module>…</Module> (or true self-close) from library.

        IMPORTANT: opening-tag attrs must use [^>]* not [^/]*. The latter matches
        newlines and wrongly ends at the first nested '/>' (e.g. <EKey …/>),
        producing unclosed <Module> that Studio rejects (Xml_TagMismatchEx).
        """
        esc = re.escape(mod_name)
        # True self-closing Module (entire element is one tag)
        m = re.search(rf'<Module\s+Name="{esc}"(?:\s[^>]*)?/>', lib)
        if m:
            return m.group(0)
        # Full element — non-greedy to first </Module> (Modules are siblings, not nested)
        m = re.search(rf'<Module\s+Name="{esc}"(?:\s[^>]*)?>.*?</Module>', lib, re.S)
        if m:
            block = m.group(0)
            # Must start with Module and end with </Module> (not a nested self-close)
            if not re.match(r"<Module\b", block) or not re.search(r"</Module>\s*$", block):
                return None
            # Exactly one Module open (no nested Module siblings swallowed)
            if len(re.findall(r"<Module\b", block)) != 1:
                return None
            return block
        return None

    # Build a clean Modules tree (matches gold Excel Autogen shape):
    #   Local + EN2T backbone + CPxRIOn AENT/children only.
    # Do NOT dump the whole library Modules section (demo PointIO / scales / VFDs
    # cause Studio import noise and Local→Local1 rename when merging into an ACD).
    # Always open L5X as a NEW project (File→Open), not Import into an existing ACD.
    #
    # Critical: 1794-AENT OutputTag is typed AB:1794_AEN_8SLOT:O:0 — Bus Size MUST be 8.
    # Using a smaller Bus Size with the 8SLOT datatype causes Studio:
    #   "Failed to set the 'Data' property (Data type mismatch...)"
    # and then every child fails with ParentModule not found.
    eip_module_names: list[str] = []
    eip_child_names: list[str] = []
    aent_tmpl = _extract_module_xml(library_text, "IO_1N90")
    child_tmpls = {
        key: _extract_module_xml(library_text, tmpl_name)
        for key, tmpl_name in EIP_CHILD_TEMPLATE.items()
    }
    enet_parent = "CPXXENET1"  # library EN2T under Local
    topology = list(getattr(inp, "eip_topology", None) or [])
    # Only emit RIO modules used by this site's word_map (Conveyor.asc banks).
    # Multi-panel RUN packages often list CP5/CP6 adapters that reuse CP1–CP4
    # IPs — Studio then shows Invalid data type on module :I/:O tags.
    used_rios = {
        str((info or {}).get("rio_name") or "").strip()
        for info in (getattr(inp, "io_word_map", None) or {}).values()
        if isinstance(info, dict) and (info or {}).get("rio_name")
    }
    if used_rios:
        filtered = [ad for ad in topology if (ad.get("rio_name") or "") in used_rios]
        if filtered:
            topology = filtered

    # --- Local controller module (fixed Ports like gold EDITED L5X) ---
    # Library Local is often L81E at chassis slot 4 — wrong for L83E open/import.
    local_mod = _extract_module_xml(library_text, "Local") or ""
    proc = (processor or "1756-L83E").strip()
    # ProductCode for common ControlLogix Ethernet processors (Studio EKey)
    _proc_product = {
        "1756-L81E": "164",
        "1756-L82E": "165",
        "1756-L83E": "166",
        "1756-L84E": "167",
        "1756-L85E": "168",
    }
    if local_mod:
        local_mod = re.sub(
            r'CatalogNumber="[^"]*"',
            f'CatalogNumber="{_xml_escape(proc)}"',
            local_mod,
            count=1,
        )
        if proc in _proc_product:
            local_mod = re.sub(
                r'ProductCode="[^"]*"',
                f'ProductCode="{_proc_product[proc]}"',
                local_mod,
                count=1,
            )
        # Chassis backplane: controller is always Address 0 on ICP bus
        local_mod = re.sub(
            r'(<Port Id="1" Address=")[^"]*(" Type="ICP")',
            r'\g<1>0\2',
            local_mod,
            count=1,
        )
        # Room for EN2T + future local cards (gold EDITED uses 17)
        local_mod = re.sub(
            r'(<Port Id="1"[^>]*>\s*<Bus Size=")[^"]*("/>)',
            r'\g<1>17\2',
            local_mod,
            count=1,
            flags=re.S,
        )
        # Ensure Ethernet port is present (empty Bus is valid for onboard ENET)
        if not re.search(r'<Port Id="2"[^>]*Type="Ethernet"', local_mod):
            local_mod = local_mod.replace(
                "</Ports>",
                '<Port Id="2" Type="Ethernet" Upstream="false">\n<Bus/>\n</Port>\n</Ports>',
                1,
            )

    enet_mod = _extract_module_xml(library_text, enet_parent) or ""
    if enet_mod:
        # Slot 1 on local chassis (gold PLC5ENET1 style)
        enet_mod = re.sub(
            r'(<Port Id="1" Address=")[^"]*(" Type="ICP")',
            r'\g<1>1\2',
            enet_mod,
            count=1,
        )
        enet_mod = re.sub(
            r'ParentModule="[^"]*"',
            'ParentModule="Local"',
            enet_mod,
            count=1,
        )

    extra_mods: list[str] = []
    if aent_tmpl and topology:
        for ad in topology:
            rio_name = ad.get("rio_name") or ""
            ip = (ad.get("ip") or "192.168.1.50").strip()
            if not rio_name:
                continue
            # Parent 1794-AENT — AB:1794_AEN_8SLOT:I:0 / :O:0 from template
            block = aent_tmpl
            block = block.replace('Name="IO_1N90"', f'Name="{rio_name}"', 1)
            block = re.sub(
                r'ParentModule="[^"]*"',
                f'ParentModule="{enet_parent}"',
                block,
                count=1,
            )
            block = re.sub(
                r'(Address=")[\d.]+(" Type="Ethernet")',
                rf'\g<1>{_xml_escape(ip)}\2',
                block,
                count=1,
            )
            # ALWAYS Bus Size="8" — datatype is AB:1794_AEN_8SLOT (gold Excel always 8)
            kids = list(ad.get("children") or [])
            block = re.sub(
                r'(<Bus Size=")[^"]*("/>)',
                r'\g<1>8\2',
                block,
                count=1,
            )
            extra_mods.append(block)
            eip_module_names.append(rio_name)

            for c in kids:
                mt = (c.get("type") or "").strip()
                tmpl = child_tmpls.get(mt)
                if not tmpl:
                    # unknown card — skip (Studio needs a real template for AB: types)
                    continue
                cname = c.get("name") or f"{rio_name}_{c.get('flex_slot')}"
                flex = int(c.get("flex_slot") or 0)
                cblock = tmpl
                # Rename module + parent + flex port address
                # Template names: IO_1N90_1 etc.
                cblock = re.sub(
                    r'Name="IO_1N90_\d+"',
                    f'Name="{cname}"',
                    cblock,
                    count=1,
                )
                cblock = re.sub(
                    r'ParentModule="[^"]*"',
                    f'ParentModule="{rio_name}"',
                    cblock,
                    count=1,
                )
                cblock = re.sub(
                    r'(<Port Id="1" Address=")[^"]*(" Type="Flex")',
                    rf'\g<1>{flex}\2',
                    cblock,
                    count=1,
                )
                # Ensure catalog matches (template already has correct /A suffix)
                catalog = c.get("catalog") or EIP_CATALOG.get(mt, mt)
                cblock = re.sub(
                    r'CatalogNumber="[^"]*"',
                    f'CatalogNumber="{_xml_escape(catalog)}"',
                    cblock,
                    count=1,
                )
                extra_mods.append(cblock)
                eip_child_names.append(cname)

    # Assemble Modules block: Local + EN2T + RIO tree only
    mod_parts: list[str] = []
    if local_mod:
        mod_parts.append(local_mod)
    if enet_mod:
        mod_parts.append(enet_mod)
    mod_parts.extend(extra_mods)
    if mod_parts:
        modules_block = "<Modules>\n" + "\n".join(mod_parts) + "\n</Modules>"
    else:
        # Fallback: library Modules wholesale (last resort)
        lib_modules = re.search(r"<Modules>.*?</Modules>", library_text, re.S)
        modules_block = lib_modules.group(0) if lib_modules else "<Modules/>"
        if processor:
            modules_block = re.sub(
                r'(<Module Name="Local"[^>]*CatalogNumber=")[^"]*"',
                rf'\g<1>{_xml_escape(processor)}"',
                modules_block,
                count=1,
            )

    dt_xml = datatypes.group(0) if datatypes else "<DataTypes/>"
    # KEEP sealed EncodedData AOIs byte-for-byte from the O'Reilly library (Excel path).
    # Stripping them left only NOP rungs. Open as new project so signatures stay valid.
    aoi_xml = aois.group(0) if aois else "<AddOnInstructionDefinitions/>"

    # Merge extra DataTypes / AOIs from gold program exports (skip names already present)
    def _merge_named_blocks(host: str, chunks: list[str], wrapper: str, item_tag: str) -> str:
        if not chunks:
            return host
        existing = set(re.findall(rf'<{item_tag}\s+Name="([^"]+)"', host))
        extras: list[str] = []
        for ch in chunks:
            for bm in re.finditer(
                rf'<{item_tag}\s+Name="([^"]+)"[^>]*>.*?</{item_tag}>', ch, re.S
            ):
                if bm.group(1) in existing:
                    continue
                existing.add(bm.group(1))
                extras.append(bm.group(0))
            # self-closing rare
            for bm in re.finditer(rf'<{item_tag}\s+Name="([^"]+)"[^>]*/>', ch):
                if bm.group(1) in existing:
                    continue
                existing.add(bm.group(1))
                extras.append(bm.group(0))
        if not extras:
            return host
        if host.rstrip().endswith(f"</{wrapper}>"):
            return host.rstrip()[: -len(f"</{wrapper}>")] + "\n" + "\n".join(extras) + f"\n</{wrapper}>"
        return host

    dt_xml = _merge_named_blocks(dt_xml, extra_dt_chunks, "DataTypes", "DataType")
    aoi_xml = _merge_named_blocks(
        aoi_xml, extra_aoi_chunks, "AddOnInstructionDefinitions", "AddOnInstructionDefinition"
    )

    # Shorten only unsealed AOI Description text. Never rewrite EncodedData.
    aoi_xml = _shorten_aoi_descriptions(aoi_xml)

    prog_names = []
    for p in programs_xml:
        m = re.search(r'<Program\s+Name="([^"]+)"', p)
        if m:
            prog_names.append(m.group(1))

    # Own tasks — match finished gold: P02_Track = IO_MAP + Sorter_Track; Sys alone.
    # WCS / ShippingSorter L3 stay on P11 Slow until area packs are site-wired.
    own_task_programs = {"IO_MAP", "Sys", "Sorter_Track"}
    optional_slow = {
        "WCS_Interface_TCP_IP", "ShippingSorter_Area_L3",
    }
    # Also catch any gold program names that aren't Sys/IO_MAP/Sorter_Track
    for gn in gold_program_names:
        if gn not in own_task_programs:
            optional_slow.add(gn)

    slow_names = [
        n for n in prog_names
        if n not in own_task_programs
        and (n.endswith("_Slow") or n in optional_slow)
    ]
    fast_names = [n for n in prog_names if n.endswith("_Fast")]
    if not slow_names:
        slow_names = [
            n for n in prog_names
            if n not in fast_names and n not in own_task_programs
        ]
    sched_slow = "".join(f'<ScheduledProgram Name="{_xml_escape(n)}"/>' for n in slow_names)
    sched_fast = "".join(f'<ScheduledProgram Name="{_xml_escape(n)}"/>' for n in fast_names)

    def _task_xml(name: str, rate: int, priority: int, watchdog: int, programs: list[str]) -> str:
        if not programs:
            return ""
        sched = "".join(f'<ScheduledProgram Name="{_xml_escape(n)}"/>' for n in programs)
        return (
            f'<Task Name="{_xml_escape(name)}" Type="PERIODIC" Rate="{rate}" '
            f'Priority="{priority}" Watchdog="{watchdog}" '
            f'DisableUpdateOutputs="false" InhibitTask="false">\n'
            f'<ScheduledPrograms>\n{sched}\n</ScheduledPrograms>\n'
            f'</Task>\n'
        )

    # Controller SFC attrs + element order match library / Rockwell schema.
    # Open this L5X as a NEW Studio project (not Import into existing .acd).
    minor_i = int(minor) if str(minor).isdigit() else 0
    tags_block = "<Tags>\n" + "".join(all_tags) + "\n</Tags>"
    programs_block = "<Programs>\n" + "".join(programs_xml) + "\n</Programs>"
    tasks_block = '<Tasks>\n'
    # P02_Track (gold finished): IO_MAP + optional Sorter_Track on same task
    p02_progs = [n for n in ("IO_MAP", "Sorter_Track") if n in prog_names]
    if p02_progs:
        tasks_block += _task_xml(
            "P02_Track_10ms", rate=10, priority=2, watchdog=50, programs=p02_progs
        )
    elif "IO_MAP" in prog_names:
        tasks_block += _task_xml("IO_MAP", rate=20, priority=12, watchdog=100, programs=["IO_MAP"])
    if "Sys" in prog_names:
        tasks_block += _task_xml("Sys", rate=100, priority=13, watchdog=250, programs=["Sys"])
    if sched_fast:
        tasks_block += (
            '<Task Name="P10_Fast_20ms" Type="PERIODIC" Rate="20" Priority="10" '
            'Watchdog="100" DisableUpdateOutputs="false" InhibitTask="false">\n'
            f'<ScheduledPrograms>\n{sched_fast}\n</ScheduledPrograms>\n'
            '</Task>\n'
        )
    if sched_slow:
        tasks_block += (
            '<Task Name="P11_Slow_200ms" Type="PERIODIC" Rate="200" Priority="11" '
            'Watchdog="500" DisableUpdateOutputs="false" InhibitTask="false">\n'
            f'<ScheduledPrograms>\n{sched_slow}\n</ScheduledPrograms>\n'
            '</Task>\n'
        )
    tasks_block += '</Tasks>'
    # Prefer library SoftwareRevision when present (helps sealed AOI host match)
    lib_sw = re.search(r'SoftwareRevision="([^"]+)"', library_text)
    sw_rev = lib_sw.group(1) if lib_sw else f"{major}.{minor}"
    l5x = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<RSLogix5000Content SchemaRevision="1.0" SoftwareRevision="{sw_rev}" TargetName="{_xml_escape(proj)}" TargetType="Controller" ContainsContext="false" Owner="SiteForge" ExportDate="{stamp}" ExportOptions="NoRawData L5KData DecoratedData ForceProtectedEncoding AllProjDocTrans">
<Controller Use="Target" Name="{_xml_escape(proj)}" ProcessorType="{_xml_escape(processor)}" MajorRev="{major}" MinorRev="{minor_i}" ProjectCreationDate="{stamp}" LastModifiedDate="{stamp}" SFCExecutionControl="CurrentActive" SFCRestartPosition="MostRecent" SFCLastScan="DontScan" ProjectSN="16#0000_0000" MatchProjectToController="false" CanUseRPIFromProducer="false" InhibitAutomaticFirmwareUpdate="0" PassThroughConfiguration="EnabledWithAppend" DownloadProjectDocumentationAndExtendedProperties="true" DownloadProjectCustomProperties="true" ReportMinorOverflow="false">
<RedundancyInfo Enabled="false" KeepTestEditsOnSwitchOver="false"/>
<Security Code="0" ChangesToDetect="16#ffff_ffff_ffff_ffff"/>
<SafetyInfo/>
{dt_xml}
{modules_block}
{aoi_xml}
{tags_block}
{programs_block}
{tasks_block}
<CST/>
<WallClockTime/>
<Trends/>
<DataLogs/>
<TimeSynchronize/>
<EthernetPorts/>
</Controller>
</RSLogix5000Content>
'''

    # Slim report for Electron IPC (full conveyor lists live in autogen_input.json on disk)
    pe_with_real = sum(1 for c in cloned if any(p != "NO_PE" for p in (c.get("jam_pes") or [])))
    report = {
        "project": proj,
        "processor": processor,
        "revision": f"{major}.{minor}",
        "library": str(library_path),
        "conveyor_count": len(cloned),
        "area_count": len(by_area),
        "tag_count": len(all_tags),
        "program_count": len(programs_xml),
        "programs": prog_names,
        "areas_summary": {a: len(items) for a, items in by_area.items()},
        "conveyor_sample": [x["conveyor"] for x in cloned[:25]],
        "template_usage": {},
        "missing_excel_templates_in_library": sorted(missing_templates),
        "io_module_count": len(inp.modules),
        "io_point_count": len(inp.io_points),
        "io_tags_in_l5x": len(io_tag_rows),
        "pe_device_count": len(getattr(inp, "pe_devices", None) or []),
        "pe_logic_rungs": pe_wired_count,
        "slow_flt_rungs": flt_count,
        "conveyors_with_real_pe": pe_with_real,
        "eip_adapter_count": len(getattr(inp, "eip_topology", None) or []),
        "eip_rio_modules": eip_module_names,
        "eip_child_modules": eip_child_names,
        "eip_child_count": len(eip_child_names),
        "eip_interface_ip": getattr(inp, "eip_interface_ip", "") or "",
        "io_word_map_count": len(getattr(inp, "io_word_map", None) or {}),
        "io_map_rungs": max(0, len(cp_i_rungs) + len(cp_o_rungs) - 2),
        "io_map_mapped": io_map_mapped,
        "io_map_unmapped": io_map_unmapped,
        "io_map_source": (
            "gold_program_excel"
            if gold_io_map_used
            else ("run_tar_gz_banks_eip" if want_io_map else "omitted")
        ),
        "io_map_note": (
            "Gold Excel IO_MAP (library) — Greensboro CP5/CP6/CP7 only"
            if gold_io_map_used
            else (
                "Omitted — IO_MAP checkbox was off"
                if not want_io_map
                else (
                    "Built from RUN Conveyor.asc banks + EIP word_map → CPxRIOn modules"
                    + (
                        " (gold Excel blocked — site is CP1–CP4)"
                        if gold_io_blocked
                        else ""
                    )
                )
            )
        ),
        "include_io_map": want_io_map or gold_io_map_used,
        "gold_io_map_blocked": gold_io_blocked,
        "eip_modules_filtered_to_word_map": sorted(used_rios) if used_rios else [],
        "gold_programs": gold_program_names,
        "sorter_build": sorter_report,
        "optional_programs_available": list(OPTIONAL_PROGRAMS.keys()),
        "task_schedule": {
            "P02_Track_10ms": "IO_MAP + Sorter_Track (live from Sorter build UI)",
            "Sys": "Task Sys @ 100ms (own task)",
            "P10_Fast_20ms": fast_names,
            "P11_Slow_200ms": slow_names,
        },
        "encoded_aois_stripped": False,
        "logic": "Fast_Conv + Slow_Jam + PE_Logic/Full_PE + Slow_Flt + IO_MAP (module:I.Data)",
        "note": (
            "Site Forge Python autogen from RUN: real PE tags (NO_PE), PE_Logic/Full_PE, Slow_Flt, "
            "full Flex I/O tree CPxRIOn + CPxRIOn_k children with AB: data types "
            "(DI_Delay16/DO8/IB16/DO16), IO_MAP XIC(CPxRIOn:I.Data[s].b)OTE(PE.I.PE_Clear) "
            "via EIPCSV Word map + octal bits. "
            "OPEN as NEW Studio project (File→Open L5X)."
        ),
        "io_tag_rows": io_tag_rows,
    }
    for item in cloned:
        t = item["template"]
        report["template_usage"][t] = report["template_usage"].get(t, 0) + 1

    # Final pass: unsealed AOI descriptions only (never EncodedData / sealed bodies)
    l5x = _shorten_aoi_descriptions(l5x)
    return l5x, report


def _emit_progress(message: str, pct: int = 0, **extra) -> None:
    """Progress for Electron (stderr) — never pollute stdout JSON."""
    payload = {"phase": "autogen", "message": message, "pct": pct, **extra}
    try:
        print(f"FORTNA_PROGRESS {json.dumps(payload, separators=(',', ':'))}", file=sys.stderr, flush=True)
    except Exception:
        pass


def generate(
    inp: AutogenInput,
    library: Path,
    out_dir: Path | None = None,
) -> dict:
    if not library.is_file():
        raise FileNotFoundError(f"Library L5X not found: {library}")
    # Folder: timestamp + archive label (track history).
    # L5X file + Controller name: site + panel ONLY (no date) — Studio rejects long dated names.
    try:
        from fortna_source_id import (
            export_label_from_meta,
            safe_fs_name,
            studio_project_stem,
        )
        export_label = export_label_from_meta()
        folder_stem = safe_fs_name(export_label) if export_label else safe_fs_name(inp.project_name)
        # OReillyDC27_ORDENCP4 — never 20260803_0815_…
        file_stem = studio_project_stem(inp.project_name, getattr(inp, "machine", "") or "")
        if not file_stem or file_stem == "Autogen_Project":
            file_stem = studio_project_stem(inp.project_name, "")
    except Exception:
        export_label = ""
        folder_stem = _safe(inp.project_name) or "Autogen_Project"
        file_stem = _safe(inp.project_name) or "Autogen_Project"
        # strip date if present
        file_stem = re.sub(r"^\d{8}_?\d{0,6}_?", "", file_stem).strip("_") or "Autogen_Project"
    if not file_stem:
        file_stem = _safe(inp.project_name) or "Autogen_Project"
    if not folder_stem:
        folder_stem = file_stem
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    # Always unique folder per run (stamp + site) so exports don't overwrite.
    folder = f"{stamp}-{folder_stem}"
    out = out_dir or (REPO_ROOT / "exports" / "autogen" / folder)
    out.mkdir(parents=True, exist_ok=True)

    # Force short controller name inside L5X (matches file stem)
    try:
        inp.project_name = file_stem
    except Exception:
        pass

    _emit_progress(f"Building L5X for {len(inp.conveyors)} conveyors…", 20, conveyor_count=len(inp.conveyors))
    l5x, report = build_l5x(inp, library)
    l5x_path = out / f"{file_stem}.L5X"
    _emit_progress("Writing L5X file…", 70)
    l5x_path.write_text(l5x, encoding="utf-8")

    # Studio 5000 Tools → Import tag CSV (controller scope) — separate from L5X
    studio_csv_path = out / f"{file_stem}_Controller_Tags.csv"
    csv_count = 0
    try:
        from fortna_plc_export import write_studio_tags_csv

        io_rows = report.pop("io_tag_rows", []) or []
        # Also list conveyor tags for review
        for c in inp.conveyors:
            n = _safe(c.clean_name)
            if not n:
                continue
            io_rows.append({
                "tag": f"{n}_Conv",
                "fortna_name": c.conveyor,
                "fortna_address": c.main_area or "",
                "description": f"Conveyor UDT placeholder {c.type}",
                "type": "BOOL",  # CSV helper is BOOL-oriented; L5X has real Conv_UDT
                "device_class": "conveyor",
            })
        csv_count = write_studio_tags_csv(
            io_rows,
            studio_csv_path,
            controller_context=file_stem,
            software_version=f"Studio 5000 v{inp.major_rev}.{inp.minor_rev}",
        )
        report["studio_tags_csv"] = str(studio_csv_path)
        report["studio_tags_csv_count"] = csv_count
    except Exception as exc:
        report.pop("io_tag_rows", None)
        report["studio_tags_csv_error"] = str(exc)

    # Save input snapshot for audit
    snap = {
        "project_name": inp.project_name,
        "processor": inp.processor,
        "major_rev": inp.major_rev,
        "minor_rev": inp.minor_rev,
        "areas": inp.areas,
        "safety_zones": inp.safety_zones,
        "conveyors": [asdict(c) for c in inp.conveyors],
        "modules": [asdict(m) for m in inp.modules],
        "io_points": [asdict(p) for p in inp.io_points],
        "pe_devices": list(getattr(inp, "pe_devices", None) or []),
        "eip_adapters": list(getattr(inp, "eip_adapters", None) or []),
        "eip_topology": list(getattr(inp, "eip_topology", None) or []),
        "io_word_map": dict(getattr(inp, "io_word_map", None) or {}),
        "eip_interface_ip": getattr(inp, "eip_interface_ip", "") or "",
        "include_programs": list(getattr(inp, "include_programs", None) or []),
        "include_sys": bool(getattr(inp, "include_sys", True)),
        "include_io_map_gold": bool(getattr(inp, "include_io_map_gold", False)),
        "engine": "fortna_autogen.py (Python — RUN/tar.gz primary; Excel gold optional)",
    }
    (out / "autogen_input.json").write_text(json.dumps(snap, indent=2), encoding="utf-8")

    # Physical I/O verification: Fortna Word.Bit → CPxRIOn:I/O.Data[s].b
    # Sources: Conveyor.asc banks + EIPCSV/EIPModules word_map (same as eipcfg/ASC in RUN).
    try:
        word_map = dict(getattr(inp, "io_word_map", None) or {})
        topo = list(getattr(inp, "eip_topology", None) or [])

        # RIO inventory for Studio module tree check
        rio_inv = {
            "note": (
                "1794 Flex I/O from RUN PROJECT/EIPAdapters + EIPModules + EIPCSV "
                "(and FORTNA/*eipcfg.xml). One RUN = one controller's network. "
                "CP1/CP4 panels need their own RUN if they are separate PLCs."
            ),
            "interface_ip": getattr(inp, "eip_interface_ip", "") or "",
            "adapters": topo,
            "word_map_count": len(word_map),
            "rio_names": [t.get("rio_name") for t in topo],
            "module_count": sum(len(t.get("children") or []) for t in topo),
        }
        (out / "rio_inventory.json").write_text(
            json.dumps(rio_inv, indent=2), encoding="utf-8"
        )
        report["rio_inventory"] = str(out / "rio_inventory.json")
        report["rio_module_count"] = rio_inv["module_count"]
        report["rio_names"] = rio_inv["rio_names"]

        def _resolve_physical(bank: str, bit: str) -> tuple[str, str]:
            """Return (module_data_ref, note)."""
            w = str(bank or "").strip()
            b = str(bit or "").strip()
            data_bit = _fortna_bit_to_data_bit(b)
            info = word_map.get(w)
            if not info:
                try:
                    info = word_map.get(str(int(float(w))))
                except Exception:
                    info = None
            # 16-bit modules sometimes list high-half words (511 for 510) in CSV only
            if not info and w.isdigit() and int(w) % 2 == 1:
                info = word_map.get(str(int(w) - 1))
            if info and data_bit is not None and 0 <= data_bit <= 15:
                rio = info.get("rio_name") or ""
                slot = int(info.get("flex_slot") or 0)
                direction = (info.get("direction") or "I").upper()
                # Studio-style: CP5RIO0:4:O.Data.3  (also valid: :I.Data[4].3)
                ref = f"{rio}:{slot}:{direction}.Data.{data_bit}"
                alt = f"{rio}:{direction}.Data[{slot}].{data_bit}"
                return ref, f"alt={alt}; type={info.get('type')}; word={w}"
            if not info:
                return "", (
                    f"UNMAPPED word {w} — not in this RUN EIPCSV "
                    f"(other panel/PLC bank, or missing module)"
                )
            return "", f"bad bit {b} for word {w}"

        map_lines = [
            "fortna_name,device_type,direction,fortna_bank,fortna_bit,"
            "module_data_ref,mapped,notes"
        ]
        mapped_n = 0
        unmapped_n = 0
        # PE devices
        for p in getattr(inp, "pe_devices", None) or []:
            ref, note = _resolve_physical(str(p.get("bank") or ""), str(p.get("bit") or ""))
            ok = "Y" if ref else "N"
            if ref:
                mapped_n += 1
            else:
                unmapped_n += 1
            map_lines.append(
                ",".join(
                    [
                        json.dumps(p.get("fortna_name") or p.get("name") or ""),
                        json.dumps("photoeye"),
                        json.dumps("I"),
                        json.dumps(str(p.get("bank") or "")),
                        json.dumps(str(p.get("bit") or "")),
                        json.dumps(ref),
                        ok,
                        json.dumps(note),
                    ]
                )
            )
        # Other IO points from io_points (beacons, etc.)
        seen_names = {
            (p.get("fortna_name") or p.get("name") or "").upper()
            for p in (getattr(inp, "pe_devices", None) or [])
        }
        for p in getattr(inp, "io_points", None) or []:
            dname = (getattr(p, "device_name", None) or "").strip()
            if not dname or dname.upper() in seen_names:
                continue
            dtype = (getattr(p, "device_type", None) or getattr(p, "io_type", None) or "").strip()
            bank = str(getattr(p, "fortna_bank", None) or "")
            bit = str(getattr(p, "fortna_bit", None) or "")
            if not bank and not bit:
                continue
            ref, note = _resolve_physical(bank, bit)
            ok = "Y" if ref else "N"
            if ref:
                mapped_n += 1
            else:
                unmapped_n += 1
            # Infer direction from type / name
            direction = "O" if re.search(
                r"beacon|horn|lamp|light|output|OB|OA", dtype + dname, re.I
            ) else "I"
            if getattr(p, "io_type", None):
                direction = "O" if str(p.io_type).upper() in ("O", "OUT", "OUTPUT") else direction
            map_lines.append(
                ",".join(
                    [
                        json.dumps(dname),
                        json.dumps(dtype or "io"),
                        json.dumps(direction),
                        json.dumps(bank),
                        json.dumps(bit),
                        json.dumps(ref),
                        ok,
                        json.dumps(note),
                    ]
                )
            )
        (out / "physical_io_map.csv").write_text("\n".join(map_lines), encoding="utf-8")
        # Keep legacy filename for older dashboards
        (out / "io_map_pending.csv").write_text("\n".join(map_lines), encoding="utf-8")
        report["physical_io_map_csv"] = str(out / "physical_io_map.csv")
        report["physical_io_mapped"] = mapped_n
        report["physical_io_unmapped"] = unmapped_n
        report["io_map_pending_csv"] = str(out / "io_map_pending.csv")
    except Exception as exc:
        report["physical_io_map_error"] = str(exc)
        report["io_map_pending_csv_error"] = str(exc)

    # Drop bulky rows from on-disk report if still present
    report.pop("io_tag_rows", None)
    (out / "autogen_report.json").write_text(json.dumps(report, indent=2), encoding="utf-8")

    # Human report
    lines = [
        f"Site Forge PLC Autogen (Python) — {stamp}",
        f"Project: {inp.project_name}",
        f"Processor: {inp.processor}  v{inp.major_rev}.{inp.minor_rev}",
        f"Library: {library.name}",
        f"Conveyors: {report['conveyor_count']}  (with real PE: {report.get('conveyors_with_real_pe', 0)}) "
        f"— scoped to this master PLC only",
        f"Areas: {report['area_count']}",
        f"Tags: {report['tag_count']}",
        f"PE devices: {report.get('pe_device_count', 0)}  PE_Logic/Full_PE rungs: {report.get('pe_logic_rungs', 0)}",
        f"Slow_Flt rungs: {report.get('slow_flt_rungs', 0)}",
        f"EIP adapters: {report.get('eip_adapter_count', 0)}  RIO: {report.get('eip_rio_modules', [])}",
        f"EIP child cards: {report.get('eip_child_count', 0)} (AB: typed C/I/O tags)",
        f"IO_MAP source: {report.get('io_map_source')} — {report.get('io_map_note', '')}",
        f"IO_MAP: {report.get('io_map_mapped', 0)} mapped, {report.get('io_map_unmapped', 0)} unmapped "
        f"(word map entries: {report.get('io_word_map_count', 0)})",
        f"RIO modules: {report.get('rio_module_count', 0)}  "
        f"names: {', '.join(report.get('rio_names') or [])}",
        f"Physical map: {report.get('physical_io_mapped', 0)} mapped / "
        f"{report.get('physical_io_unmapped', 0)} unmapped → physical_io_map.csv",
        f"Studio tag CSV: {studio_csv_path.name if csv_count else '(none)'} ({csv_count} rows)",
        f"Programs: {', '.join(report['programs'])}",
        f"Templates: {report['template_usage']}",
        f"Output: {l5x_path}",
        "",
        report["note"],
        "Engine: fortna_autogen.py — IO_MAP from RUN/tar.gz banks + EIP modules by default.",
        "Gold Excel IO_MAP is optional (--io-map-gold / UI checkbox); it replaces RUN mapping.",
        "Logic: Fast_Conv + Slow_Jam + PE_Logic/Full_PE + Slow_Flt + IO_MAP CP_I/CP_O.",
        "1794 tree: RUN EIPAdapters/EIPModules/EIPCSV + eipcfg.xml → CPxRIOn modules in L5X.",
        "Verify: open physical_io_map.csv + rio_inventory.json next to the L5X.",
        "OPEN as NEW project in Studio (File→Open). Do not Import into existing .acd.",
    ]
    if report["missing_excel_templates_in_library"]:
        lines.append(
            "Excel templates not in library (used fallbacks): "
            + ", ".join(report["missing_excel_templates_in_library"])
        )
    (out / "autogen_report.txt").write_text("\n".join(lines), encoding="utf-8")

    # Copy library alongside for Studio import reference (best-effort; OneDrive can lock)
    try:
        shutil.copy2(library, out / library.name)
    except Exception:
        pass

    # Push export into PRISM (same site as tar.gz; skips re-index noise via upsert)
    prism_info: dict = {}
    try:
        from fortna_prism_ingest import after_export
        prism_info = after_export(export_dir=out, kind="autogen", site=file_stem)
    except Exception as exc:
        prism_info = {"ok": False, "error": str(exc)}

    result = {
        "ok": True,
        "engine": "python",
        "export_name": file_stem,
        "source_label": file_stem,
        "out_dir": str(out),
        "l5x": str(l5x_path),
        "studio_tags_csv": str(studio_csv_path) if csv_count else "",
        "report": report,
        "report_txt": str(out / "autogen_report.txt"),
        "l5x_bytes": l5x_path.stat().st_size if l5x_path.is_file() else 0,
        "prism": prism_info,
    }
    # Persist full result for Electron recovery if stdout/IPC fails
    try:
        (out / "autogen_result.json").write_text(
            json.dumps(result, separators=(",", ":")), encoding="utf-8"
        )
    except Exception:
        pass
    _emit_progress(
        f"Done — {report.get('conveyor_count', 0)} conveyors, {report.get('tag_count', 0)} tags",
        100,
    )
    return result


def main() -> int:
    ap = argparse.ArgumentParser(description="Site Forge PLC Autogen (Excel → Python)")
    sub = ap.add_subparsers(dest="cmd", required=True)

    p_ins = sub.add_parser("inspect-excel", help="Explain sheets / why Excel feels locked")
    p_ins.add_argument("excel")

    p_ex = sub.add_parser("from-excel", help="Generate L5X from autogen .xlsm/.xlsx")
    p_ex.add_argument("excel")
    p_ex.add_argument("--library", default=str(DEFAULT_LIBRARY))
    p_ex.add_argument("--out-dir", default="")

    p_js = sub.add_parser("from-json", help="Generate L5X from JSON input snapshot")
    p_js.add_argument("json_path")
    p_js.add_argument("--library", default=str(DEFAULT_LIBRARY))
    p_js.add_argument("--out-dir", default="")

    p_demo = sub.add_parser("demo", help="Run against bundled sample Excel + library")
    p_demo.add_argument("--library", default=str(DEFAULT_LIBRARY))
    p_demo.add_argument("--excel", default=str(DEFAULT_SAMPLE_XLS))
    p_demo.add_argument("--out-dir", default="")

    p_run = sub.add_parser(
        "from-run",
        help="Build autogen input from Fortna RUN (tar.gz extract) and generate L5X — no Excel typing",
    )
    p_run.add_argument("--run-dir", default=str(DEFAULT_RUN), help="Path to RUN folder (default workspace/active/RUN)")
    p_run.add_argument("--library", default=str(DEFAULT_LIBRARY))
    p_run.add_argument("--processor", default="1756-L83E")
    p_run.add_argument("--out-dir", default="")
    p_run.add_argument(
        "--preview-only",
        action="store_true",
        help="Only dump autogen_input JSON (no L5X) to inspect mapping",
    )
    p_run.add_argument(
        "--include-programs",
        default="",
        help=(
            "Comma-separated optional programs to merge: "
            "ShippingSorter_Area_L3,WCS_Interface_TCP_IP,Sorter_Track"
        ),
    )
    p_run.add_argument(
        "--no-sys",
        action="store_true",
        help="Do not merge gold Sys_Program.L5X",
    )
    p_run.add_argument(
        "--with-io-map",
        action="store_true",
        help="Include RUN/tar.gz IO_MAP program (default when --no-io-map not set)",
    )
    p_run.add_argument(
        "--no-io-map",
        action="store_true",
        help="Omit IO_MAP program from the L5X entirely",
    )
    p_run.add_argument(
        "--no-io-map-gold",
        action="store_true",
        help="Do not merge gold Excel IO_MAP_Program.L5X (default)",
    )
    p_run.add_argument(
        "--io-map-gold",
        action="store_true",
        help="OPTIONAL CLI: merge gold Excel IO_MAP_Program.L5X (Greensboro CP5–CP7)",
    )
    p_run.add_argument(
        "--workbook",
        default="",
        help="Path to Site Forge autogen_workbook.json (dashboard edits). Applied over RUN before L5X.",
    )

    args = ap.parse_args()
    try:
        def _out(obj: dict, pretty: bool = False) -> None:
            if pretty:
                print(json.dumps(obj, indent=2))
            else:
                print(json.dumps(obj, separators=(",", ":")))

        if args.cmd == "inspect-excel":
            _out(inspect_excel(Path(args.excel)), pretty=True)
            return 0
        if args.cmd == "from-excel":
            inp = load_from_excel(Path(args.excel))
            result = generate(inp, Path(args.library), Path(args.out_dir) if args.out_dir else None)
            _out(result)
            return 0 if result.get("ok") else 1
        if args.cmd == "from-json":
            inp = load_from_json(Path(args.json_path))
            result = generate(inp, Path(args.library), Path(args.out_dir) if args.out_dir else None)
            _out(result)
            return 0 if result.get("ok") else 1
        if args.cmd == "from-run":
            _emit_progress(f"Loading RUN from {args.run_dir}…", 5)
            inp = load_from_run(Path(args.run_dir), processor=args.processor)
            # Dashboard workbook (Inputdata replacement) — human edits over RUN
            wb_path = (getattr(args, "workbook", "") or "").strip()
            if wb_path:
                try:
                    from fortna_workbook import apply_workbook_to_input, load_workbook
                    wb = load_workbook(Path(wb_path))
                    if wb:
                        inp = apply_workbook_to_input(inp, wb)
                        sb = wb.get("sorter_build")
                        if isinstance(sb, dict) and sb:
                            inp.sorter_build = sb
                        _emit_progress(
                            f"Applied workbook ({len(wb.get('conveyors') or [])} rows)…",
                            12,
                        )
                except Exception as wb_exc:
                    _emit_progress(f"Workbook apply skipped: {wb_exc}", 12)
            # Optional gold programs (ShippingSorter / WCS) + live Sorter_Track
            raw_inc = getattr(args, "include_programs", "") or ""
            inp.include_programs = [
                x.strip() for x in str(raw_inc).replace(";", ",").split(",") if x.strip()
            ]
            inp.include_sys = not bool(getattr(args, "no_sys", False))
            # IO_MAP on/off (RUN banks). Default ON unless --no-io-map.
            if bool(getattr(args, "no_io_map", False)):
                inp.include_io_map = False
            else:
                inp.include_io_map = True
            # Gold Excel only with --io-map-gold (CLI advanced)
            if bool(getattr(args, "io_map_gold", False)):
                inp.include_io_map_gold = True
            else:
                inp.include_io_map_gold = False
            if args.preview_only:
                vfd_n = sum(1 for c in inp.conveyors if "vfd" in (c.type or "").lower())
                with_pe = sum(1 for c in inp.conveyors if c.all_pe_tags or c.exit_pe_tag)
                snap = {
                    "ok": True,
                    "engine": "python",
                    "project_name": inp.project_name,
                    "processor": inp.processor,
                    "areas": inp.areas,
                    "conveyor_count": len(inp.conveyors),
                    "vfd_conveyor_count": vfd_n,
                    "ms_conveyor_count": len(inp.conveyors) - vfd_n,
                    "conveyors_with_pe": with_pe,
                    "pe_device_count": len(inp.pe_devices or []),
                    "eip_adapter_count": len(inp.eip_adapters or []),
                    "eip_interface_ip": inp.eip_interface_ip or "",
                    "conveyors": [asdict(c) for c in inp.conveyors[:40]],
                    "pe_sample": (inp.pe_devices or [])[:15],
                    "io_point_count": len(inp.io_points),
                    "include_programs": inp.include_programs,
                    "include_sys": inp.include_sys,
                    "include_io_map_gold": inp.include_io_map_gold,
                    "optional_programs_available": list(OPTIONAL_PROGRAMS.keys()),
                    "run_dir": str(Path(args.run_dir).resolve()),
                    "note": (
                        "Python autogen preview — real PE tags + EIP from RUN. "
                        "Full list written on Generate → autogen_input.json"
                    ),
                }
                _out(snap)
                return 0
            result = generate(inp, Path(args.library), Path(args.out_dir) if args.out_dir else None)
            _out(result)
            return 0 if result.get("ok") else 1
        if args.cmd == "demo":
            xl = Path(args.excel)
            if not xl.is_file():
                _out({"ok": False, "error": f"Sample Excel not found: {xl}"})
                return 1
            inp = load_from_excel(xl)
            result = generate(inp, Path(args.library), Path(args.out_dir) if args.out_dir else None)
            _out(result)
            return 0 if result.get("ok") else 1
    except Exception as exc:
        print(json.dumps({"ok": False, "error": str(exc), "engine": "python"}, separators=(",", ":")))
        return 1
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
